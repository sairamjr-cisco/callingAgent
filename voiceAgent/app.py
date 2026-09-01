from flask import Flask, request, jsonify, Response, send_from_directory
from flask_cors import CORS
from twilio.rest import Client
from twilio.request_validator import RequestValidator
import os
import html
import json
import logging
import re
import random
import base64
import urllib.request
import urllib.error
import urllib.parse
import difflib
from datetime import datetime, timezone
from openpyxl import load_workbook

# Manually load environment variables from .env.local if it exists
_base_path = os.path.dirname(os.path.abspath(__file__))
_env_local_path = os.path.join(_base_path, '.env.local')
if os.path.exists(_env_local_path):
    try:
        with open(_env_local_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    if key:
                        os.environ[key] = value
    except Exception as exc:
        print(f"Failed to load .env.local: {exc}")

app = Flask(__name__)

APP_PORT = int(os.environ.get('PORT', '5501'))
APP_VERSION_BASE = (os.environ.get('APP_VERSION', 'v0') or 'v0').strip()
APP_DEPLOY_STAMP = (os.environ.get('APP_DEPLOY_STAMP', '') or '').strip()
if not APP_DEPLOY_STAMP:
    APP_DEPLOY_STAMP = datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')
APP_VERSION = f"{APP_VERSION_BASE}-{APP_DEPLOY_STAMP}"
POD_NAME = (os.environ.get('HOSTNAME', '') or '').strip()
PUBLIC_BASE_URL = os.environ.get('PUBLIC_BASE_URL', '').strip().rstrip('/')
MAX_REQUEST_BYTES = int(os.environ.get('MAX_REQUEST_BYTES', '65536'))
MAX_TWIML_TEXT_LENGTH = int(os.environ.get('MAX_TWIML_TEXT_LENGTH', '2000'))
WEBHOOK_SIGNATURE_VALIDATION = os.environ.get('TWILIO_WEBHOOK_SIGNATURE_VALIDATION', 'true').lower() in {'1', 'true', 'yes', 'on'}
DEFAULT_VOICE = os.environ.get('DEFAULT_VOICE', 'Polly.Joanna').strip() or 'Polly.Joanna'
ENABLE_CALL_RECORDING_DEFAULT = os.environ.get('ENABLE_CALL_RECORDING_DEFAULT', 'false').lower() in {'1', 'true', 'yes', 'on'}
ENABLE_CALL_TRANSCRIPT_DEFAULT = os.environ.get('ENABLE_CALL_TRANSCRIPT_DEFAULT', 'false').lower() in {'1', 'true', 'yes', 'on'}
ENABLE_AUTO_IVR_NAVIGATION_DEFAULT = os.environ.get('ENABLE_AUTO_IVR_NAVIGATION_DEFAULT', 'false').lower() in {'1', 'true', 'yes', 'on'}
TELEPHONY_PROVIDER = (os.environ.get('TELEPHONY_PROVIDER', 'twilio') or 'twilio').strip().lower()
try:
    CUSTOMER_CARE_SPEECH_TIMEOUT_SECONDS = str(max(1, min(int(os.environ.get('CUSTOMER_CARE_SPEECH_TIMEOUT_SECONDS', '3')), 10)))
except (TypeError, ValueError):
    CUSTOMER_CARE_SPEECH_TIMEOUT_SECONDS = '3'

# Replace these with your actual Twilio credentials
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '').strip()
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '').strip()
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')
CALL_STATE_DIR = os.environ.get('CALL_STATE_DIR', '/tmp/calling-agent-state')
CALL_ARTIFACTS_DIR = os.environ.get('CALL_ARTIFACTS_DIR', '/tmp/calling-agent-artifacts')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CUSTOMER_CARE_QA_FILE = os.environ.get('CUSTOMER_CARE_QA_FILE', os.path.join(BASE_DIR, 'AyjHceJGegZV3Ma9uIP2_qa.json')).strip()
CUSTOMER_DATA_FILE = os.environ.get('CUSTOMER_DATA_FILE', os.path.join(BASE_DIR, 'input', 'customer_data.json')).strip()
IVR_PROFILE_FILE = os.environ.get('IVR_PROFILE_FILE', os.path.join(BASE_DIR, 'ivr.json')).strip()
PROMPTS_DIR = os.environ.get('PROMPTS_DIR', os.path.join(BASE_DIR, 'prompts')).strip()
CONTACTS_SHEET_TEMPLATE = os.environ.get('CONTACTS_SHEET_TEMPLATE', os.path.join(BASE_DIR, 'input', 'contacts_sheet.xlsx')).strip()
PAYER_PROFILE_FILE = os.environ.get('PAYER_PROFILE_FILE', os.path.join(BASE_DIR, 'payer_profiles.json')).strip()
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '').strip()

if PUBLIC_BASE_URL:
    CORS(app, resources={r"/api/*": {"origins": [PUBLIC_BASE_URL]}})
else:
    CORS(app)

class TwilioTelephonyProvider:
    name = 'twilio'

    @staticmethod
    def create_client(account_sid, auth_token):
        return Client(account_sid, auth_token)

    @staticmethod
    def create_call(provider_client, call_params):
        return provider_client.calls.create(**call_params)

    @staticmethod
    def fetch_call(provider_client, call_sid):
        return provider_client.calls(call_sid).fetch()

    @staticmethod
    def update_call_twiml(provider_client, call_sid, twiml):
        return provider_client.calls(call_sid).update(twiml=twiml)

    @staticmethod
    def fetch_recording(provider_client, recording_sid):
        return provider_client.recordings(recording_sid).fetch()

    @staticmethod
    def delete_recording(provider_client, recording_sid):
        return provider_client.recordings(recording_sid).delete()

    @staticmethod
    def create_recording_transcription(provider_client, recording_sid, callback_url):
        return provider_client.recordings(recording_sid).transcriptions.create(
            status_callback=callback_url,
            status_callback_method='POST',
        )

    @staticmethod
    def validate_webhook_signature(auth_token, url, post_vars, signature):
        dynamic_validator = RequestValidator(auth_token)
        return dynamic_validator.validate(url, post_vars, signature)


def get_telephony_provider():
    if TELEPHONY_PROVIDER == 'twilio':
        return TwilioTelephonyProvider
    raise ValueError(f"Unsupported TELEPHONY_PROVIDER '{TELEPHONY_PROVIDER}'")


client = None
telephony_provider = get_telephony_provider()
if TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN:
    client = telephony_provider.create_client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
validator = RequestValidator(TWILIO_AUTH_TOKEN) if TWILIO_AUTH_TOKEN else None
call_twilio_credentials = {}

APP_LOG_FILE = os.environ.get('APP_LOG_FILE', '').strip()


def configure_logging():
    handlers = [logging.StreamHandler()]
    if APP_LOG_FILE:
        log_dir = os.path.dirname(APP_LOG_FILE)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)
        handlers.append(logging.FileHandler(APP_LOG_FILE, encoding='utf-8'))
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s %(levelname)s:%(name)s:%(message)s',
        handlers=handlers,
        force=True,
    )


configure_logging()
logger = logging.getLogger('calling-agent')

app.config['MAX_CONTENT_LENGTH'] = MAX_REQUEST_BYTES

os.makedirs(CALL_STATE_DIR, exist_ok=True)
os.makedirs(CALL_ARTIFACTS_DIR, exist_ok=True)


if WEBHOOK_SIGNATURE_VALIDATION and not TWILIO_AUTH_TOKEN:
    logger.warning('twilio_signature_validation_disabled_missing_auth_token')

if not client:
    logger.warning('twilio_client_not_initialized_missing_credentials')


def get_twilio_client_or_error():
    if client:
        return client, None
    return None, (jsonify({
        'error': 'Twilio credentials are missing. Set TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN.',
        'code': 'TWILIO_CREDENTIALS_MISSING',
    }), 500)


def safe_text(value):
    text = (value or '')[:MAX_TWIML_TEXT_LENGTH]
    return html.escape(text)


def get_state_file_path(call_sid):
    safe_sid = ''.join(ch for ch in (call_sid or '') if ch.isalnum() or ch in ('-', '_'))
    return os.path.join(CALL_STATE_DIR, f"{safe_sid}.json")


def read_call_state(call_sid):
    path = get_state_file_path(call_sid)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as handle:
            return json.load(handle)
    except Exception:
        return {}


def write_call_state(call_sid, state):
    import uuid
    path = get_state_file_path(call_sid)
    temp_path = f"{path}.tmp.{uuid.uuid4().hex}"
    try:
        with open(temp_path, 'w', encoding='utf-8') as handle:
            json.dump(state, handle)
        os.replace(temp_path, path)
    except Exception as e:
        logger.exception("failed_to_write_call_state call_sid=%s path=%s", call_sid, path)
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
        raise e


def update_call_state(call_sid, updates):
    state = read_call_state(call_sid)
    state.update(updates)
    write_call_state(call_sid, state)
    return state


def strip_internal_state_fields(state):
    if not isinstance(state, dict):
        return {}
    sanitized = dict(state)
    for key in ('_twilio_auth_token', '_twilio_account_sid'):
        sanitized.pop(key, None)
    return sanitized


def clear_internal_call_secrets(call_sid):
    state = read_call_state(call_sid)
    if not isinstance(state, dict) or not state:
        return
    changed = False
    for key in ('_twilio_auth_token', '_twilio_account_sid'):
        if key in state:
            state.pop(key, None)
            changed = True
    if changed:
        write_call_state(call_sid, state)


def append_interaction_transcript(call_sid, role, text, source='runtime'):
    if not call_sid or not text:
        return
    state = read_call_state(call_sid)
    entries = state.get('interaction_transcript', [])
    speaker_map = {
        'agent': 'calling_agent',
        'customer': 'representative',
        'customer_care_ivr': 'representative',
    }
    speaker = speaker_map.get(str(role or '').strip().lower(), 'system')
    entries.append({
        'ts': datetime.now().astimezone().isoformat(),
        'role': role,
        'speaker': speaker,
        'text': text,
        'source': source,
    })
    if len(entries) > 100:
        entries = entries[-100:]
    state['interaction_transcript'] = entries
    write_call_state(call_sid, state)


def normalize_interaction_transcript(entries):
    if not isinstance(entries, list):
        return []

    def classify_speaker(item):
        role = str(item.get('role') or '').strip().lower()
        speaker = str(item.get('speaker') or '').strip().lower()
        if speaker in {'calling_agent', 'representative', 'system'}:
            return speaker
        if role in {'agent'}:
            return 'calling_agent'
        if role in {'customer', 'customer_care_ivr'}:
            return 'representative'
        return 'system'

    def format_calling_agent_text(text, source):
        raw = str(text or '').strip()
        src = str(source or '').strip().lower()
        if not raw:
            return ''
        if src == 'campaign_member_switch' or src == 'claim_details_audit':
            return f"system: {raw}"
        
        # Clean up double pressed/said prefixes
        raw_lower = raw.lower()
        if raw_lower.startswith('pressed:'):
            return raw
        if raw_lower.startswith('said:'):
            return raw
            
        is_pressed = raw_lower.startswith('pressed ') or 'auto_dtmf' in src or 'ivr_navigation' in src
        if is_pressed:
            cleaned = raw[8:].strip() if raw_lower.startswith('pressed ') else raw
            return f"pressed: {cleaned}"
            
        cleaned = raw[5:].strip() if raw_lower.startswith('said ') else raw
        return f"said: {cleaned}"

    normalized = []
    for item in entries:
        if not isinstance(item, dict):
            continue
        ts = str(item.get('ts') or item.get('timestamp') or datetime.now().astimezone().isoformat()).strip()
        role = str(item.get('role') or '').strip()
        source = str(item.get('source') or 'runtime').strip()
        text = str(item.get('text') or '').strip()
        speaker = classify_speaker(item)
        if speaker == 'representative':
            normalized.append({
                'timestamp': ts,
                'representative': text,
                'calling_agent': '',
                'source': source,
            })
            continue

        if speaker == 'calling_agent':
            formatted_agent = format_calling_agent_text(text, source)
            if source == 'claim_details_audit':
                normalized.append({
                    'timestamp': ts,
                    'representative': '',
                    'calling_agent': formatted_agent,
                    'source': source,
                    'audits': item.get('audits')
                })
            elif normalized and not normalized[-1].get('calling_agent'):
                normalized[-1]['calling_agent'] = formatted_agent
            else:
                normalized.append({
                    'timestamp': ts,
                    'representative': '',
                    'calling_agent': formatted_agent,
                    'source': source,
                })
            continue

        normalized.append({
            'timestamp': ts,
            'representative': '',
            'calling_agent': f"said: {text}" if text else '',
            'source': source,
        })
    return normalized


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


def mask_secret(value):
    value = (value or '').strip()
    if not value:
        return ''
    if len(value) <= 6:
        return '*' * len(value)
    return f"{value[:3]}{'*' * (len(value) - 6)}{value[-3:]}"


def validate_twilio_credentials(account_sid, auth_token, from_number):
    if not account_sid or not auth_token or not from_number:
        return 'Missing Twilio credentials. Account SID, Auth Token, and Phone Number are required.'
    if not re.fullmatch(r'AC[a-zA-Z0-9]{32}', account_sid):
        return 'Invalid Twilio Account SID format.'
    if not re.fullmatch(r'\+[1-9]\d{7,14}', from_number):
        return 'Invalid Twilio Phone Number format. Use E.164 format like +12605973908.'
    return None


def get_twilio_context_for_call(call_sid):
    context = call_twilio_credentials.get(call_sid)
    if context:
        return telephony_provider.create_client(context['account_sid'], context['auth_token']), context
    twilio_client, error_response = get_twilio_client_or_error()
    if error_response:
        return None, None
    return twilio_client, {
        'account_sid': TWILIO_ACCOUNT_SID,
        'auth_token': TWILIO_AUTH_TOKEN,
        'from_number': TWILIO_PHONE_NUMBER,
    }


def sanitize_subfolder(value):
    value = (value or '').strip().replace('\\', '/').strip('/')
    value = re.sub(r'[^a-zA-Z0-9_./-]', '-', value)
    value = re.sub(r'/+', '/', value)
    parts = [part for part in value.split('/') if part and part != '.' and part != '..']
    if not parts:
        return 'default'
    safe = '/'.join(parts)
    return safe[:120]


def build_storage_options(payload):
    options = payload or {}
    mode = (options.get('mode') or 'twilio_only').strip()
    if mode not in {'twilio_only', 'server_keep_twilio', 'server_delete_twilio'}:
        mode = 'twilio_only'
    subfolder = sanitize_subfolder(options.get('subfolder') or 'default')
    return {
        'mode': mode,
        'subfolder': subfolder,
        'base_dir': CALL_ARTIFACTS_DIR,
        'save_enabled': mode in {'server_keep_twilio', 'server_delete_twilio'},
        'delete_from_twilio': mode == 'server_delete_twilio',
    }


def get_artifact_call_dir(call_sid, storage_options):
    call_safe = ''.join(ch for ch in (call_sid or '') if ch.isalnum() or ch in ('-', '_')) or 'unknown_call'
    subfolder = sanitize_subfolder((storage_options or {}).get('subfolder') or 'default')
    path = os.path.normpath(os.path.join(CALL_ARTIFACTS_DIR, subfolder, call_safe))
    base_norm = os.path.normpath(CALL_ARTIFACTS_DIR)
    if not path.startswith(base_norm):
        raise ValueError('Invalid artifact path')
    os.makedirs(path, exist_ok=True)
    return path


def list_artifact_subfolders(max_depth=3):
    root = os.path.normpath(CALL_ARTIFACTS_DIR)
    folders = {'default'}
    if not os.path.isdir(root):
        return ['default']

    for dirpath, dirnames, _ in os.walk(root):
        rel = os.path.relpath(dirpath, root)
        if rel == '.':
            pass
        else:
            depth = rel.count(os.sep) + 1
            if depth <= max_depth:
                safe_rel = sanitize_subfolder(rel)
                if safe_rel and safe_rel != 'default':
                    folders.add(safe_rel)
        depth_current = 0 if rel == '.' else rel.count(os.sep) + 1
        if depth_current >= max_depth:
            dirnames[:] = []
    return sorted(folders)


def save_json_file(path, data):
    tmp_path = f"{path}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as handle:
        json.dump(data, handle, ensure_ascii=True, indent=2)
    os.replace(tmp_path, path)


def save_text_file(path, text):
    tmp_path = f"{path}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as handle:
        handle.write(text or '')
    os.replace(tmp_path, path)


def is_passive_announcement_or_disclaimer(text):
    text_norm = re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9\s]', ' ', str(text or '').lower())).strip()
    if not text_norm:
        return False
    
    # Common passive IVR announcements, disclaimers, language selects, or waiting statements
    disclaimer_phrases = [
        'monitored or recorded',
        'recorded for quality',
        'quality or account security',
        'quality purposes',
        'account security purposes',
        'quality assurance',
        'calls may be recorded',
        'calls may be monitored',
        'have your npi',
        'have your tax id',
        'have your tax identifier',
        'have your member id',
        'have your member i d',
        'have this information available',
        'having this information available',
        'may also be asked for',
        'may be asked for',
        'listed on the back of',
        'refer to the phone number',
        'just remain on the line',
        'remain on the line',
        'for english just remain',
        'for english remain',
        'for english just stay',
        'de espanol',
        'will now be disconnected',
        'appreciate your call and wish you',
        'wish you a pleasant day',
    ]
    
    for phrase in disclaimer_phrases:
        if phrase in text_norm:
            return True
            
    return False


def resolve_contact_answer(question_text, contact_context):
    question = (question_text or '').strip().lower()
    if not question or not isinstance(contact_context, dict):
        return ''

    if is_passive_announcement_or_disclaimer(question_text):
        return ''

    normalized = []
    for key, value in contact_context.items():
        text_value = str(value or '').strip()
        if not text_value:
            continue
        key_text = str(key or '').strip().lower()
        if not key_text:
            continue
        if key_text in {'aetna_main_menu_choice', 'purpose_of_call', 'purpose of call', 'callback_number'}:
            continue
        normalized.append((key_text, text_value))

    direct_map = {
        'customer name': ['member_name', 'name', 'patient_name'],
        'patient name': ['member_name', 'name', 'patient_name'],
        'patient id': ['member_id', 'memberid', 'patient_id', 'patientid'],
        'claim status': ['claim_status_answer', 'claim_status', 'claimstatusanswer'],
        'claim amount': ['billing_answer', 'billed_amount', 'allowed_amount', 'paid_amount'],
        'billing': ['billing_answer', 'billing', 'billinginfo'],
        'copay amount': ['copay', 'co_pay', 'copay_amount'],
        'copay': ['copay', 'co_pay', 'copay_amount'],
        'coinsurance': ['coinsurance', 'coinsurance_amount'],
        'authorization': ['auth_answer', 'authorization_answer', 'authorization'],
        'payer id': ['payer_id', 'payerid', 'payer_id_answer'],
        'member id': ['member_id', 'memberid', 'member_id_answer'],
        'member name': ['member_name', 'name', 'patient_name'],
        'date of birth': ['member_dob', 'dob', 'date_of_birth'],
        'dob': ['member_dob', 'dob', 'date_of_birth'],
        'provider npi': ['npi', 'provider_npi', 'provider_id'],
        'npi': ['npi', 'provider_npi', 'provider_id'],
        'provider name': ['provider_name', 'billing_provider', 'servicing_provider'],
        'date of service': ['date_of_service', 'dos', 'service_date'],
        'coverage due date': ['coverage_termination_date', 'coverage_effective_date'],
        'coverage termination': ['coverage_termination_date'],
        'coverage effective': ['coverage_effective_date'],
        'call reference': ['call_reference_number', 'reference_number', 'call_reference'],
        'appeal': ['appeal_address', 'appeal_deadline'],
        'denial': ['denial_reason', 'authorization_answer', 'denial_date'],
        'trace': ['payer_id_answer', 'trace_number', 'eft_trace', 'transaction_number'],
        'eft': ['payer_id_answer', 'payment_method', 'paid_date'],
        'paid amount': ['billing_answer', 'paid_amount', 'total_paid'],
        'allowed amount': ['billing_answer', 'allowed_amount'],
    }
    for phrase, candidate_keys in direct_map.items():
        if phrase in question:
            for ck in candidate_keys:
                for key_text, value_text in normalized:
                    if ck == key_text or ck in key_text:
                        return value_text

    question_tokens = {tok for tok in re.split(r'[^a-z0-9]+', question) if len(tok) >= 2}
    best_value = ''
    best_score = 0
    for key_text, value_text in normalized:
        key_tokens = {tok for tok in re.split(r'[^a-z0-9]+', key_text) if len(tok) >= 2}
        score = len(question_tokens.intersection(key_tokens))
        if score > best_score and score >= 2:
            best_score = score
            best_value = value_text
    return best_value


def build_human_fallback_answer(asked_text, contact_context, profile_name=''):
    text_norm = _normalize_text(asked_text)
    
    greeting_tmpl = get_template_value(profile_name, 'greeting', 'Hello, this is {provider_name}. I am calling about this member claim.')
    thanks_tmpl = get_template_value(profile_name, 'thanks_goodbye', 'You are welcome. Thank you for your help today.')
    fallback_misunderstood = get_template_value(profile_name, 'fallback_misunderstood', 'I want to make sure I answer correctly. Could you rephrase that once, or ask for member, claim, provider, or date details?')

    if any(phrase in text_norm for phrase in ('thank you', 'thanks', 'that is all', "that's it", 'ok thank', 'okay thank')):
        return thanks_tmpl
    if any(phrase in text_norm for phrase in ('hello', 'hi', 'good morning', 'good afternoon', 'good evening')):
        provider_name = _ctx_first(contact_context, ['provider_name', 'billing_provider']) or 'Harbor Medical Group'
        return greeting_tmpl.format(provider_name=provider_name)

    available = []
    mapping = [
        ('patient name', ['member_name', 'patient_name', 'name']),
        ('patient id', ['member_id', 'patient_id']),
        ('date of birth', ['member_dob', 'dob', 'date_of_birth']),
        ('claim status', ['claim_status_answer', 'claim_status']),
        ('provider npi', ['npi', 'provider_npi', 'provider_id']),
        ('date of service', ['date_of_service', 'dos', 'service_date']),
    ]
    ctx = contact_context if isinstance(contact_context, dict) else {}
    for label, keys in mapping:
        for key in keys:
            value = str(ctx.get(key) or '').strip()
            if value:
                available.append(label)
                break
    if available:
        short_list = ', '.join(list(dict.fromkeys(available))[:4])
        return f'I want to make sure I give you the right detail. I can confirm {short_list}. Which one should I pull up first?'
    return fallback_misunderstood


def infer_question_topic(asked_text):
    text_norm = _normalize_text(asked_text)
    if any(token in text_norm for token in ('patient id', 'member id', 'member number', 'patient number')):
        return 'patient_id'
    if any(token in text_norm for token in ('patient name', 'member name', 'name')):
        return 'patient_name'
    if any(token in text_norm for token in ('date of birth', 'dob', 'birth')):
        return 'dob'
    if any(token in text_norm for token in ('npi', 'provider npi', 'provider id')):
        return 'npi'
    if any(token in text_norm for token in ('date of service', 'dos', 'service date')):
        return 'dos'
    if any(token in text_norm for token in ('claim status', 'status', 'approved', 'denied', 'pending')):
        return 'claim_status'
    return 'general'


def available_detail_labels(contact_context):
    ctx = contact_context if isinstance(contact_context, dict) else {}
    mapping = [
        ('patient name', ['member_name', 'patient_name', 'name']),
        ('patient id', ['member_id', 'patient_id']),
        ('date of birth', ['member_dob', 'dob', 'date_of_birth']),
        ('claim status', ['claim_status_answer', 'claim_status']),
        ('provider npi', ['npi', 'provider_npi', 'provider_id']),
        ('date of service', ['date_of_service', 'dos', 'service_date']),
    ]
    labels = []
    for label, keys in mapping:
        for key in keys:
            value = str(ctx.get(key) or '').strip()
            if value:
                labels.append(label)
                break
    return list(dict.fromkeys(labels))


def choose_non_repeating_fallback_reply(call_sid, asked_text, contact_context):
    state = read_call_state(call_sid)
    profile_name = str(state.get('selected_profile_name') or '').strip()
    fallback_count = int(state.get('fallback_reply_count') or 0)
    last_fallback_reply = str(state.get('last_fallback_reply') or '').strip()

    default_variants = [
        "I'm sorry, I didn't catch that. Could you please repeat that once?",
        'Sorry, the line was unclear. Could you please say that again?',
        'Apologies, I missed part of that. Could you repeat it slowly?',
        'I want to make sure I respond correctly. Could you repeat the question?',
        "I'm sorry, I didn't fully understand. Can you please rephrase that?",
        'Thanks, one moment please while I confirm that.',
        'Could you please repeat the requested detail one more time?',
        "Understood. I'm listening.",
        'Okay, please continue.',
    ]
    variants = get_template_value(profile_name, 'fallback_variants', default_variants)
    if not isinstance(variants, list):
        variants = default_variants

    choices = [item for item in variants if item != last_fallback_reply] or list(variants)
    selected = random.choice(choices)

    update_call_state(call_sid, {
        'fallback_reply_count': fallback_count + 1,
        'last_fallback_reply': selected,
    })
    return selected


def _normalize_text(value):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9\s]', ' ', str(value or '').lower())).strip()


def _token_set(value):
    return {tok for tok in _normalize_text(value).split(' ') if len(tok) >= 2}


def _digitize_numeric_phrases(text):
    def replace_number(match):
        number = match.group(0)
        if len(number) >= 4:
            return ' '.join(list(number))
        return number

    value = str(text or '')
    value = re.sub(r'(?<![/-])\b\d{4,}\b(?![/-])', replace_number, value)
    return value


def spell_digits_for_tts(value):
    text = str(value or '').strip()
    if not text:
        return ''
    digits_only = re.sub(r'\D+', '', text)
    if digits_only:
        return ' '.join(list(digits_only))
    return text


def _normalize_profile_id(value):
    text = str(value or '').strip().lower()
    if not text:
        return ''
    text = re.sub(r'[^a-z0-9_-]+', '_', text).strip('_')
    return text[:80]


def _normalize_customer_care_number(value):
    text = str(value or '').strip()
    if not text:
        return ''
    text = text.replace(' ', '')
    text = re.sub(r'[^0-9+]', '', text)
    if text.startswith('00'):
        text = f'+{text[2:]}'
    if text and not text.startswith('+') and len(text) >= 10:
        text = f'+{text}'
    return text[:20]


def _context_key_norm(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def get_contact_context_value(contact_context, aliases):
    if not isinstance(contact_context, dict) or not contact_context:
        return ''
    alias_set = {_context_key_norm(alias) for alias in (aliases or []) if str(alias or '').strip()}
    if not alias_set:
        return ''
    for key, value in contact_context.items():
        if _context_key_norm(key) in alias_set:
            return value
    return ''


def _parse_keyword_list(value):
    if isinstance(value, list):
        parts = value
    else:
        text = str(value or '').strip()
        if not text:
            return []
        parts = re.split(r'[,;\n]+', text)
    return [str(item or '').strip().lower() for item in parts if str(item or '').strip()][:20]


def normalize_single_ivr_profile(raw_profile, fallback_profile_name='default'):
    default = {
        'profile_name': fallback_profile_name or 'default',
        'welcome_text': '',
        'language': {
            'preferred': 'english',
            'keywords': ['english'],
        },
        'menu_goals': ['claim status', 'talk to customer care agent'],
        'representative_keywords': list(DEFAULT_IVR_REPRESENTATIVE_KEYWORDS),
        'static_route': {},
    }

    parsed = raw_profile if isinstance(raw_profile, dict) else {}
    profile = dict(default)
    profile_name = _normalize_profile_id(parsed.get('profile_name') or parsed.get('name') or fallback_profile_name) or default['profile_name']
    profile['profile_name'] = profile_name
    profile['welcome_text'] = str(parsed.get('welcome_text') or '').strip()

    language = parsed.get('language') if isinstance(parsed.get('language'), dict) else {}
    preferred = str(language.get('preferred') or parsed.get('preferred_language') or default['language']['preferred']).strip().lower() or 'english'
    keywords = language.get('keywords')
    if isinstance(keywords, str):
        keywords = [part.strip().lower() for part in keywords.split(',') if part.strip()]
    if not isinstance(keywords, list):
        keywords = []
    keywords = [str(k).strip().lower() for k in keywords if str(k).strip()][:12]
    if preferred and preferred not in keywords:
        keywords.insert(0, preferred)
    profile['language'] = {'preferred': preferred, 'keywords': keywords or ['english']}

    menu_goals = parsed.get('menu_goals')
    if isinstance(menu_goals, str):
        menu_goals = [part.strip() for part in menu_goals.split(',') if part.strip()]
    if not isinstance(menu_goals, list):
        menu_goals = []
    profile['menu_goals'] = [str(item).strip() for item in menu_goals if str(item).strip()][:20]

    rep_keywords = parsed.get('representative_keywords')
    if isinstance(rep_keywords, str):
        rep_keywords = [part.strip().lower() for part in rep_keywords.split(',') if part.strip()]
    if not isinstance(rep_keywords, list):
        rep_keywords = []
    rep_keywords = [str(item).strip().lower() for item in rep_keywords if str(item).strip()][:20]
    profile['representative_keywords'] = rep_keywords or list(DEFAULT_IVR_REPRESENTATIVE_KEYWORDS)
    profile['static_route'] = normalize_static_route(parsed.get('static_route'))
    return profile


def load_ivr_profile():
    default = {
        'profile_name': 'default',
        'profiles': {},
        'default_profile': 'aetna_representative',
        'number_map': {}
    }

    if IVR_PROFILE_FILE and os.path.exists(IVR_PROFILE_FILE):
        try:
            with open(IVR_PROFILE_FILE, 'r', encoding='utf-8') as handle:
                parsed = json.load(handle)
                if isinstance(parsed, dict) and isinstance(parsed.get('profiles'), dict):
                    for raw_profile_id, raw_profile in parsed.get('profiles', {}).items():
                        profile_id = _normalize_profile_id(raw_profile_id)
                        if profile_id:
                            default['profiles'][profile_id] = normalize_single_ivr_profile(raw_profile, fallback_profile_name=profile_id)
                    default['default_profile'] = _normalize_profile_id(parsed.get('default_profile')) or 'aetna_representative'
                    raw_number_map = parsed.get('number_map') if isinstance(parsed.get('number_map'), dict) else {}
                    for raw_number, raw_profile_id in raw_number_map.items():
                        profile_id = _normalize_profile_id(raw_profile_id)
                        number = _normalize_customer_care_number(raw_number)
                        if number and profile_id in default['profiles']:
                            default['number_map'][number] = profile_id
        except Exception as exc:
            logger.warning('failed_to_load_ivr_json_fallback path=%s error=%s', IVR_PROFILE_FILE, str(exc))

    prompts_dir = os.path.join(BASE_DIR, 'prompts')
    if os.path.exists(prompts_dir):
        try:
            for f in os.listdir(prompts_dir):
                if f.endswith('_prompts.json'):
                    payer_id = f[:-13]
                    normalized_id = _normalize_profile_id(payer_id)
                    if normalized_id and normalized_id not in default['profiles']:
                        default['profiles'][normalized_id] = {
                            'profile_name': normalized_id,
                            'welcome_text': f"Welcome to {normalized_id}.",
                            'language': {
                                'preferred': 'english',
                                'keywords': ['english']
                            },
                            'menu_goals': ['dental', 'claims', 'claim status', 'talk to customer care agent'],
                            'representative_keywords': [
                                'representative', 'agent', 'customer care', 'customer service', 'operator', 'talk to someone'
                            ],
                            'static_route': {
                                'language': {
                                    'digit': '1',
                                    'prompt_keywords': ['press 1 for english', 'for english press 1']
                                },
                                'department': {
                                    'digit': '1',
                                    'prompt_keywords': ['claims', 'claim status', 'press 1']
                                },
                                'representative': {
                                    'digit': '0',
                                    'prompt_keywords': ['speak with a representative', 'press zero']
                                }
                            }
                        }
        except Exception as exc:
            logger.warning('failed_to_scan_prompts_directory error=%s', str(exc))

    if default['profiles']:
        if default['default_profile'] not in default['profiles']:
            default['default_profile'] = next(iter(default['profiles'].keys()))
        first_key = default['default_profile']
        default['profile_name'] = default['profiles'][first_key].get('profile_name', 'default')
        default['welcome_text'] = default['profiles'][first_key].get('welcome_text', '')
        default['language'] = default['profiles'][first_key].get('language', {'preferred': 'english', 'keywords': ['english']})
        default['menu_goals'] = default['profiles'][first_key].get('menu_goals', [])
        default['representative_keywords'] = default['profiles'][first_key].get('representative_keywords', [])
        default['static_route'] = default['profiles'][first_key].get('static_route', {})

    logger.info('ivr_profiles_loaded default=%s profiles=%s', default['default_profile'], list(default['profiles'].keys()))
    return default


def load_payer_profiles():
    defaults = {
        'aetna_representative': {
            'profile_name': 'aetna_representative',
            'display_name': 'Aetna',
            'phone_number': '',
        },
        'united_health_care': {
            'profile_name': 'united_health_care',
            'display_name': 'United Health Care',
            'phone_number': '',
        },
        'icici_lombard': {
            'profile_name': 'icici_lombard',
            'display_name': 'ICICI Lombard',
            'phone_number': '',
        },
    }
    if not PAYER_PROFILE_FILE or not os.path.exists(PAYER_PROFILE_FILE):
        logger.warning('payer_profile_file_missing path=%s using_defaults=%s', PAYER_PROFILE_FILE, len(defaults))
        return defaults

    try:
        with open(PAYER_PROFILE_FILE, 'r', encoding='utf-8') as handle:
            parsed = json.load(handle)
    except Exception as exc:
        logger.warning('payer_profile_load_failed path=%s error=%s using_defaults=%s', PAYER_PROFILE_FILE, str(exc), len(defaults))
        return defaults

    rows = []
    if isinstance(parsed, dict):
        if isinstance(parsed.get('profiles'), list):
            rows = [row for row in parsed.get('profiles') if isinstance(row, dict)]
        elif isinstance(parsed.get('profiles'), dict):
            for profile_id, payload in parsed.get('profiles').items():
                if isinstance(payload, dict):
                    row = dict(payload)
                    row['profile_name'] = row.get('profile_name') or profile_id
                    rows.append(row)
        else:
            for profile_id, payload in parsed.items():
                if isinstance(payload, dict):
                    row = dict(payload)
                    row['profile_name'] = row.get('profile_name') or profile_id
                    rows.append(row)
    elif isinstance(parsed, list):
        rows = [row for row in parsed if isinstance(row, dict)]

    resolved = dict(defaults)
    for row in rows:
        profile_name = _normalize_profile_id(row.get('profile_name') or row.get('id') or row.get('name'))
        if not profile_name:
            continue
        display_name = str(row.get('display_name') or row.get('label') or profile_name.replace('_', ' ').title()).strip()
        phone_number = _normalize_customer_care_number(row.get('phone_number') or row.get('number') or row.get('customer_care_number'))
        resolved[profile_name] = {
            'profile_name': profile_name,
            'display_name': display_name,
            'phone_number': phone_number,
        }

    logger.info('payer_profile_loaded path=%s profiles=%s', PAYER_PROFILE_FILE, len(resolved))
    return resolved


def extract_claim_status_from_text(answer_text):
    text = str(answer_text or '').strip()
    if not text:
        return ''
    low = text.lower()
    if any(token in low for token in ('denied', 'partially denied')):
        return 'Denied/Partially Denied'
    if any(token in low for token in ('pending', 'in process', 'under review', 'medical review')):
        return 'Pending'
    if any(token in low for token in ('paid', 'final', 'completed', 'approved')):
        return 'Paid/Final'
    return text[:120]


def persist_claim_status_to_contacts_sheet(call_sid, member_id, claim_status, storage_options):
    safe_member_id = normalize_member_id(member_id)
    safe_status = str(claim_status or '').strip()
    if not safe_member_id or not safe_status:
        return '', 'MISSING_MEMBER_OR_STATUS'
    if not CONTACTS_SHEET_TEMPLATE or not os.path.exists(CONTACTS_SHEET_TEMPLATE):
        return '', 'CONTACTS_SHEET_TEMPLATE_MISSING'

    try:
        workbook = load_workbook(CONTACTS_SHEET_TEMPLATE)
        sheet = workbook[workbook.sheetnames[0]]
        header_row_idx = 1
        header_map = {}
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row=header_row_idx, column=col).value or '').strip().lower()
            if header:
                header_map[header] = col

        member_col = header_map.get('member_id') or header_map.get('member id')
        claim_col = header_map.get('claim_status') or header_map.get('claim status') or header_map.get('claim_status_answer')
        if not member_col:
            return '', 'CONTACTS_SHEET_MEMBER_COLUMN_MISSING'
        if not claim_col:
            claim_col = sheet.max_column + 1
            sheet.cell(row=header_row_idx, column=claim_col, value='claim_status')

        matched_row = 0
        for row in range(header_row_idx + 1, sheet.max_row + 1):
            value = normalize_member_id(sheet.cell(row=row, column=member_col).value)
            if value and value == safe_member_id:
                matched_row = row
                break
        if not matched_row:
            return '', 'CONTACTS_SHEET_MEMBER_NOT_FOUND'

        sheet.cell(row=matched_row, column=claim_col, value=safe_status)
        artifact_dir = get_artifact_call_dir(call_sid, storage_options)
        output_path = os.path.join(artifact_dir, f'contacts_sheet_updated_{safe_member_id}.xlsx')
        workbook.save(output_path)
        logger.info('contacts_sheet_claim_status_updated call_sid=%s member_id=%s path=%s', call_sid, safe_member_id, output_path)
        return output_path, ''
    except Exception as exc:
        logger.exception('contacts_sheet_claim_status_update_failed call_sid=%s member_id=%s error=%s', call_sid, safe_member_id, str(exc))
        return '', str(exc)


def maybe_update_claim_status(call_sid, state, asked_text, answer_text, active_member_id):
    question_norm = _normalize_text(asked_text)
    if not question_norm:
        return
    claim_prompts = ('claim status', 'status of claim', 'status of my claim')
    if not any(token in question_norm for token in claim_prompts):
        return

    claim_status = extract_claim_status_from_text(answer_text)
    if not claim_status:
        return

    if str(state.get('claim_status_value') or '').strip() == claim_status and str(state.get('contacts_sheet_updated_path') or '').strip():
        return

    storage_options = state.get('storage_options') or build_storage_options({})
    updated_path, update_error = persist_claim_status_to_contacts_sheet(
        call_sid,
        active_member_id,
        claim_status,
        storage_options,
    )

    updates = {'claim_status_value': claim_status}
    artifacts = state.get('artifacts', {})
    if updated_path:
        updates['contacts_sheet_updated_path'] = updated_path
        artifacts['contacts_sheet_updated_path'] = updated_path
        artifacts['last_error'] = ''
        logger.info('claim_status_updated call_sid=%s member_id=%s status=%s path=%s', call_sid, active_member_id, claim_status, updated_path)
    elif update_error:
        artifacts['last_error'] = f'CLAIM_STATUS_UPDATE_FAILED:{update_error}'
        logger.warning('claim_status_update_failed call_sid=%s member_id=%s error=%s', call_sid, active_member_id, update_error)
    updates['artifacts'] = artifacts
    update_call_state(call_sid, updates)


def infer_claim_status_from_conversation(asked_text, answer_text=''):
    asked_norm = _normalize_text(asked_text)
    answer_norm = _normalize_text(answer_text)
    status = ''
    if any(token in asked_norm or token in answer_norm for token in ('denied', 'rejected', 'rejection', 'partially denied')):
        status = 'Denied/Partially Denied'
    elif any(token in asked_norm or token in answer_norm for token in ('pending', 'under review', 'in process', 'medical review')):
        status = 'Pending'
    elif any(token in asked_norm or token in answer_norm for token in ('approved', 'paid', 'final', 'completed', 'processed')):
        status = 'Paid/Final'

    if not status:
        return ''
    return status


def build_claim_status_followup(asked_text, answer_text='', contact_context=None, state=None):
    profile_name = ""
    if state and isinstance(state, dict):
        profile_name = str(state.get('selected_profile_name') or '').strip()
    
    purpose = ""
    if isinstance(contact_context, dict):
        purpose = str(
            contact_context.get('purpose_of_call') or 
            contact_context.get('Purpose of Call') or 
            contact_context.get('Purpose Of Call') or 
            ''
        ).strip().lower()
        asked_norm = _normalize_text(asked_text)
        if 'coverage' in purpose or 'eligibility' in purpose or 'benefits' in purpose:
            if any(token in asked_norm for token in ('no longer has active coverage', 'medical coverage began', 'terminated on', 'active coverage', 'effective', 'began')):
                is_automated_ivr = ('automated' in purpose) or (profile_name == 'aetna_representative')
                default_reply = 'Thank you for the coverage and eligibility information. We appreciate your help. Have a good day.'
                reply = get_template_value(profile_name, 'coverage_verified', default_reply)
                return {
                    'claim_status': 'Coverage Verified',
                    'reply_text': reply if not is_automated_ivr else '',
                    'end_call': False if is_automated_ivr else True,
                    'await_rejection_reason': False,
                }

    claim_status = infer_claim_status_from_conversation(asked_text, answer_text)
    if not claim_status:
        return {'claim_status': '', 'reply_text': '', 'end_call': False, 'await_rejection_reason': False}

    # If we are interacting with an automated IVR, do not unilaterally trigger a spoken end_call transition.
    # Instead, we stay silent and let the IVR read out details and present the "press 3" menu naturally.
    is_automated_ivr = ('automated' in purpose) or (profile_name == 'aetna_representative')

    if claim_status == 'Paid/Final':
        default_reply = 'Thank you for the update. The claim is approved. We appreciate your help. Have a good day.'
        reply = get_template_value(profile_name, 'claim_approved', default_reply)
        return {
            'claim_status': claim_status,
            'reply_text': reply if not is_automated_ivr else '',
            'end_call': False if is_automated_ivr else True,
            'await_rejection_reason': False,
        }
    if claim_status == 'Denied/Partially Denied':
        return {
            'claim_status': claim_status,
            'reply_text': '',
            'end_call': False,
            'await_rejection_reason': False,
        }
    if claim_status == 'Pending':
        default_reply = 'Thank you for the update. Could you please share the expected ETA for approval? We will follow up accordingly. Thank you, goodbye.'
        reply = get_template_value(profile_name, 'claim_pending', default_reply)
        return {
            'claim_status': claim_status,
            'reply_text': reply if not is_automated_ivr else '',
            'end_call': False if is_automated_ivr else True,
            'await_rejection_reason': False,
        }
    return {'claim_status': '', 'reply_text': '', 'end_call': False, 'await_rejection_reason': False}


def transition_call_to_next_member_context(call_sid, state, voice, final_text):
    profile_name = str(state.get('selected_profile_name') or '').strip() if state else ''
    if profile_name == 'aetna_representative':
        logger.info("transition_call_to_next_member_context bypassed for aetna_representative to prevent double context-switching.")
        return False, build_customer_care_twiml(final_text, voice, followup=False)

    queue = normalize_campaign_contact_queue(state.get('campaign_contact_queue'))
    has_next_member = bool(queue) or bool(state.get('campaign_has_next_member'))
    if has_next_member:
        default_switch = "Now, I would like to check another claim for a different patient."
        switch_suffix = get_template_value(profile_name, 'next_member_switch', default_switch)
        switch_speech = f"{final_text} {switch_suffix}".strip()
        if queue:
            next_contact = dict(queue[0])
            remaining_queue = queue[1:]
            next_member_id = normalize_member_id(next_contact.get('member_id') or extract_member_id_from_context(next_contact))
            merged_next_context = resolve_contact_context(next_contact)
            all_bank = normalize_uploaded_answer_bank(state.get('uploaded_answer_bank_all') or state.get('uploaded_answer_bank') or [])
            next_bank = filter_uploaded_answer_bank_for_member(all_bank, next_member_id)
            update_call_state(call_sid, {
                'status': 'customer_care_listening',
                'contact_context': merged_next_context,
                'active_member_id': next_member_id,
                'campaign_member_switched': True,
                'uploaded_answer_bank': next_bank,
                'uploaded_answer_bank_all': all_bank,
                'campaign_contact_queue': remaining_queue,
                'campaign_remaining_contacts_after_current': len(remaining_queue),
                'campaign_has_next_member': bool(remaining_queue),
                'awaiting_rejection_reason': False,
                'fallback_reply_count': 0,
                'last_fallback_reply': '',
                'last_speech': '',
                'detected_intent': '',
                'next_action': '',
                'last_answer_text': '',
            })
            append_interaction_transcript(call_sid, 'agent', f'Switched context to next member {next_member_id} on same call.', source='campaign_member_switch')
        return True, build_customer_care_twiml(switch_speech, voice, followup=False)
    return False, build_end_call_twiml(final_text, voice)


def looks_like_rejection_reason_detail(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    markers = [
        'reason', 'because', 'denied due', 'rejected due', 'non covered',
        'not covered', 'documentation', 'authorization', 'eligibility',
        'timely filing', 'invalid', 'mismatch', 'missing', 'duplicate claim',
        'incorrect coding', 'medical necessity',
    ]
    return any(marker in normalized for marker in markers)


def extract_inline_audit_from_text(text, member_id, claim_status):
    """Extract the 9 audit fields from a spoken claim detail text block."""
    import re
    t = re.sub(r'\s+', ' ', str(text or ''))
    def amt(pattern):
        m = re.search(pattern, t, re.I)
        return float(m.group(1)) if m else None
    def dte(pattern):
        m = re.search(pattern, t, re.I)
        return m.group(1).strip() if m else None
    comp_date  = dte(r'completed on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})')
    allowable  = amt(r'allowable amount is\s*(?:\$)?(\d+(?:\.\d{2})?)')
    paid       = amt(r'paid to the provider[^0-9]*(?:\$)?(\d+(?:\.\d{2})?)')
    copay      = amt(r'co-payment amount[\s.]*(?:\$)?(\d+(?:\.\d{2})?)')
    coins      = amt(r'co-insurance amount[\s.]*(?:\$)?(\d+(?:\.\d{2})?)')
    issued     = dte(r'issued electronically on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})')
    settled    = dte(r'settled on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})')
    eft_m      = re.search(r'EFT[\s.]*Trace number is\s*([0-9\s,\.]+)', t, re.I)
    eft        = re.sub(r'[^0-9]', '', eft_m.group(1)).strip() if eft_m else None
    clm_m      = re.search(r'claim number is[,\s]*([A-Za-z0-9][A-Za-z0-9\s,\.]+?)(?:\s+say|\.\.\.|   |$)', t, re.I)
    claim_num  = re.sub(r'[^A-Za-z0-9]', '', clm_m.group(1)).upper() if clm_m else None
    return {
        'member_id': str(member_id or '').strip().upper(),
        'claim_status': claim_status,
        'claim_number': claim_num,
        'claim_completion_date': comp_date,
        'allowable_amount': allowable,
        'total_paid_amount': paid,
        'co_payment_amount': copay,
        'co_insurance_amount': coins,
        'payment_issued_date': issued,
        'settlement_date': settled,
        'eft_trace_number': eft,
    }


def persist_claim_status_from_conversation(call_sid, state, asked_text, answer_text, active_member_id):
    """Detect claim status from spoken text and persist to call state only.
    Audit JSON writing is handled separately at context-switch time."""
    claim_status = infer_claim_status_from_conversation(asked_text, answer_text)
    if not claim_status:
        return

    current_status = str(state.get('claim_status_value') or '').strip()
    if current_status == claim_status and str(state.get('contacts_sheet_updated_path') or '').strip():
        return

    storage_options = state.get('storage_options') or build_storage_options({})
    updated_path, update_error = persist_claim_status_to_contacts_sheet(
        call_sid,
        active_member_id,
        claim_status,
        storage_options,
    )
    updates = {'claim_status_value': claim_status}
    if updated_path:
        updates['contacts_sheet_updated_path'] = updated_path
        artifacts = state.get('artifacts', {})
        artifacts['contacts_sheet_updated_path'] = updated_path
        updates['artifacts'] = artifacts
        logger.info('claim_status_conversation_updated call_sid=%s member_id=%s status=%s path=%s', call_sid, active_member_id, claim_status, updated_path)
    elif update_error:
        logger.warning('claim_status_conversation_update_failed call_sid=%s member_id=%s error=%s', call_sid, active_member_id, update_error)
    update_call_state(call_sid, updates)


def build_customer_care_intro(contact_context):
    ctx = contact_context if isinstance(contact_context, dict) else {}
    provider_name = str(ctx.get('provider_name') or ctx.get('provider') or '').strip()
    member_name = str(ctx.get('member_name') or '').strip()
    member_id = normalize_member_id(ctx.get('member_id') or ctx.get('Member ID'))
    claim_number = str(ctx.get('call_reference_number') or ctx.get('claim_number') or '').strip()
    claim_number_spoken = spell_digits_for_tts(claim_number)
    safe_provider = provider_name or 'the provider office'
    if member_name:
        patient_ref = f'patient {member_name}'
    elif member_id:
        patient_ref = f'member id {member_id}'
    else:
        patient_ref = 'the patient'

    if claim_number_spoken:
        return f'Hi, I am calling from {safe_provider}. I am following up on claim reference {claim_number_spoken} for {patient_ref}. I need a quick status update on this claim.'
    return f'Hi, I am calling from {safe_provider}. I am following up for {patient_ref}. I need a quick status update on this claim.'


def _digits_only(value):
    return re.sub(r'\D+', '', str(value or '').strip())


def _date_to_dtmf(value):
    text = str(value or '').strip()
    if not text:
        return ''
    
    formats_to_try = [
        '%m/%d/%Y',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%m-%d-%Y',
        '%Y/%m/%d',
    ]
    
    text_clean = text.split(' ')[0] if ' ' in text else text

    for fmt in formats_to_try:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime('%m%d%Y')
        except ValueError:
            pass
        try:
            dt = datetime.strptime(text_clean, fmt)
            return dt.strftime('%m%d%Y')
        except ValueError:
            pass

    digits = _digits_only(text)
    if len(digits) == 8:
        return digits
    if len(digits) == 6:
        return f"{digits[:2]}{digits[2:4]}20{digits[4:]}"
    return ''


def _ctx_first(ctx, keys):
    if not isinstance(ctx, dict):
        return ''
    for key in keys:
        value = str(ctx.get(key) or '').strip()
        if value:
            return value
    return ''


def _context_value_by_key(contact_context, key_name):
    key_norm = _normalize_text(key_name).replace(' ', '_')
    aliases = {
        'npi': ['npi', 'provider_npi', 'tax_id'],
        'ssn': ['ssn', 'subscriber_ssn', 'subscriber_social_security_number', 'social_security_number'],
        'member_id': ['member_id', 'patient_id', 'member_id_answer', 'aetna_patient_id', 'aetna_member_id'],
        'member_dob': ['member_dob', 'dob', 'date_of_birth'],
        'date_of_service': ['date_of_service', 'dos', 'service_date'],
        'callback_number': ['callback_number', 'caller_phone', 'phone', 'provider_phone'],
        'fax_number': ['fax_number', 'provider_fax', 'fax', 'callback_number'],
        'provider_name': ['provider_name', 'provider'],
        'servicing_address': ['servicing_address', 'service_address', 'provider_address', 'address'],
        'member_name': ['member_name', 'patient_name', 'name'],
        'billed_amount': ['billed_amount', 'billing_answer', 'submitted_amount'],
        'allowed_amount': ['allowed_amount'],
        'paid_amount': ['paid_amount', 'total_paid'],
        'copay_amount': ['copay', 'co_pay', 'copay_amount'],
        'coinsurance_amount': ['coinsurance', 'coinsurance_amount'],
        'denial_reason': ['denial_reason'],
        'denial_date': ['denial_date'],
        'appeal_deadline': ['appeal_deadline'],
        'appeal_address': ['appeal_address'],
        'payment_method': ['payment_method'],
        'eft_trace': ['eft_trace', 'trace_number', 'transaction_number'],
        'paid_date': ['paid_date'],
        'call_reference_number': ['call_reference_number', 'reference_number', 'call_reference'],
        'claim_number': ['claim_number', 'call_reference_number', 'reference_number', 'call_reference'],
        'claim_number_suffix': ['claim_number_suffix', 'claim_last4', 'claim_last_four'],
        'aetna_main_menu_choice': ['aetna_main_menu_choice', 'main_menu_choice', 'aetna_menu_choice', 'claim_menu_choice'],
        'caller_name': ['caller_name', 'agent_name'],
    }
    keys = aliases.get(key_norm, [key_name])
    return _ctx_first(contact_context, keys)


def _dtmf_digits_for_field(field_name, field_value):
    field_norm = _normalize_text(field_name).replace(' ', '_')
    if field_norm in {'member_dob', 'date_of_service', 'dob', 'dos'}:
        return _date_to_dtmf(field_value)
    return _digits_only(field_value)


def infer_enter_field_from_prompt(prompt_norm):
    text = str(prompt_norm or '').strip().lower()
    if not text:
        return ''
    if 'entered' in text or 'keyed' in text or 'you entered' in text:
        return ''
    if 'enter' not in text and 'key in' not in text and 'keyed in' not in text:
        return ''
    if any(token in text for token in ('social security', 'ssn')):
        return 'ssn'
    if any(token in text for token in ('npi', 'tax id', 'tax identifier')):
        return 'npi'
    if any(token in text for token in ('date of birth', 'dob', 'birth')):
        return 'member_dob'
    if any(token in text for token in ('date of service', 'dos', 'service date')):
        return 'date_of_service'
    if 'fax' in text and 'number' in text:
        return 'fax_number'
    if any(token in text for token in ('patient id', 'member id', 'aetna id', 'patient i d', 'member i d')):
        return 'member_id'
    if 'id' in text and not any(token in text for token in ('npi', 'tax id')):
        return 'member_id'
    return ''


def load_prompt_rules(profile_name):
    safe_profile = _normalize_profile_id(profile_name)
    if not safe_profile:
        return []
    prompt_file = os.path.join(PROMPTS_DIR, f'{safe_profile}_prompts.json')
    if not os.path.exists(prompt_file):
        return []
    try:
        with open(prompt_file, 'r', encoding='utf-8') as handle:
            payload = json.load(handle)
    except Exception as exc:
        logger.warning('prompt_rules_load_failed profile=%s path=%s error=%s', safe_profile, prompt_file, str(exc))
        return []
    rules = payload.get('rules') if isinstance(payload, dict) else []
    if not isinstance(rules, list):
        return []
    logger.info('prompt_rules_loaded profile=%s path=%s rules=%s', safe_profile, prompt_file, len(rules))
    return [rule for rule in rules if isinstance(rule, dict)]


def load_payer_templates(profile_name):
    safe_profile = _normalize_profile_id(profile_name)
    if not safe_profile:
        return {}
    prompt_file = os.path.join(PROMPTS_DIR, f'{safe_profile}_prompts.json')
    if not os.path.exists(prompt_file):
        return {}
    try:
        with open(prompt_file, 'r', encoding='utf-8') as handle:
            payload = json.load(handle)
            if isinstance(payload, dict) and 'templates' in payload:
                return payload['templates']
    except Exception:
        pass
    return {}


def get_template_value(profile_name, key, default):
    templates = load_payer_templates(profile_name)
    return templates.get(key, default)


def _build_speak_from_template(template, contact_context):
    text = str(template or '').strip()
    placeholders = re.findall(r'\{([a-zA-Z0-9_ ]+)\}', text)
    for placeholder in placeholders:
        value = _context_value_by_key(contact_context, placeholder)
        if not value and placeholder.strip().lower() == 'caller_name':
            value = 'Casey Lane'
        text = text.replace(f'{{{placeholder}}}', str(value or '').strip())
    return re.sub(r'\s+', ' ', text).strip()


def _normalize_digits_in_text(text):
    word_to_digit = {
        'zero': '0', 'one': '1', 'two': '2', 'three': '3', 'four': '4',
        'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9',
    }
    res = str(text or '').lower()
    for word, digit in word_to_digit.items():
        res = re.sub(rf'\b{word}\b', digit, res)
    return res


def _rule_matches_prompt(rule, prompt_norm):
    rule_id = str(rule.get('id') or '').strip().lower()
    
    # If the rule is listen_one_moment, change prompt matching from 'contains' to exact equality '=='
    if rule_id == 'listen_one_moment':
        equals_list = ["one moment", "just a moment", "please hold", "thank you for your patience"]
        if prompt_norm not in equals_list:
            return False

    # Normalize both prompt and rule criteria to digits to make it 100% resilient to transcription differences
    prompt_digit = _normalize_digits_in_text(prompt_norm)

    contains = [str(item or '').strip().lower() for item in (rule.get('prompt_contains') or []) if str(item or '').strip()]
    if contains:
        matched = False
        for fragment in contains:
            frag_norm = fragment.lower()
            frag_digit = _normalize_digits_in_text(frag_norm)
            if '*' in frag_norm:
                regex_pattern = re.escape(frag_norm).replace(r'\*', '.*')
                regex_digit = re.escape(frag_digit).replace(r'\*', '.*')
                if re.search(regex_pattern, prompt_norm) or re.search(regex_digit, prompt_digit):
                    matched = True
                    break
            else:
                if frag_norm in prompt_norm or frag_digit in prompt_digit:
                    matched = True
                    break
        if not matched:
            return False

    requires_any = [str(item or '').strip().lower() for item in (rule.get('requires_any') or []) if str(item or '').strip()]
    if requires_any:
        matched = False
        for fragment in requires_any:
            frag_norm = fragment.lower()
            frag_digit = _normalize_digits_in_text(frag_norm)
            if '*' in frag_norm:
                regex_pattern = re.escape(frag_norm).replace(r'\*', '.*')
                regex_digit = re.escape(frag_digit).replace(r'\*', '.*')
                if re.search(regex_pattern, prompt_norm) or re.search(regex_digit, prompt_digit):
                    matched = True
                    break
            else:
                if frag_norm in prompt_norm or frag_digit in prompt_digit:
                    matched = True
                    break
        if not matched:
            return False

    return True


def resolve_prompt_action_from_rules(asked_text, contact_context, rules):
    prompt_norm = _normalize_text(asked_text)

    enter_field_hint = infer_enter_field_from_prompt(prompt_norm)
    if enter_field_hint:
        hinted_value = _context_value_by_key(contact_context, enter_field_hint)
        if enter_field_hint == 'member_id' and hinted_value and any(c.isalpha() for c in str(hinted_value)):
            spelled_id = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', str(hinted_value)))).upper()
            return {'type': 'speak', 'text': spelled_id, 'source': f'enter_intent_{enter_field_hint}_spoken_alphabetic'}

        hinted_digits = _dtmf_digits_for_field(enter_field_hint, hinted_value)
        if hinted_digits:
            return {'type': 'dtmf', 'digits': hinted_digits, 'source': f'enter_intent_{enter_field_hint}'}

    for rule in rules or []:
        if not _rule_matches_prompt(rule, prompt_norm):
            continue
        action_type = str(rule.get('type') or '').strip().lower()
        source = str(rule.get('id') or 'prompt_rule').strip()
        field_name = str(rule.get('field') or '').strip()
        field_value = _context_value_by_key(contact_context, field_name) if field_name else ''
        enter_mode = ' enter ' in f' {prompt_norm} ' or ' key in ' in f' {prompt_norm} ' or ' keyed in ' in f' {prompt_norm} '

        if enter_mode and field_name:
            enter_digits = _dtmf_digits_for_field(field_name, field_value)
            if enter_digits:
                return {'type': 'dtmf', 'digits': enter_digits, 'source': f'{source}_enter_mode'}

        if action_type == 'dtmf':
            if field_name == 'member_id' and field_value and any(c.isalpha() for c in str(field_value)):
                spelled_id = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', str(field_value)))).upper()
                return {'type': 'speak', 'text': spelled_id, 'source': f'{source}_spoken_alphabetic'}

            digits = str(rule.get('value') or '').strip()
            if not digits and field_name:
                digits = _dtmf_digits_for_field(field_name, field_value)
                if not digits and field_name in {'aetna_main_menu_choice', 'main_menu_choice', 'aetna_menu_choice', 'claim_menu_choice'}:
                    choice_norm = _normalize_text(field_value)
                    main_menu_map = {
                        'claims': '1',
                        'claim': '1',
                        'coverage and benefits': '2',
                        'coverage': '2',
                        'benefits': '2',
                        'precertification': '3',
                        'pre certification': '3',
                        'contact information': '4',
                        'contact': '4',
                        'appeal inquiries': '5',
                        'appeal': '5',
                        'join the network': '6',
                        'network': '6',
                    }
                    for label, digit in main_menu_map.items():
                        if label in choice_norm:
                            digits = digit
                            break
            if digits:
                return {'type': 'dtmf', 'digits': digits, 'source': source}

        elif action_type == 'digits_speak':
            digits = _digits_only(field_value)
            if digits:
                return {'type': 'speak', 'text': spell_digits_for_tts(digits), 'source': source}

        elif action_type == 'speak_field':
            if field_value:
                return {'type': 'speak', 'text': str(field_value).strip(), 'source': source}

        elif action_type == 'spell':
            value = str(field_value or '').strip() or 'Casey Lane'
            spelled = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', value))).upper()
            if spelled:
                return {'type': 'speak', 'text': f'It is {spelled}.', 'source': source}

        elif action_type == 'speak':
            template = str(rule.get('template') or '').strip()
            if template:
                text = _build_speak_from_template(template, contact_context)
                if text:
                    return {'type': 'speak', 'text': text, 'source': source}

        elif action_type == 'listen':
            return {'type': 'listen', 'source': source}
    return None


def build_aetna_claim_inquiry(contact_context):
    member_name = _ctx_first(contact_context, ['member_name', 'patient_name', 'name'])
    dos = _ctx_first(contact_context, ['date_of_service', 'dos', 'service_date'])
    if member_name and dos:
        return f"I am calling to check the status of a claim for member {member_name}, for date of service {dos}. Could you confirm whether the claim was received and the current status?"
    if dos:
        return f"I am calling to check the status of a claim for date of service {dos}. Could you confirm whether the claim was received and the current status?"
    return 'I am calling to check the status of a claim. Could you please help me with the current status?'


def is_active_question_or_command(asked_text):
    prompt_norm = _normalize_text(asked_text)
    if not prompt_norm:
        return False

    # Interrogatives & direct commands (must be present for it to be an active question/command)
    active_tokens = {
        'who', 'what', 'when', 'where', 'why', 'how', 'which', 'whom',
        'please', 'enter', 'say', 'press', 'provide', 'confirm', 'spell', 'share', 'verify', 'give', 'tell',
        'rephrase', 'repeat', 'restate', 'help'
    }
    active_phrases = (
        'is it', 'was it', 'can you', 'could you', 'may i', 'do you', 'does it', 'if you', 'are you', 'should i', 'would you',
        'is that', 'was that', 'is there'
    )
    
    # Check words
    tokens = set(prompt_norm.split(' '))
    if tokens.intersection(active_tokens):
        return True
    
    # Check phrases
    for phrase in active_phrases:
        if phrase in prompt_norm:
            return True
            
    return False


def resolve_universal_prompt_action(asked_text, contact_context, state):
    ivr_cfg = (state.get('ivr_navigation') or {}) if isinstance(state, dict) else {}
    profile_name = _normalize_profile_id(ivr_cfg.get('profile_name'))
    if not profile_name:
        return None

    prompt = _normalize_text(asked_text)
    if not prompt:
        return None

    prompt_rules = load_prompt_rules(profile_name)
    # Filter out main_menu_claims rule if we are already in the customer_care_qna phase
    if isinstance(state, dict) and state.get('ivr_phase') == 'customer_care_qna':
        prompt_rules = [rule for rule in prompt_rules if str(rule.get('id') or '').strip().lower() != 'main_menu_claims']

    # Extract common context fields
    member_id = _digits_only(_ctx_first(contact_context, ['member_id', 'patient_id', 'member_id_answer']))
    npi = _digits_only(_ctx_first(contact_context, ['npi', 'provider_npi', 'tax_id']))
    dob = _date_to_dtmf(_ctx_first(contact_context, ['member_dob', 'dob', 'date_of_birth']))
    dos = _date_to_dtmf(_ctx_first(contact_context, ['date_of_service', 'dos', 'service_date']))
    ssn = _digits_only(_context_value_by_key(contact_context, 'ssn'))
    purpose = str(
        contact_context.get('Purpose of Call') or 
        contact_context.get('purpose_of_call') or 
        contact_context.get('Purpose Of Call') or 
        ''
    ).strip()
    callback = _digits_only(_ctx_first(contact_context, ['callback_number', 'caller_phone', 'phone', 'provider_phone']))
    provider_name = _ctx_first(contact_context, ['provider_name', 'provider'])
    service_address = _ctx_first(contact_context, ['servicing_address', 'service_address', 'provider_address', 'address'])
    caller_name = _ctx_first(contact_context, ['caller_name', 'agent_name']) or 'Casey Lane'
    member_name = _ctx_first(contact_context, ['member_name', 'patient_name', 'name'])
    billed_amount = _ctx_first(contact_context, ['billed_amount', 'billing_answer', 'submitted_amount'])
    claim_number = _ctx_first(contact_context, ['claim_number', 'call_reference_number', 'reference_number', 'call_reference'])
    claim_number_suffix = str(_ctx_first(contact_context, ['claim_number_suffix', 'claim_last4', 'claim_last_four'])).strip()
    if not claim_number_suffix and claim_number:
        compact_claim = re.sub(r'[^A-Za-z0-9]+', '', str(claim_number))
        if len(compact_claim) >= 4:
            claim_number_suffix = compact_claim[-4:]

    queue = normalize_campaign_contact_queue(state.get('campaign_contact_queue')) if isinstance(state, dict) else []
    has_next_member = bool(queue) or bool(state.get('campaign_has_next_member')) if isinstance(state, dict) else False

    # Check matching enter field hints dynamically
    enter_field_hint = infer_enter_field_from_prompt(prompt)
    if enter_field_hint:
        hinted_value = _context_value_by_key(contact_context, enter_field_hint)
        if enter_field_hint == 'member_id' and hinted_value and any(c.isalpha() for c in str(hinted_value)):
            spelled_id = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', str(hinted_value)))).upper()
            return {'type': 'speak', 'text': spelled_id, 'source': f'enter_intent_{enter_field_hint}_spoken_alphabetic'}

        hinted_digits = _dtmf_digits_for_field(enter_field_hint, hinted_value)
        if hinted_digits:
            return {'type': 'dtmf', 'digits': hinted_digits, 'source': f'enter_intent_{enter_field_hint}'}

    for rule in prompt_rules:
        if not _rule_matches_prompt(rule, prompt):
            continue

        action_type = str(rule.get('type') or '').strip().lower()
        source = str(rule.get('id') or 'prompt_rule').strip()
        field_name = str(rule.get('field') or '').strip()
        field_value = _context_value_by_key(contact_context, field_name) if field_name else ''

        # 1. Custom Action: Fax Document Request
        if action_type == 'fax_request':
            digits = str(rule.get('value') or '2').strip()
            if purpose == "Automated Claim Status Inquiry & Document Request":
                return {'type': 'dtmf', 'digits': digits, 'source': f'{source}_fax_doc_request'}
            else:
                return {'type': 'dtmf', 'digits': digits, 'source': f'{source}_skip_fax'}

        # 2. Custom Action: SSN Keypad Input
        elif action_type == 'ssn_input':
            if ssn:
                return {'type': 'dtmf', 'digits': ssn, 'source': f'{source}_ssn'}

        # 3. Custom Action: Campaign Queue Rollover / Try Another Patient
        elif action_type == 'campaign_rollover':
            digits = str(rule.get('value') or '1').strip()
            if has_next_member:
                return {'type': 'dtmf', 'digits': digits, 'source': f'{source}_try_another_patient_press_{digits}'}
            else:
                goodbye_text = 'No more patients to check. Thank you for your help today. Goodbye.'
                return {'type': 'speak', 'text': goodbye_text, 'end_call': True, 'source': f'{source}_campaign_end'}

        # 4. Standard DTMF / digits / text resolved actions
        elif action_type == 'dtmf':
            if field_name == 'member_id' and field_value and any(c.isalpha() for c in str(field_value)):
                spelled_id = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', str(field_value)))).upper()
                return {'type': 'speak', 'text': spelled_id, 'source': f'{source}_spoken_alphabetic'}

            digits = str(rule.get('value') or '').strip()
            if not digits and field_name:
                digits = _dtmf_digits_for_field(field_name, field_value)
                if not digits and field_name in {'aetna_main_menu_choice', 'main_menu_choice', 'aetna_menu_choice', 'claim_menu_choice'}:
                    choice_norm = _normalize_text(field_value)
                    main_menu_map = {
                        'claims': '1',
                        'claim': '1',
                        'coverage and benefits': '2',
                        'coverage': '2',
                        'benefits': '2',
                        'precertification': '3',
                        'pre certification': '3',
                        'contact information': '4',
                        'contact': '4',
                        'appeal': '5',
                        'appeals': '5',
                    }
                    digits = main_menu_map.get(choice_norm, '')
            if digits:
                return {'type': 'dtmf', 'digits': digits, 'source': source}

        elif action_type == 'digits_speak':
            digits = _digits_only(field_value)
            if digits:
                return {'type': 'speak', 'text': spell_digits_for_tts(digits), 'source': source}

        elif action_type == 'speak_field':
            if field_value:
                return {'type': 'speak', 'text': str(field_value).strip(), 'source': source}

        elif action_type == 'spell':
            value = str(field_value or '').strip() or 'Casey Lane'
            spelled = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', value))).upper()
            if spelled:
                return {'type': 'speak', 'text': f'It is {spelled}.', 'source': source}

        elif action_type == 'speak':
            template = str(rule.get('template') or '').strip()
            if template:
                text = _build_speak_from_template(template, contact_context)
                if text:
                    return {'type': 'speak', 'text': text, 'source': source}

        elif action_type == 'listen':
            return {'type': 'listen', 'source': source}

    # Spoken info fallback checks (in case some keys aren't in JSON prompts yet)
    if 'may i have your name' in prompt:
        return {'type': 'speak', 'text': f'My name is {caller_name}.', 'source': 'universal_caller_name'}
    if 'spell it out' in prompt:
        spelled = ' '.join(list(re.sub(r'[^A-Za-z0-9]', '', caller_name))).upper()
        return {'type': 'speak', 'text': f'It is {spelled}.', 'source': 'universal_caller_name_spelling'}
    if 'callback number' in prompt and callback:
        return {'type': 'speak', 'text': spell_digits_for_tts(callback), 'source': 'universal_callback_number'}
    if 'direct line' in prompt:
        return {'type': 'speak', 'text': 'Yes, this is a direct line.', 'source': 'universal_direct_line'}
    if "provider's name" in prompt and provider_name:
        return {'type': 'speak', 'text': provider_name, 'source': 'universal_provider_name'}
    if 'servicing address' in prompt and service_address:
        return {'type': 'speak', 'text': service_address, 'source': 'universal_servicing_address'}
    if "member's name" in prompt and member_name:
        return {'type': 'speak', 'text': member_name, 'source': 'universal_member_name'}
    if 'how may i help you' in prompt:
        return {'type': 'speak', 'text': build_aetna_claim_inquiry(contact_context), 'source': 'universal_claim_inquiry'}
    if 'total bill amount' in prompt and billed_amount:
        return {'type': 'speak', 'text': billed_amount, 'source': 'universal_billed_amount'}
    if 'claim number' in prompt and ('last four' in prompt or 'last 4' in prompt or 'last four characters' in prompt or 'last 4 characters' in prompt) and claim_number_suffix:
        return {'type': 'speak', 'text': spell_digits_for_tts(claim_number_suffix), 'source': 'universal_claim_number_suffix'}
    if 'claim number' in prompt and claim_number:
        return {'type': 'speak', 'text': claim_number, 'source': 'universal_claim_number'}
    if 'final' in prompt and not any(token in prompt for token in ('paid', 'denied', 'partially', 'deductible')):
        return {'type': 'speak', 'text': 'When you say final, was it paid, denied, or applied to deductible?', 'source': 'universal_final_status_clarify'}

    return None


def load_customer_care_qa_pairs():
    if not CUSTOMER_CARE_QA_FILE:
        return []
    if not os.path.exists(CUSTOMER_CARE_QA_FILE):
        logger.warning('customer_care_qa_file_missing path=%s', CUSTOMER_CARE_QA_FILE)
        return []
    try:
        with open(CUSTOMER_CARE_QA_FILE, 'r', encoding='utf-8') as handle:
            raw = json.load(handle)
    except Exception as exc:
        logger.warning('customer_care_qa_file_load_failed path=%s error=%s', CUSTOMER_CARE_QA_FILE, str(exc))
        return []

    pairs = []
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            question = str(item.get('question') or '').strip()
            answer = str(item.get('answer') or '').strip()
            if question and answer:
                pairs.append({'question': question, 'answer': answer})
    logger.info('customer_care_qa_pairs_loaded path=%s count=%s', CUSTOMER_CARE_QA_FILE, len(pairs))
    return pairs


def normalize_member_id(value):
    return str(value or '').strip().upper()


def member_id_candidates(value):
    member_id = normalize_member_id(value)
    if not member_id:
        return []
    candidates = [member_id]
    if member_id.isdigit():
        compact = str(int(member_id))
        if compact != member_id:
            candidates.append(compact)
    return candidates


def extract_member_id_from_context(contact_context):
    if not isinstance(contact_context, dict):
        return ''
    for key, val in contact_context.items():
        clean_key = str(key).strip().lower().replace('_', '').replace(' ', '')
        if clean_key in {'memberid', 'patientid'}:
            parsed_val = normalize_member_id(val)
            if parsed_val:
                return parsed_val
    return ''


def load_customer_data_index():
    if not CUSTOMER_DATA_FILE:
        return {}
    if not os.path.exists(CUSTOMER_DATA_FILE):
        logger.warning('customer_data_file_missing path=%s', CUSTOMER_DATA_FILE)
        return {}

    try:
        with open(CUSTOMER_DATA_FILE, 'r', encoding='utf-8') as handle:
            raw = json.load(handle)
    except Exception as exc:
        logger.warning('customer_data_file_load_failed path=%s error=%s', CUSTOMER_DATA_FILE, str(exc))
        return {}

    rows = []
    if isinstance(raw, dict):
        if 'member_id' in raw:
            rows = [raw]
        else:
            nested_keys = ('members', 'data', 'rows', 'items')
            for key in nested_keys:
                nested = raw.get(key)
                if isinstance(nested, list):
                    rows.extend([row for row in nested if isinstance(row, dict)])
                elif isinstance(nested, dict):
                    rows.extend([row for row in nested.values() if isinstance(row, dict)])
            for maybe_row in raw.values():
                if isinstance(maybe_row, dict):
                    rows.append(maybe_row)
    elif isinstance(raw, list):
        rows = [row for row in raw if isinstance(row, dict)]

    index = {}
    duplicate_count = 0
    for row in rows:
        mid = extract_member_id_from_context(row)
        if mid:
            if mid in index:
                duplicate_count += 1
            index[mid] = row

    logger.info('customer_data_index_loaded path=%s members=%s duplicates=%s', CUSTOMER_DATA_FILE, len(index), duplicate_count)
    return index


def resolve_contact_context(contact_context):
    base_context = contact_context if isinstance(contact_context, dict) else {}
    member_id = extract_member_id_from_context(base_context)
    if not member_id:
        return base_context

    customer_row = None
    for candidate in member_id_candidates(member_id):
        maybe_row = CUSTOMER_DATA_INDEX.get(candidate)
        if isinstance(maybe_row, dict):
            customer_row = maybe_row
            break
    if not isinstance(customer_row, dict):
        logger.warning('customer_data_member_not_found member_id=%s', member_id)
        return base_context

    merged = dict(base_context)
    merged.update(customer_row)
    return merged


def resolve_json_answer(question_text, qa_pairs):
    question_raw = str(question_text or '').strip()
    if not question_raw or not qa_pairs:
        return ''

    question_norm = _normalize_text(question_raw)
    if not question_norm:
        return ''

    for row in qa_pairs:
        row_q = _normalize_text(row.get('question'))
        if not row_q:
            continue
        if question_norm == row_q or question_norm in row_q or row_q in question_norm:
            return row.get('answer') or ''

    question_tokens = _token_set(question_norm)
    best_answer = ''
    best_score = 0.0
    for row in qa_pairs:
        row_q = _normalize_text(row.get('question'))
        if not row_q:
            continue
        row_tokens = _token_set(row_q)
        overlap = len(question_tokens.intersection(row_tokens))
        ratio = difflib.SequenceMatcher(None, question_norm, row_q).ratio()
        score = (overlap * 1.0) + (ratio * 4.0)
        if score > best_score:
            best_score = score
            best_answer = row.get('answer') or ''

    if best_score >= 3.4:
        return best_answer
    return ''


def normalize_uploaded_answer_bank(value):
    pairs = []

    def add_pair(question, answer, member_id=''):
        q = str(question or '').strip()
        a = str(answer or '').strip()
        if q and a:
            pairs.append({'question': q, 'answer': a, 'member_id': str(member_id or '').strip()})

    def scan_item(item, inherited_member_id=''):
        if not isinstance(item, dict):
            return
        local_member_id = extract_member_id_from_context(item) or inherited_member_id
        add_pair(item.get('question') or item.get('Question'), item.get('answer') or item.get('Answer'), local_member_id)
        nested_keys = ('qa', 'questions', 'answers', 'items', 'data')
        for key in nested_keys:
            nested = item.get(key)
            if isinstance(nested, list):
                for row in nested:
                    scan_item(row, local_member_id)

    if isinstance(value, list):
        for row in value:
            scan_item(row)
    elif isinstance(value, dict):
        scan_item(value)
        for maybe_row in value.values():
            if isinstance(maybe_row, dict):
                scan_item(maybe_row)
            elif isinstance(maybe_row, list):
                for row in maybe_row:
                    scan_item(row)

    seen = set()
    unique_pairs = []
    for pair in pairs:
        key = (_normalize_text(pair.get('question')), pair.get('answer'), pair.get('member_id'))
        if key in seen:
            continue
        seen.add(key)
        unique_pairs.append(pair)
    return unique_pairs[:500]


def filter_uploaded_answer_bank_for_member(uploaded_answer_bank, member_id=''):
    bank = normalize_uploaded_answer_bank(uploaded_answer_bank)
    candidate_ids = set(member_id_candidates(member_id))
    if not candidate_ids:
        return bank
    scoped = [
        item for item in bank
        if not str(item.get('member_id') or '').strip()
        or bool(candidate_ids.intersection(member_id_candidates(item.get('member_id'))))
    ]
    return scoped


def normalize_campaign_contact_queue(value):
    if not isinstance(value, list):
        return []
    queue = []
    for row in value:
        if not isinstance(row, dict):
            continue
        member_id = normalize_member_id(row.get('member_id') or extract_member_id_from_context(row))
        if not member_id:
            continue
        queue_row = dict(row)
        queue_row['member_id'] = member_id
        queue.append(queue_row)
    return queue


def gemini_answer_with_context(question_text, contact_context):
    if not GEMINI_API_KEY:
        return '', 'gemini_api_key_missing'

    ctx_desc = []
    if isinstance(contact_context, dict):
        for k, v in contact_context.items():
            if v:
                ctx_desc.append(f"- {k}: {v}")
    ctx_text = "\n".join(ctx_desc)

    system_instruction = (
        "You are Casey Lane, a professional, human billing representative calling from the provider office to follow up on patient claims. "
        "Under absolutely no circumstances should you ever say or imply that you are an AI, a machine, a robot, a model, or an automated voice assistant. "
        "If asked who or what you are, say your name is Casey Lane and you are calling from the billing department of the provider office. "
        "Your task is to answer the representative's question naturally and conversational based strictly on the provided caller and patient context. "
        "Keep your answers extremely short (typically 1 to 3 words, or at most 1 short sentence), plain text, and completely human. "
        "EXAMPLES:\n"
        "- If asked what you are calling about and the context shows claim status, say 'Claims'.\n"
        "- If asked what you are calling about and the context shows eligibility verification, say: 'I'm a member' or 'Member Eligibility'.\n"
        "- If asked if you are calling as a health care professional, say: 'Yes'.\n"
        "- If asked what type of claim you are calling about, say: 'Medical'."
    )

    prompt = f"""
Patient & Campaign Context:
{ctx_text}

Payer IVR Asked:
"{question_text}"

Intelligently formulate the absolute best, most direct, and shortest response to proceed through the menu.
"""

    payload = {
        'systemInstruction': {
            'parts': [{
                'text': system_instruction
            }]
        },
        'contents': [{
            'parts': [{
                'text': prompt
            }]
        }],
        'tools': [{
            'google_search': {}
        }]
    }

    endpoint = (
        'https://generativelanguage.googleapis.com/v1beta/models/'
        'gemini-2.5-flash:generateContent'
    )
    url = f'{endpoint}?key={urllib.parse.quote(GEMINI_API_KEY)}'
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            body = resp.read().decode('utf-8', errors='ignore')
    except urllib.error.HTTPError as exc:
        err_body = exc.read().decode('utf-8', errors='ignore') if hasattr(exc, 'read') else ''
        logger.warning('customer_care_gemini_http_error status=%s body=%s', getattr(exc, 'code', 'unknown'), err_body[:500])
        return '', f'gemini_http_{getattr(exc, "code", "error")}'
    except Exception as exc:
        logger.warning('customer_care_gemini_request_failed error=%s', str(exc))
        return '', 'gemini_request_failed'

    try:
        parsed = json.loads(body)
    except Exception:
        return '', 'gemini_parse_failed'

    candidates = parsed.get('candidates') if isinstance(parsed, dict) else None
    if not isinstance(candidates, list) or not candidates:
        return '', 'gemini_empty_candidates'

    parts = (((candidates[0] or {}).get('content') or {}).get('parts') or [])
    for part in parts:
        text = str((part or {}).get('text') or '').strip()
        if text:
            return text, ''
    return '', 'gemini_empty_text'


def _to_bool_or_none(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'y', 'confirm'}:
        return True
    if text in {'0', 'false', 'no', 'n', 'deny'}:
        return False
    return None


def infer_confirmation_preference(contact_context, last_answer_text=''):
    if isinstance(contact_context, dict):
        for key in ('confirmation_expected', 'confirm_response', 'should_confirm', 'expected_confirmation'):
            parsed = _to_bool_or_none(contact_context.get(key))
            if parsed is not None:
                return parsed

    last_text = str(last_answer_text or '').strip().lower()
    if last_text:
        if any(token in last_text for token in ('deny', 'denied', 'decline', 'incorrect', 'not correct', 'no')):
            return False
        if any(token in last_text for token in ('confirm', 'confirmed', 'correct', 'yes')):
            return True

    return True


def resolve_confirmation_digit(asked_text, contact_context, last_answer_text='', state=None):
    raw_text = str(asked_text or '').strip().lower()
    if not raw_text:
        return ''

    # If the representative is asking to use the subscriber's SSN instead
    if 'social security number' in raw_text or 'ssn' in raw_text or 'subscriber ssn' in raw_text:
        ssn_value = _digits_only(_context_value_by_key(contact_context, 'ssn'))
        if ssn_value:
            return '1'
        else:
            return '2'

    # If the representative is asking "same patient or different patient"
    if 'same patient' in raw_text or 'different patient' in raw_text or ('same' in raw_text and 'different' in raw_text):
        if state and isinstance(state, dict):
            queue = normalize_campaign_contact_queue(state.get('campaign_contact_queue'))
            has_next_member = bool(queue) or bool(state.get('campaign_has_next_member'))
            if has_next_member:
                return '2'
            else:
                return '1'
        return '2'

    word_to_digit = {
        'zero': '0', 'one': '1', 'two': '2', 'three': '3', 'four': '4',
        'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9',
    }
    text = raw_text
    for word, digit in word_to_digit.items():
        text = re.sub(rf'\b{word}\b', digit, text)

    if 'press' not in text:
        return ''

    press_hits = list(re.finditer(r'press\s*([0-9])\b', text))
    if not press_hits:
        return ''

    yes_keywords = r'yes|right|correct|confirm|confirmed|accurate|valid'
    no_keywords = r'no|wrong|incorrect|deny|denied|invalid|not\s+correct'

    yes_digit = ''
    no_digit = ''

    for hit in press_hits:
        digit = hit.group(1)
        start = max(0, hit.start() - 60)
        end = min(len(text), hit.end() + 90)
        window = text[start:end]
        if not yes_digit and re.search(yes_keywords, window):
            yes_digit = digit
        if not no_digit and re.search(no_keywords, window):
            no_digit = digit

    if not yes_digit:
        m = re.search(rf'\b({yes_keywords})\b[^.\n]{{0,70}}press\s*([0-9])\b', text)
        if m:
            yes_digit = m.group(2)
    if not no_digit:
        m = re.search(rf'\b({no_keywords})\b[^.\n]{{0,70}}press\s*([0-9])\b', text)
        if m:
            no_digit = m.group(2)

    # Only match confirmation digits if we have an explicit yes/no keyword context.
    # Falling back to arbitrary press digits in general menus is extremely dangerous!
    if not yes_digit:
        return ''

    if not no_digit and yes_digit:
        for hit in press_hits:
            candidate = hit.group(1)
            if candidate != yes_digit:
                no_digit = candidate
                break

    if no_digit and re.search(r'\b(if\s+)?(correct|confirm|yes)\b', text) and re.search(r'\b(else|otherwise|if\s+not|if\s+no|wrong|incorrect|cancel)\b', text):
        compare_target = ''
        if any(token in text for token in ('member id', 'patient id', 'aetna id', 'member i d', 'patient i d', 'you entered id')):
            compare_target = _digits_only(_ctx_first(contact_context, ['member_id', 'patient_id', 'member_id_answer', 'aetna_patient_id', 'aetna_member_id']))
        elif any(token in text for token in ('date of birth', 'dob')):
            compare_target = _date_to_dtmf(_ctx_first(contact_context, ['member_dob', 'dob', 'date_of_birth']))
        elif any(token in text for token in ('date of service', 'dos')):
            compare_target = _date_to_dtmf(_ctx_first(contact_context, ['date_of_service', 'dos', 'service_date']))
        elif 'fax' in text and 'number' in text:
            compare_target = _digits_only(_ctx_first(contact_context, ['fax_number', 'provider_fax', 'fax', 'callback_number']))
        elif any(token in text for token in ('npi', 'tax id')):
            compare_target = _digits_only(_ctx_first(contact_context, ['npi', 'provider_npi', 'tax_id']))

        in_prompt_digits = ''.join(re.findall(r'\d+', text))
        if compare_target:
            if not in_prompt_digits or compare_target not in in_prompt_digits:
                return no_digit

    wants_confirm = infer_confirmation_preference(contact_context, last_answer_text=last_answer_text)
    if wants_confirm:
        return yes_digit
    return no_digit or yes_digit


DEFAULT_IVR_REPRESENTATIVE_KEYWORDS = [
    'representative', 'agent', 'customer service', 'customer care',
    'operator', 'speak to someone', 'live person', 'talk to someone',
]
DEFAULT_IVR_MAX_ATTEMPTS = 6
IVR_STATIC_PHASES = ('language', 'department', 'submenu', 'representative')


def find_digit_for_keywords(text, keywords, window_before=60, window_after=90):
    word_to_digit = {
        'zero': '0', 'one': '1', 'two': '2', 'three': '3', 'four': '4',
        'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9',
    }
    normalized_text = str(text or '').strip().lower()
    if not normalized_text:
        return '', ''
    for word, digit in word_to_digit.items():
        normalized_text = re.sub(rf'\b{word}\b', digit, normalized_text)

    if 'press' not in normalized_text:
        return '', ''

    press_hits = list(re.finditer(r'press\s*([0-9*#])\b', normalized_text))
    if not press_hits:
        return '', ''

    safe_keywords = [str(kw or '').strip().lower() for kw in (keywords or []) if str(kw or '').strip()]
    if not safe_keywords:
        return '', ''

    for hit in press_hits:
        digit = hit.group(1)
        start = max(0, hit.start() - window_before)
        end = min(len(normalized_text), hit.end() + window_after)
        window = normalized_text[start:end]
        for keyword in safe_keywords:
            if keyword in window:
                return digit, keyword
    return '', ''


def extract_first_press_digit(text):
    word_to_digit = {
        'zero': '0', 'one': '1', 'two': '2', 'three': '3', 'four': '4',
        'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9',
    }
    normalized_text = str(text or '').strip().lower()
    if not normalized_text:
        return ''
    for word, digit in word_to_digit.items():
        normalized_text = re.sub(rf'\b{word}\b', digit, normalized_text)
    m = re.search(r'press\s*([0-9*#])\b', normalized_text)
    return m.group(1) if m else ''


def normalize_static_route(raw_route):
    route = {}
    if not isinstance(raw_route, dict):
        return route
    for phase in IVR_STATIC_PHASES:
        raw_step = raw_route.get(phase)
        if isinstance(raw_step, str):
            raw_step = {'digit': raw_step}
        if not isinstance(raw_step, dict):
            continue
        digit = str(raw_step.get('digit') or '').strip()
        if not re.fullmatch(r'[0-9#*]', digit):
            continue
        prompt_keywords = _parse_keyword_list(raw_step.get('prompt_keywords'))
        route[phase] = {
            'digit': digit,
            'prompt_keywords': prompt_keywords,
        }
    return route


def should_trigger_static_route_step(step, speech_result):
    if not isinstance(step, dict):
        return False
    digit = str(step.get('digit') or '').strip()
    if not re.fullmatch(r'[0-9#*]', digit):
        return False
    speech_norm = _normalize_text(speech_result)
    if not speech_norm:
        return False
    prompt_keywords = [kw for kw in (step.get('prompt_keywords') or []) if kw]
    if not prompt_keywords:
        return True
    return any(_normalize_text(keyword) in speech_norm for keyword in prompt_keywords)


IVR_CATEGORY_CATALOG = [
    {
        'id': 'eligibility',
        'label': 'Eligibility and benefits verification',
        'keywords': ['eligibility', 'benefit verification', 'benefits verification', 'eligibility and benefits', 'verify benefits'],
        'context_signals': ['eligibility_answer', 'benefits_answer', 'coverage_answer', 'plan_type'],
        'submenus': [],
    },
    {
        'id': 'prior_authorization',
        'label': 'Prior authorization or precertification requests',
        'keywords': ['prior authorization', 'preauthorization', 'pre-authorization', 'precertification', 'pre-certification', 'authorization'],
        'context_signals': ['auth_answer', 'authorization_answer', 'prior_auth_answer', 'precert_answer'],
        'submenus': [
            {'id': 'auth_status', 'keywords': ['status of an authorization', 'authorization status', 'check on an authorization'], 'context_signals': ['auth_answer', 'authorization_answer']},
            {'id': 'auth_new_request', 'keywords': ['new authorization', 'new precertification', 'submit a request'], 'context_signals': []},
        ],
    },
    {
        'id': 'claims_status',
        'label': 'Claims status, denials, and appeals',
        'keywords': ['claims status', 'claim status', 'claims', 'denials', 'denial', 'appeals', 'appeal'],
        'context_signals': ['claim_status_answer', 'denial_reason', 'appeal_deadline', 'appeal_address'],
        'submenus': [
            {'id': 'denial', 'keywords': ['denial', 'denied claim'], 'context_signals': ['denial_reason']},
            {'id': 'appeal', 'keywords': ['appeal'], 'context_signals': ['appeal_deadline', 'appeal_address']},
            {'id': 'claim_status', 'keywords': ['claim status', 'status of a claim'], 'context_signals': ['claim_status_answer']},
        ],
    },
    {
        'id': 'provider_enrollment',
        'label': 'Provider enrollment, credentialing, or demographics updates',
        'keywords': ['provider enrollment', 'credentialing', 'demographics update', 'demographic update', 'provider demographics'],
        'context_signals': ['npi', 'provider_id', 'credentialing_answer', 'tax_id'],
        'submenus': [],
    },
    {
        'id': 'pharmacy_benefits',
        'label': 'Pharmacy benefits and formulary exceptions',
        'keywords': ['pharmacy benefits', 'pharmacy benefit', 'formulary exception', 'formulary', 'pharmacy'],
        'context_signals': ['pharmacy_answer', 'formulary_answer', 'rx_answer'],
        'submenus': [],
    },
]


def _match_ivr_category_by_reason(reason_text):
    reason_norm = _normalize_text(reason_text)
    if not reason_norm:
        return None
    for category in IVR_CATEGORY_CATALOG:
        candidates = [category['id'].replace('_', ' '), category['label']] + category['keywords']
        for candidate in candidates:
            candidate_norm = _normalize_text(candidate)
            if candidate_norm and (candidate_norm in reason_norm or reason_norm in candidate_norm):
                return category
    return None


def infer_ivr_category(contact_context, explicit_reason=''):
    explicit_match = _match_ivr_category_by_reason(explicit_reason)
    if explicit_match:
        return explicit_match, 'explicit_reason'

    if not isinstance(contact_context, dict) or not contact_context:
        return None, ''

    best_category = None
    best_score = 0
    for category in IVR_CATEGORY_CATALOG:
        score = sum(1 for signal in category.get('context_signals', []) if str(contact_context.get(signal) or '').strip())
        if score > best_score:
            best_score = score
            best_category = category
    if best_category:
        return best_category, 'context_signals'
    return None, ''


def infer_ivr_submenu(category, contact_context):
    if not category or not category.get('submenus'):
        return None
    if not isinstance(contact_context, dict):
        contact_context = {}
    best_submenu = None
    best_score = 0
    for submenu in category['submenus']:
        score = sum(1 for signal in submenu.get('context_signals', []) if str(contact_context.get(signal) or '').strip())
        if score > best_score:
            best_score = score
            best_submenu = submenu
    return best_submenu


def normalize_ivr_navigation_config(payload):
    payload = payload if isinstance(payload, dict) else {}

    def as_keyword_list(value):
        if isinstance(value, str):
            value = [part.strip() for part in value.split(',')]
        if not isinstance(value, list):
            return []
        return [str(kw).strip().lower() for kw in value if str(kw or '').strip()][:20]

    department_keywords = as_keyword_list(payload.get('department_keywords'))
    submenu_keywords = as_keyword_list(payload.get('submenu_keywords'))
    representative_keywords = as_keyword_list(payload.get('representative_keywords'))

    try:
        max_attempts = int(payload.get('max_attempts') or DEFAULT_IVR_MAX_ATTEMPTS)
    except (TypeError, ValueError):
        max_attempts = DEFAULT_IVR_MAX_ATTEMPTS
    max_attempts = max(1, min(max_attempts, 15))

    enabled = _to_bool(payload.get('enabled'), ENABLE_AUTO_IVR_NAVIGATION_DEFAULT)

    return {
        'enabled': enabled,
        'department_keywords': department_keywords,
        'submenu_keywords': submenu_keywords,
        'representative_keywords': representative_keywords,
        'language_keywords': as_keyword_list(payload.get('language_keywords')),
        'language_preferred': str(payload.get('language_preferred') or '').strip().lower(),
        'welcome_text': str(payload.get('welcome_text') or '').strip(),
        'welcome_detected': _to_bool(payload.get('welcome_detected'), False),
        'language_selected': _to_bool(payload.get('language_selected'), False),
        'phase': str(payload.get('phase') or 'ivr_welcome').strip() or 'ivr_welcome',
        'max_attempts': max_attempts,
        'matched_category': '',
        'matched_category_source': '',
        'profile_name': '',
        'profile_source': '',
        'customer_care_number': '',
        'static_route': normalize_static_route(payload.get('static_route')),
        'allowed_profiles': [],
        'ivr_profile_error': '',
    }


def apply_contact_ivr_overrides(config, contact_context):
    config = dict(config or {})
    ctx = contact_context if isinstance(contact_context, dict) else {}

    row_welcome = get_contact_context_value(ctx, ['ivr_welcome_text', 'ivr welcome text'])
    row_language = get_contact_context_value(ctx, ['ivr_language_keywords', 'ivr language keywords'])
    row_department = get_contact_context_value(ctx, ['ivr_department_keywords', 'ivr department keywords'])
    row_submenu = get_contact_context_value(ctx, ['ivr_submenu_keywords', 'ivr submenu keywords'])
    row_representative = get_contact_context_value(ctx, ['ivr_representative_keywords', 'ivr representative keywords'])
    row_max_attempts = get_contact_context_value(ctx, ['ivr_max_attempts', 'ivr max attempts'])

    source_tags = []
    if str(row_welcome or '').strip():
        config['welcome_text'] = str(row_welcome).strip()
        source_tags.append('row_welcome')

    language_keywords = _parse_keyword_list(row_language)
    if language_keywords:
        config['language_keywords'] = language_keywords
        config['language_preferred'] = language_keywords[0]
        source_tags.append('row_language')

    department_keywords = _parse_keyword_list(row_department)
    if department_keywords:
        config['department_keywords'] = department_keywords
        source_tags.append('row_department')

    submenu_keywords = _parse_keyword_list(row_submenu)
    if submenu_keywords:
        config['submenu_keywords'] = submenu_keywords
        source_tags.append('row_submenu')

    representative_keywords = _parse_keyword_list(row_representative)
    if representative_keywords:
        config['representative_keywords'] = representative_keywords
        source_tags.append('row_representative')

    if str(row_max_attempts or '').strip():
        try:
            max_attempts = int(str(row_max_attempts).strip())
            config['max_attempts'] = max(1, min(max_attempts, 15))
            source_tags.append('row_max_attempts')
        except (TypeError, ValueError):
            pass

    if source_tags:
        config['profile_source'] = ','.join(source_tags)
    return config


def resolve_ivr_profile_for_contact(contact_context):
    profile = IVR_PROFILE if isinstance(IVR_PROFILE, dict) else {}
    profiles = profile.get('profiles') if isinstance(profile.get('profiles'), dict) else {}

    profile_alias_map = {}
    for profile_id, profile_data in profiles.items():
        normalized_id = _normalize_profile_id(profile_id)
        if normalized_id:
            profile_alias_map[normalized_id] = normalized_id
        embedded_name = _normalize_profile_id((profile_data or {}).get('profile_name'))
        if embedded_name and embedded_name not in profile_alias_map:
            profile_alias_map[embedded_name] = normalized_id

    allowed_profiles = sorted(list(profiles.keys()))
    source = 'contact_row_ivr_profile'

    row_profile_raw = get_contact_context_value(contact_context, ['ivr_profile', 'ivr profile', 'profile_name', 'profile name', 'ivr_profile_name', 'ivr profile name'])
    row_profile = _normalize_profile_id(row_profile_raw)
    resolved_row_profile = profile_alias_map.get(row_profile, row_profile)
    if not resolved_row_profile:
        return None, source, {
            'code': 'IVR_PROFILE_REQUIRED',
            'message': 'Missing ivr profile in contact row. Provide ivr_profile/profile_name that matches ivr.json profiles.',
            'allowed_profiles': allowed_profiles,
        }
    if resolved_row_profile not in profiles:
        return None, source, {
            'code': 'IVR_PROFILE_NOT_FOUND',
            'message': f'Unknown ivr profile "{row_profile_raw}". It must match one of the profiles in ivr.json.',
            'allowed_profiles': allowed_profiles,
        }

    selected = profiles[resolved_row_profile]
    selected_profile = normalize_single_ivr_profile(selected, fallback_profile_name=selected.get('profile_name') if isinstance(selected, dict) else resolved_row_profile)
    return selected_profile, source, None


def is_welcome_prompt_detected(expected_welcome_text, observed_text):
    expected_tokens = [tok for tok in _normalize_text(expected_welcome_text).split(' ') if len(tok) >= 3]
    observed_tokens = set(tok for tok in _normalize_text(observed_text).split(' ') if len(tok) >= 3)
    if not expected_tokens or not observed_tokens:
        return False
    overlap = sum(1 for tok in expected_tokens if tok in observed_tokens)
    ratio = overlap / max(1, len(set(expected_tokens)))
    return ratio >= 0.35 and overlap >= 2


def is_common_hold_message(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    hold_markers = [
        '1 moment',
        'one moment',
        'just a moment',
        'please hold',
        'put your call on hold',
        'please stay on the line',
        'did not hear anything from you',
        'your call is on hold',
        'speaking with has put your call on hold',
    ]
    return any(marker in normalized for marker in hold_markers)


def build_ivr_navigation_config(payload, contact_context, call_reason=''):
    config = normalize_ivr_navigation_config(payload)

    if not config['enabled']:
        return config

    profile, profile_source, profile_error = resolve_ivr_profile_for_contact(contact_context)
    if profile_error:
        config['allowed_profiles'] = profile_error.get('allowed_profiles') or []
        config['ivr_profile_error'] = profile_error.get('message') or 'Invalid ivr profile.'
        config['profile_source'] = profile_source
        config['profile_name'] = ''
        return config
    config['profile_name'] = str(profile.get('profile_name') or '').strip()
    config['profile_source'] = profile_source
    config['customer_care_number'] = _normalize_customer_care_number(
        get_contact_context_value(contact_context, ['customer_care', 'customer care', 'customer_care_number', 'customer care number'])
    )

    config = apply_contact_ivr_overrides(config, contact_context)

    if not config['welcome_text']:
        config['welcome_text'] = str(profile.get('welcome_text') or '').strip()
    if not config['language_keywords']:
        config['language_keywords'] = [str(k).strip().lower() for k in ((profile.get('language') or {}).get('keywords') or []) if str(k).strip()]
    if not config.get('language_preferred'):
        config['language_preferred'] = str((profile.get('language') or {}).get('preferred') or '').strip().lower()
    if not config['representative_keywords']:
        config['representative_keywords'] = [str(k).strip().lower() for k in (profile.get('representative_keywords') or []) if str(k).strip()]
    if not config['representative_keywords']:
        config['representative_keywords'] = list(DEFAULT_IVR_REPRESENTATIVE_KEYWORDS)
    if not isinstance(config.get('static_route'), dict) or not config.get('static_route'):
        config['static_route'] = normalize_static_route(profile.get('static_route'))

    menu_goals = [str(item).strip() for item in (profile.get('menu_goals') or []) if str(item).strip()]
    combined_reason = call_reason
    if menu_goals:
        joined = ', '.join(menu_goals)
        combined_reason = f'{combined_reason}, {joined}'.strip(', ').strip()

    if config['department_keywords']:
        return config

    matched_category, match_source = infer_ivr_category(contact_context, combined_reason)
    if not matched_category:
        return config

    config['department_keywords'] = list(matched_category['keywords'])
    config['matched_category'] = matched_category['id']
    config['matched_category_source'] = match_source

    if not config['submenu_keywords'] and matched_category.get('submenus'):
        matched_submenu = infer_ivr_submenu(matched_category, contact_context)
        if matched_submenu:
            config['submenu_keywords'] = list(matched_submenu['keywords'])
        else:
            config['submenu_keywords'] = list(dict.fromkeys(
                kw for submenu in matched_category['submenus'] for kw in submenu['keywords']
            ))

    return config


def build_ivr_navigation_twiml(play_digits=''):
    callback_url = get_public_base_url()
    digit_block = f'<Play digits="{play_digits}"/>' if play_digits else ''
    return f'''
    <Response>
        {digit_block}
        <Gather input="speech dtmf" action="{callback_url}/api/webhook/ivr_menu" method="POST" timeout="6" speechTimeout="3">
            <Pause length="20"/>
        </Gather>
        <Redirect method="POST">{callback_url}/api/webhook/ivr_menu</Redirect>
    </Response>
    '''


def build_customer_care_twiml(prompt_text, voice, followup=True, play_digits=''):
    callback_url = get_public_base_url()
    safe_prompt = safe_text(prompt_text)
    digit_block = f'<Play digits="{play_digits}"/>' if play_digits else ''
    prompt_block = f'<Say voice="{voice}">{safe_prompt}</Say>' if safe_prompt else ''

    return f'''
    <Response>
        {digit_block}
        <Gather input="speech dtmf" action="{callback_url}/api/webhook/speech" method="POST" timeout="8" speechTimeout="3">
            {prompt_block}
            <Pause length="60"/>
        </Gather>
        <Redirect method="POST">{callback_url}/api/webhook/speech</Redirect>
    </Response>
    '''


def build_end_call_twiml(prompt_text, voice):
    safe_prompt = safe_text(prompt_text)
    return f'''
    <Response>
        <Say voice="{voice}">{safe_prompt}</Say>
        <Hangup/>
    </Response>
    '''


def resolve_customer_care_answer(asked_text, contact_context, uploaded_answer_bank=None, active_member_id='', call_sid=''):
    if is_passive_announcement_or_disclaimer(asked_text):
        return '', 'statement_silent_listen'

    answer_text = resolve_contact_answer(asked_text, contact_context)
    if answer_text:
        return _digitize_numeric_phrases(answer_text), 'contacts_sheet'

    scoped_bank = filter_uploaded_answer_bank_for_member(uploaded_answer_bank, active_member_id or extract_member_id_from_context(contact_context))
    json_answer = resolve_json_answer(asked_text, scoped_bank)
    if json_answer:
        return _digitize_numeric_phrases(json_answer), 'uploaded_json'

    profile_name = ''
    if call_sid:
        try:
            state = read_call_state(call_sid)
            profile_name = str(state.get('selected_profile_name') or '').strip()
        except Exception:
            pass

    if profile_name in {'aetna_representative', 'united_health_care', 'cigna', 'Humana'}:
        logger.info('no_rule_matched_automated_profile_silent call_sid=%s profile=%s', call_sid, profile_name)
        return '', 'statement_silent_listen'

    gemini_answer, gemini_error = gemini_answer_with_context(asked_text, contact_context)
    if gemini_answer:
        return _digitize_numeric_phrases(gemini_answer), 'gemini_google'

    fallback_answer = choose_non_repeating_fallback_reply(call_sid, asked_text, contact_context) if call_sid else build_human_fallback_answer(asked_text, contact_context, profile_name=profile_name)
    logger.warning('customer_care_answer_fallback asked_text=%s gemini_error=%s fallback=%s', asked_text, gemini_error, fallback_answer)
    return fallback_answer, 'fallback_human'


def build_answer_consistency_check(asked_text, answer_text, answer_source, contact_context, active_member_id='', uploaded_answer_bank=None):
    flags = []
    safe_source = str(answer_source or '').strip()
    safe_active_member_id = normalize_member_id(active_member_id)
    context_member_id = normalize_member_id(extract_member_id_from_context(contact_context))

    if safe_active_member_id and context_member_id:
        active_candidates = set(member_id_candidates(safe_active_member_id))
        context_candidates = set(member_id_candidates(context_member_id))
        if active_candidates and context_candidates and not active_candidates.intersection(context_candidates):
            flags.append('member_id_mismatch')

    if safe_source in {'contacts_sheet', 'uploaded_json', 'gemini_google'} and not str(answer_text or '').strip():
        flags.append('empty_answer_text')

    if safe_source == 'contacts_sheet':
        expected = resolve_contact_answer(asked_text, contact_context)
        if not expected:
            flags.append('contacts_source_without_match')

    if safe_source == 'uploaded_json':
        member_scope = safe_active_member_id or context_member_id
        scoped_bank = filter_uploaded_answer_bank_for_member(uploaded_answer_bank or [], member_scope)
        expected = resolve_json_answer(asked_text, scoped_bank)
        if not expected:
            flags.append('uploaded_json_source_without_match')

    status = 'ok' if not flags else 'warning'
    return {
        'status': status,
        'flags': flags,
        'answer_source': safe_source,
        'active_member_id': safe_active_member_id,
        'context_member_id': context_member_id,
    }


# Runtime answering uses per-call uploaded_answer_bank from UI uploads.
CUSTOMER_CARE_QA_PAIRS = []
CUSTOMER_DATA_INDEX = load_customer_data_index()
IVR_PROFILE = load_ivr_profile()
PAYER_PROFILES = load_payer_profiles()


def save_recording_media_to_local(call_sid, recording_sid, storage_options):
    twilio_client, context = get_twilio_context_for_call(call_sid)
    if not twilio_client or not context:
        return None, 'TWILIO_CONTEXT_MISSING'

    recording = telephony_provider.fetch_recording(twilio_client, recording_sid)
    media_url = (getattr(recording, 'media_url', '') or '').strip()
    if not media_url:
        media_url = f"https://api.twilio.com/2010-04-01/Accounts/{context['account_sid']}/Recordings/{recording_sid}"
    media_url = f"{media_url}.mp3" if not media_url.endswith('.mp3') else media_url

    creds = f"{context['account_sid']}:{context['auth_token']}".encode('utf-8')
    auth_header = f"Basic {base64.b64encode(creds).decode('ascii')}"
    req = urllib.request.Request(media_url, headers={'Authorization': auth_header, 'User-Agent': 'calling-agent/1.0'})
    logger.info('artifact_download_start call_sid=%s recording_sid=%s media_url=%s mode=%s subfolder=%s', call_sid, recording_sid, media_url, storage_options.get('mode'), storage_options.get('subfolder'))
    with urllib.request.urlopen(req, timeout=30) as resp:
        payload = resp.read()
        logger.info('artifact_download_ok call_sid=%s recording_sid=%s bytes=%s status=%s', call_sid, recording_sid, len(payload), getattr(resp, 'status', 'unknown'))

    artifact_dir = get_artifact_call_dir(call_sid, storage_options)
    logger.info('artifact_directory call_sid=%s path=%s', call_sid, artifact_dir)
    file_name = f"recording-{recording_sid}.mp3"
    full_path = os.path.join(artifact_dir, file_name)
    with open(full_path, 'wb') as handle:
        handle.write(payload)
    logger.info('artifact_recording_saved call_sid=%s recording_sid=%s path=%s', call_sid, recording_sid, full_path)

    metadata_path = os.path.join(artifact_dir, f"recording-{recording_sid}.json")
    save_json_file(metadata_path, {
        'call_sid': call_sid,
        'recording_sid': recording_sid,
        'saved_at': datetime.now(timezone.utc).isoformat(),
        'source_url': media_url,
        'file_name': file_name,
        'bytes': len(payload),
    })
    logger.info('artifact_recording_metadata_saved call_sid=%s recording_sid=%s path=%s', call_sid, recording_sid, metadata_path)
    return full_path, None


def gemini_transcribe_audio(file_path):
    if not GEMINI_API_KEY:
        return '', 'gemini_api_key_missing'
    if not os.path.exists(file_path):
        return '', 'audio_file_not_found'
    
    try:
        with open(file_path, 'rb') as f:
            audio_data = f.read()
        
        encoded_data = base64.b64encode(audio_data).decode('utf-8')
        
        payload = {
            'contents': [{
                'parts': [
                    {
                        'inlineData': {
                            'mimeType': 'audio/mp3',
                            'data': encoded_data
                        }
                    },
                    {
                        'text': 'Please provide a complete and accurate word-for-word transcript of this phone call recording. Do not add any conversational intro, outro, or meta-comments. Just output the clean transcript text.'
                    }
                ]
            }]
        }
        
        endpoint = (
            'https://generativelanguage.googleapis.com/v1beta/models/'
            'gemini-2.5-flash:generateContent'
        )
        url = f'{endpoint}?key={urllib.parse.quote(GEMINI_API_KEY)}'
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
            method='POST',
        )
        
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode('utf-8', errors='ignore')
            
        parsed = json.loads(body)
        candidates = parsed.get('candidates') if isinstance(parsed, dict) else None
        if isinstance(candidates, list) and candidates:
            parts = candidates[0].get('content', {}).get('parts', [])
            for part in parts:
                text = str(part.get('text') or '').strip()
                if text:
                    return text, ''
        return '', 'gemini_empty_text'
    except Exception as exc:
        logger.warning('gemini_transcription_failed error=%s', str(exc))
        return '', str(exc)


def request_recording_transcription(call_sid, recording_sid):
    import threading

    def _perform_transcription():
        logger.info('gemini_transcription_thread_start call_sid=%s recording_sid=%s', call_sid, recording_sid)
        state = read_call_state(call_sid)
        artifacts = state.get('artifacts', {})
        storage_options = state.get('storage_options') or build_storage_options({})
        
        saved_path = artifacts.get('recording_saved_path')
        temp_download = False
        
        # If recording was not saved locally, download it temporarily for transcription
        if not saved_path or not os.path.exists(saved_path):
            try:
                saved_path, download_error = save_recording_media_to_local(call_sid, recording_sid, storage_options)
                if download_error:
                    logger.warning('gemini_transcription_download_failed call_sid=%s error=%s', call_sid, download_error)
                    update_call_state(call_sid, {
                        'transcription': {
                            'enabled': True,
                            'status': 'unavailable',
                            'source': 'gemini',
                            'transcription_sid': '',
                            'text': '',
                            'error': f"Download failed: {download_error}",
                        }
                    })
                    return
                temp_download = not storage_options.get('save_enabled')
            except Exception as exc:
                logger.exception('gemini_transcription_download_exception call_sid=%s error=%s', call_sid, str(exc))
                update_call_state(call_sid, {
                    'transcription': {
                        'enabled': True,
                        'status': 'unavailable',
                        'source': 'gemini',
                        'transcription_sid': '',
                        'text': '',
                        'error': f"Download exception: {str(exc)}",
                    }
                })
                return

        # Perform Gemini transcription
        try:
            logger.info('gemini_transcription_api_start call_sid=%s path=%s', call_sid, saved_path)
            transcript_text, transcribe_error = gemini_transcribe_audio(saved_path)
            
            if transcribe_error:
                logger.warning('gemini_transcription_api_failed call_sid=%s error=%s', call_sid, transcribe_error)
                update_call_state(call_sid, {
                    'transcription': {
                        'enabled': True,
                        'status': 'unavailable',
                        'source': 'gemini',
                        'transcription_sid': '',
                        'text': '',
                        'error': transcribe_error,
                    }
                })
                return
            
            logger.info('gemini_transcription_api_success call_sid=%s text_len=%s', call_sid, len(transcript_text))
            
            transcription_state = {
                'enabled': True,
                'status': 'completed',
                'source': 'gemini',
                'transcription_sid': f"GM{recording_sid}",
                'recording_sid': recording_sid,
                'text': transcript_text,
                'error': '',
            }
            
            # Save the transcript file if saving is enabled
            if storage_options.get('save_enabled'):
                try:
                    artifact_dir = get_artifact_call_dir(call_sid, storage_options)
                    transcript_path = os.path.join(artifact_dir, f"transcript-{recording_sid}.txt")
                    save_text_file(transcript_path, transcript_text)
                    artifacts['transcription_saved_path'] = transcript_path
                    artifacts['last_error'] = ''
                    logger.info('gemini_transcript_saved_to_artifacts call_sid=%s path=%s', call_sid, transcript_path)
                except Exception as exc:
                    artifacts['last_error'] = str(exc)
                    logger.exception('gemini_transcript_save_exception call_sid=%s error=%s', call_sid, str(exc))

            # Also save normalized interaction transcript if there is one
            interaction = state.get('interaction_transcript', [])
            if storage_options.get('save_enabled') and interaction:
                try:
                    artifact_dir = get_artifact_call_dir(call_sid, storage_options)
                    interaction_path = os.path.join(artifact_dir, 'interaction-transcript.json')
                    save_json_file(interaction_path, normalize_interaction_transcript(interaction))
                    artifacts['interaction_saved_path'] = interaction_path
                    logger.info('gemini_interaction_saved_to_artifacts call_sid=%s path=%s entries=%s', call_sid, interaction_path, len(interaction))
                    
                    try:
                        logger.info('triggering_excel_audit_update call_sid=%s', call_sid)
                        dialogue_lines = []
                        for entry in interaction:
                            speaker = str(entry.get('speaker') or 'unknown')
                            text = str(entry.get('text') or '').strip()
                            dialogue_lines.append(f"{speaker}: {text}")
                        dialogue_text = "\n".join(dialogue_lines)
                        
                        extracted_audits = extract_claim_details_via_gemini(dialogue_text)
                        if extracted_audits and isinstance(extracted_audits, dict):
                            audits_list = extracted_audits.get('audits') if isinstance(extracted_audits.get('audits'), list) else [extracted_audits]
                            update_contacts_excel_with_audits(audits_list, call_sid=call_sid)
                            
                            # Append structured audits to the interaction transcript JSON!
                            if audits_list:
                                audit_data = audits_list[0]
                                audit_msg = f"[AUDIT RESULTS] claim_status: {audit_data.get('claim_status')} | allowable_amount: {audit_data.get('allowable_amount')} | total_paid_amount: {audit_data.get('total_paid_amount')} | co_payment_amount: {audit_data.get('co_payment_amount')} | co_insurance_amount: {audit_data.get('co_insurance_amount')} | eft_trace_number: {audit_data.get('eft_trace_number')}"
                                
                                # Read fresh state & append
                                fresh_state = read_call_state(call_sid)
                                fresh_transcript = fresh_state.get('interaction_transcript') or []
                                fresh_transcript.append({
                                    'ts': datetime.now().astimezone().isoformat(),
                                    'role': 'agent',
                                    'speaker': 'calling_agent',
                                    'text': audit_msg,
                                    'source': 'claim_details_audit',
                                    'audits': audit_data
                                })
                                fresh_state['interaction_transcript'] = fresh_transcript
                                write_call_state(call_sid, fresh_state)
                                
                                # Re-save the interaction-transcript.json file with the audit block included!
                                save_json_file(interaction_path, normalize_interaction_transcript(fresh_transcript))
                    except Exception as excel_exc:
                        logger.warning('failed_to_process_excel_audit_in_background call_sid=%s error=%s', call_sid, str(excel_exc))
                except Exception as exc:
                    artifacts['last_error'] = str(exc)
                    logger.exception('gemini_interaction_save_exception call_sid=%s error=%s', call_sid, str(exc))

            update_call_state(call_sid, {
                'transcription': transcription_state,
                'artifacts': artifacts
            })
            
        except Exception as exc:
            logger.exception('gemini_transcription_exception call_sid=%s error=%s', call_sid, str(exc))
            update_call_state(call_sid, {
                'transcription': {
                    'enabled': True,
                    'status': 'unavailable',
                    'source': 'gemini',
                    'transcription_sid': '',
                    'text': '',
                    'error': str(exc),
                }
            })
        finally:
            if temp_download and saved_path and os.path.exists(saved_path):
                try:
                    os.remove(saved_path)
                    logger.info('gemini_transcription_temp_file_removed path=%s', saved_path)
                except Exception as exc:
                    logger.warning('gemini_transcription_temp_file_remove_failed path=%s error=%s', saved_path, str(exc))

    # Mark transcription as requested/in-progress
    update_call_state(call_sid, {
        'transcription': {
            'enabled': True,
            'status': 'requested',
            'source': 'gemini',
            'transcription_sid': f"GM{recording_sid}",
            'text': '',
            'error': '',
        }
    })
    
    # Start the transcription in a background thread so webhooks don't block
    threading.Thread(target=_perform_transcription).start()


def get_public_base_url():
    if PUBLIC_BASE_URL:
        return PUBLIC_BASE_URL
    proto = request.headers.get('X-Forwarded-Proto', request.scheme)
    host = request.headers.get('X-Forwarded-Host', request.host)
    return f"{proto}://{host}".rstrip('/')


def _request_url_for_signature():
    if PUBLIC_BASE_URL:
        query = request.query_string.decode('utf-8') if request.query_string else ''
        base_url = f"{PUBLIC_BASE_URL}{request.path}"
        return f"{base_url}?{query}" if query else base_url
    return request.url


def _is_valid_twilio_webhook_request():
    if not WEBHOOK_SIGNATURE_VALIDATION:
        return True
    if not validator:
        return False

    twilio_signature = request.headers.get('X-Twilio-Signature', '')
    if not twilio_signature:
        return False

    post_vars = request.form.to_dict(flat=True)
    url = _request_url_for_signature()

    call_sid = post_vars.get('CallSid')
    context = call_twilio_credentials.get(call_sid or '')
    if context and context.get('auth_token'):
        return telephony_provider.validate_webhook_signature(context['auth_token'], url, post_vars, twilio_signature)

    persisted_token = ''
    if call_sid:
        state = read_call_state(call_sid)
        persisted_token = str((state or {}).get('_twilio_auth_token') or '').strip()
    if persisted_token:
        return telephony_provider.validate_webhook_signature(persisted_token, url, post_vars, twilio_signature)

    return telephony_provider.validate_webhook_signature(TWILIO_AUTH_TOKEN, url, post_vars, twilio_signature)


@app.before_request
def verify_twilio_signature_for_webhooks():
    if request.path.startswith('/api/webhook/') and not _is_valid_twilio_webhook_request():
        logger.warning('api_webhook_invalid_signature path=%s remote_addr=%s', request.path, request.remote_addr)
        return jsonify({'error': 'Invalid webhook signature.'}), 403


@app.after_request
def set_security_headers(response):
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "base-uri 'self'; "
        "object-src 'none'; "
        "frame-ancestors 'none'; "
        "form-action 'self'; "
        "img-src 'self' data: blob:; "
        "style-src 'self'; "
        "font-src 'self' data:; "
        "script-src 'self'; "
        "connect-src 'self' https://generativelanguage.googleapis.com; "
        "upgrade-insecure-requests"
    )
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
    return response


@app.route('/', methods=['GET'])
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')


@app.route('/static/<path:asset_path>', methods=['GET'])
def serve_static_asset(asset_path):
    static_dir = os.path.join(BASE_DIR, 'static')
    full_path = os.path.join(static_dir, asset_path)
    if os.path.isfile(full_path):
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        return send_from_directory(directory, filename)
    return jsonify({'error': 'Not found'}), 404

def load_campaign_properties():
    props_path = os.path.join(BASE_DIR, 'campaign.properties')
    props = {
        'call_delay': '5',
        'agent_voice': 'Polly.Joanna',
        'payer_profile': 'aetna_representative',
        'enable_recording': 'true',
        'enable_transcript': 'true',
        'storage_mode': 'server_delete_twilio',
        'artifact_subfolder': 'default',
        'contacts_file': 'contacts_sheet.xlsx',
        'script_file': 'script.json'
    }
    if os.path.exists(props_path):
        try:
            with open(props_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    if '=' in line:
                        key, value = line.split('=', 1)
                        props[key.strip()] = value.strip().strip('"').strip("'")
        except Exception as exc:
            logger.warning('failed_to_load_campaign_properties path=%s error=%s', props_path, str(exc))
    return props


def load_contacts_from_excel(file_path):
    contacts = []
    if not os.path.exists(file_path):
        return contacts
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        
        # Read headers
        header_row = 1
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=header_row, column=col).value or '').strip()
            headers.append(val)
            
        # Read rows
        for row in range(header_row + 1, sheet.max_row + 1):
            row_data = {}
            has_data = False
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=row, column=col).value
                if val is not None:
                    has_data = True
                header = headers[col - 1]
                if header:
                    row_data[header] = str(val).strip() if val is not None else ""
            if has_data and row_data.get('member_id'):
                contacts.append(row_data)
    except Exception as exc:
        logger.warning('failed_to_load_contacts_from_excel path=%s error=%s', file_path, str(exc))
    return contacts


def load_script_from_file(file_path):
    script_data = []
    if not os.path.exists(file_path):
        return script_data
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            script_data = json.load(f) or []
    except Exception as exc:
        logger.warning('failed_to_load_script_from_file path=%s error=%s', file_path, str(exc))
    return script_data


def extract_claim_details_via_gemini(dialogue_text):
    if not GEMINI_API_KEY:
        logger.warning('gemini_extraction_api_key_missing')
        return None

    prompt = f"""
Analyze the following phone conversation transcript between a health provider voice agent and a payer representative.
Extract the claim auditing details requested in the exact JSON format specified below.
If a field is not mentioned, use null. Remove "$" symbols, commas, or extra text from numbers.

Intelligently reconstruct currency decimals:
- Raw transcribed values often lack decimal points (e.g., "6158" means "61.58", "5572" means "55.72", "$168" means "1.68", "4668" means "46.68", "55927" means "559.27"). Intelligently divide currency fields by 100 where applicable to output correct float values in dollars and cents.
- Standard co-payments (like "45") are whole dollars and should remain whole (e.g., 45.0).
- Extract the EFT Trace number strictly and fully as a raw string of digits. Do NOT truncate, do NOT use scientific notation, and do NOT convert it to a float.
- Long numbers like EFT Trace numbers are sometimes split across dialogue turns (e.g. ending in "88261780100" and the next turn starting with "7 5"). In such cases, intelligently combine them into the single, full EFT trace number string (e.g., "8826178010075").

Dialogue:
{dialogue_text}

JSON Output Schema:
{{
  "audits": [
    {{
      "member_id": "string",
      "claim_status": "Completed" or "Pending" or "Denied" or null,
      "claim_completion_date": "MM/DD/YYYY" or "Month Day, Year" or null,
      "allowable_amount": float or null,
      "total_paid_amount": float or null,
      "co_payment_amount": float or null,
      "co_insurance_amount": float or null,
      "payment_issued_date": "MM/DD/YYYY" or "Month Day, Year" or null,
      "settlement_date": "MM/DD/YYYY" or "Month Day, Year" or null,
      "eft_trace_number": "string" or null
    }}
  ]
}}
"""

    payload = {
        'systemInstruction': {
            'parts': [{
                'text': 'You are a healthcare billing data extraction assistant. Return strictly valid JSON matching the requested schema. Do not output any markdown wrappers, ticks or extra explanations.'
            }]
        },
        'contents': [{
            'parts': [{
                'text': prompt
            }]
        }]
    }

    endpoint = (
        'https://generativelanguage.googleapis.com/v1beta/models/'
        'gemini-2.5-flash:generateContent'
    )
    url = f'{endpoint}?key={urllib.parse.quote(GEMINI_API_KEY)}'
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            body = resp.read().decode('utf-8', errors='ignore')
            parsed = json.loads(body)
            candidates = parsed.get('candidates', [])
            if candidates:
                text = candidates[0].get('content', {}).get('parts', [{}])[0].get('text', '').strip()
                if text.startswith('```'):
                    text = text.split('```json')[-1].split('```')[0].strip()
                extracted = json.loads(text)
                return extracted
    except Exception as exc:
        logger.warning('failed_to_extract_claim_details_via_gemini: %s', str(exc))
    return None


def save_audit_results_json(audits, call_sid=None):
    """
    Primary audit writer: saves audit_results.json inside the call's artifact folder.
    Pure Python — no pandas/openpyxl dependencies, atomic write, crash-safe.
    """
    if not audits:
        return
    try:
        state = read_call_state(call_sid) if call_sid else {}
        storage_options = state.get('storage_options') or build_storage_options({})
        artifact_dir = get_artifact_call_dir(call_sid, storage_options) if call_sid else None
        if not artifact_dir:
            logger.warning('save_audit_results_json_no_artifact_dir call_sid=%s', call_sid)
            return

        output_path = os.path.join(artifact_dir, 'audit_results.json')

        # Merge with existing file if present (multiple calls may flush partial audits)
        existing = []
        if os.path.exists(output_path):
            try:
                with open(output_path, 'r', encoding='utf-8') as fh:
                    existing = json.load(fh)
                    if not isinstance(existing, list):
                        existing = []
            except Exception:
                existing = []

        # Merge: update existing entry for same member_id or append new
        existing_index = {
            str(e.get('member_id') or '').strip().upper(): i
            for i, e in enumerate(existing)
        }
        for audit in audits:
            mid = str(audit.get('member_id') or '').strip().upper()
            if not mid:
                continue
            if mid in existing_index:
                # Update existing entry — prefer non-null incoming values
                idx = existing_index[mid]
                for k, v in audit.items():
                    if v is not None:
                        existing[idx][k] = v
            else:
                existing.append(audit)
                existing_index[mid] = len(existing) - 1

        import uuid as _uuid
        tmp_path = f"{output_path}.tmp.{_uuid.uuid4().hex}"
        with open(tmp_path, 'w', encoding='utf-8') as fh:
            json.dump(existing, fh, indent=2, default=str)
        os.replace(tmp_path, output_path)

        logger.info('AUDIT_JSON_SAVED call_sid=%s path=%s members=%s', call_sid, output_path, len(existing))

        # Update call state artifacts so the UI shows the JSON path
        if call_sid:
            fresh_state = read_call_state(call_sid)
            artifacts = fresh_state.get('artifacts', {}) or {}
            artifacts['audit_results_json_path'] = output_path
            artifacts['audit_json_status'] = f"Audit JSON saved: {os.path.basename(output_path)}"
            update_call_state(call_sid, {'artifacts': artifacts})

    except Exception as exc:
        logger.error('AUDIT_JSON_SAVE_FAILED call_sid=%s error=%s', call_sid, str(exc))


def run_in_process_excel_converter():
    input_dir = os.path.join(BASE_DIR, "input")
    if not os.path.exists(input_dir):
        logger.warning('excel_converter_missing_input_dir path=%s', input_dir)
        return
        
    try:
        excel_files = [f for f in os.listdir(input_dir) if f.endswith((".xlsx", ".xls"))]
        if not excel_files:
            logger.warning('excel_converter_no_excel_files_found path=%s', input_dir)
            return

        # Prioritize files containing 'customer_data' or 'customer'
        target_file = None
        for f in excel_files:
            if "customer" in f.lower():
                target_file = f
                break
        if not target_file:
            target_file = excel_files[0]

        excel_file_path = os.path.join(input_dir, target_file)
        json_file_path = os.path.join(input_dir, "customer_data.json")

        logger.info('excel_converter_in_process_start excel=%s json=%s', excel_file_path, json_file_path)
        from convertExcelToJson import convert_excel_to_json
        convert_excel_to_json(excel_file_path, json_file_path)
        logger.info('excel_converter_in_process_success excel=%s json=%s', excel_file_path, json_file_path)
    except Exception as exc:
        logger.exception('excel_converter_in_process_failed error=%s', str(exc))


@app.route('/api/call', methods=['POST'])
def make_call():
    # 1. Load campaign configurations from properties file first
    props = load_campaign_properties()

    # 2. Check if we should dynamically run the Excel converter
    run_converter = props.get('run_excel_converter', 'true').lower() in {'1', 'true', 'yes', 'on'}
    if run_converter:
        try:
            # Run the excel parser in-process (fully blocking & synchronized!)
            run_in_process_excel_converter()
            
            # Point to input/customer_data.json and reload in memory
            new_customer_data_file = os.path.join(BASE_DIR, 'input', 'customer_data.json')
            if os.path.exists(new_customer_data_file):
                global CUSTOMER_DATA_FILE, CUSTOMER_DATA_INDEX
                CUSTOMER_DATA_FILE = new_customer_data_file
                CUSTOMER_DATA_INDEX = load_customer_data_index()
                logger.info('api_call_customer_data_reloaded path=%s members=%s', CUSTOMER_DATA_FILE, len(CUSTOMER_DATA_INDEX))
        except Exception as exc:
            logger.exception('api_call_excel_conversion_failed error=%s', str(exc))
    
    # 4. Extract campaign settings
    call_delay = int(props.get('call_delay', '5'))
    voice = props.get('agent_voice', 'Polly.Joanna')
    selected_profile_name = props.get('payer_profile', 'aetna_representative')
    enable_recording = props.get('enable_recording', 'true').lower() in {'1', 'true', 'yes', 'on'}
    enable_transcript = props.get('enable_transcript', 'true').lower() in {'1', 'true', 'yes', 'on'}
    
    storage_mode = props.get('storage_mode', 'server_delete_twilio')
    subfolder = props.get('artifact_subfolder', 'default')
    storage_options = {
        'mode': storage_mode,
        'subfolder': subfolder,
        'base_dir': CALL_ARTIFACTS_DIR,
        'save_enabled': storage_mode in {'server_keep_twilio', 'server_delete_twilio'},
        'delete_from_twilio': storage_mode == 'server_delete_twilio',
    }

    # 5. Load Twilio credentials directly from .env variables
    account_sid = TWILIO_ACCOUNT_SID
    auth_token = TWILIO_AUTH_TOKEN
    from_number = TWILIO_PHONE_NUMBER

    selected_profile = PAYER_PROFILES.get(selected_profile_name)
    if not isinstance(selected_profile, dict):
        return jsonify({'error': f'Unknown payer profile "{selected_profile_name}".', 'code': 'PAYER_PROFILE_UNKNOWN', 'allowed_profiles': sorted(list(PAYER_PROFILES.keys()))}), 400
    to_number = str(selected_profile.get('phone_number') or '').strip()

    # 6. Load contacts workbook and answers script from the input/ folder
    contacts_path = os.path.join(BASE_DIR, 'input', props.get('contacts_file', 'contacts_sheet.xlsx'))
    contacts = load_contacts_from_excel(contacts_path)
    if not contacts:
        return jsonify({'error': f'No contacts found in {contacts_path}. Please place your Excel/CSV workbook in the input folder.', 'code': 'CONTACTS_FILE_MISSING'}), 400

    script_path = os.path.join(BASE_DIR, 'input', props.get('script_file', 'script.json'))
    all_answers_bank = load_script_from_file(script_path)

    # 7. Formulate campaign structures
    contact_context = resolve_contact_context(contacts[0])
    if selected_profile_name:
        contact_context = dict(contact_context)
        contact_context['ivr_profile'] = selected_profile_name
        contact_context['profile_name'] = selected_profile_name

    campaign_contact_queue = contacts[1:]
    active_member_id = normalize_member_id(contact_context.get('member_id') or extract_member_id_from_context(contact_context))
    uploaded_answer_bank = filter_uploaded_answer_bank_for_member(all_answers_bank, active_member_id)

    call_reason = str(contact_context.get('call_reason') or '').strip()
    ivr_navigation_config = build_ivr_navigation_config({'enabled': True}, contact_context, call_reason)

    remaining_contacts_after_current = len(campaign_contact_queue)
    campaign_total_contacts = len(contacts)

    logger.info('api_call_request to_number=%s profile=%s remaining_contacts_after_current=%s', to_number, selected_profile_name, remaining_contacts_after_current)

    if not to_number:
        return jsonify({'error': 'Phone number is required for selected profile.', 'code': 'PAYER_PHONE_MISSING'}), 400

    creds_error = validate_twilio_credentials(account_sid, auth_token, from_number)
    if creds_error:
        return jsonify({'error': creds_error, 'code': 'TWILIO_CREDENTIALS_INVALID'}), 400

    if ivr_navigation_config.get('enabled') and ivr_navigation_config.get('ivr_profile_error'):
        logger.warning('api_call_ivr_profile_error to_number=%s error=%s allowed_profiles=%s', to_number, ivr_navigation_config.get('ivr_profile_error'), ','.join(ivr_navigation_config.get('allowed_profiles') or []))
        return jsonify({
            'error': ivr_navigation_config.get('ivr_profile_error'),
            'code': 'IVR_PROFILE_INVALID',
            'allowed_profiles': ivr_navigation_config.get('allowed_profiles') or [],
        }), 400

    try:
        twilio_client = telephony_provider.create_client(account_sid, auth_token)
    except Exception as exc:
        return jsonify({'error': f'Unable to initialize Twilio client: {str(exc)}', 'code': 'TWILIO_CLIENT_INIT_FAILED'}), 400

    logger.info('api_call_credentials sid=%s from=%s token=%s', mask_secret(account_sid), from_number, mask_secret(auth_token))
    logger.info(
        'api_call_ivr_config enabled=%s profile=%s source=%s customer_care=%s lang=%s dept=%s submenu=%s rep=%s static_digits=%s max_attempts=%s matched_category=%s matched_source=%s',
        ivr_navigation_config.get('enabled'),
        ivr_navigation_config.get('profile_name', ''),
        ivr_navigation_config.get('profile_source', ''),
        ivr_navigation_config.get('customer_care_number', ''),
        ','.join(ivr_navigation_config.get('language_keywords') or []),
        ','.join(ivr_navigation_config.get('department_keywords') or []),
        ','.join(ivr_navigation_config.get('submenu_keywords') or []),
        ','.join(ivr_navigation_config.get('representative_keywords') or []),
        ','.join([f"{phase}:{(ivr_navigation_config.get('static_route') or {}).get(phase, {}).get('digit', '')}" for phase in IVR_STATIC_PHASES if (ivr_navigation_config.get('static_route') or {}).get(phase)]),
        ivr_navigation_config.get('max_attempts'),
        ivr_navigation_config.get('matched_category', ''),
        ivr_navigation_config.get('matched_category_source', ''),
    )

    callback_url = get_public_base_url()
    logger.info('api_call_callback_base_url=%s', callback_url)

    try:
        twiml_instructions = f'''
        <Response>
        <Gather input="speech dtmf" action="{callback_url}/api/webhook/speech" method="POST" timeout="8" speechTimeout="3">
                <Pause length="20"/>
            </Gather>
            <Redirect method="POST">{callback_url}/api/webhook/speech</Redirect>
        </Response>
        '''
        initial_status = 'customer_care_listening'

        call_params = {
            'twiml': twiml_instructions,
            'to': to_number,
            'from_': from_number,
            'status_callback': f"{callback_url}/api/webhook/call_status",
            'status_callback_method': 'POST',
            'status_callback_event': ['initiated', 'ringing', 'answered', 'completed'],
        }
        if enable_recording:
            call_params.update({
                'record': True,
                'recording_status_callback': f"{callback_url}/api/webhook/recording_status",
                'recording_status_callback_method': 'POST',
                'recording_status_callback_event': ['in-progress', 'completed', 'absent'],
            })

        call = telephony_provider.create_call(twilio_client, call_params)
        call_twilio_credentials[call.sid] = {
            'account_sid': account_sid,
            'auth_token': auth_token,
            'from_number': from_number,
        }
        
        # Initialize state for this call
        write_call_state(call.sid, {
            'status': initial_status,
            'last_speech': '',
            'interaction_transcript': [],
            'contact_context': contact_context,
            'active_member_id': active_member_id,
            'uploaded_answer_bank': uploaded_answer_bank,
            'uploaded_answer_bank_all': all_answers_bank,
            'campaign_contact_queue': campaign_contact_queue,
            'voice': voice,
            'selected_profile_name': selected_profile_name,
            'run_excel_converter': run_converter,
            'campaign_total_contacts': campaign_total_contacts,
            'campaign_remaining_contacts_after_current': remaining_contacts_after_current,
            'campaign_has_next_member': bool(remaining_contacts_after_current > 0),
            'ivr_navigation': ivr_navigation_config,
            'ivr_department_selected': False,
            'ivr_submenu_selected': False,
            'ivr_welcome_detected': False,
            'ivr_language_selected': False,
            'ivr_phase': 'customer_care_intro',
            'ivr_attempts': 0,
            'claim_status_value': '',
            'contacts_sheet_updated_path': '',
            'fallback_reply_count': 0,
            'last_fallback_reply': '',
            'awaiting_rejection_reason': False,
            '_twilio_account_sid': account_sid,
            '_twilio_auth_token': auth_token,
            'storage_options': storage_options,
            'artifacts': {
                'base_dir': CALL_ARTIFACTS_DIR,
                'subfolder': storage_options.get('subfolder'),
                'mode': storage_options.get('mode'),
                'recording_saved_path': '',
                'recording_saved': False,
                'twilio_deleted': False,
                'transcription_saved_path': '',
                'interaction_saved_path': '',
                'last_error': '',
            },
            'recording': {
                'enabled': enable_recording,
                'status': 'not_started' if enable_recording else 'disabled',
                'recording_sid': '',
                'recording_url': '',
                'duration': '',
            },
            'transcription': {
                'enabled': enable_transcript,
                'status': 'not_requested' if enable_transcript else 'disabled',
                'source': 'twilio_native',
                'transcription_sid': '',
                'text': '',
                'error': '',
            }
        })
        logger.info('api_call_success call_sid=%s', call.sid)

        return jsonify({
            'message': 'Call initiated successfully',
            'call_sid': call.sid,
            'selected_profile_name': selected_profile_name,
            'recording_enabled': enable_recording,
            'transcript_enabled': enable_transcript,
            'storage_mode': storage_options.get('mode'),
            'artifact_subfolder': storage_options.get('subfolder'),
            'uploaded_answer_count': len(uploaded_answer_bank),
            'active_member_id': active_member_id,
            'target_number': to_number,
            'ivr_navigation_enabled': ivr_navigation_config['enabled'],
            'ivr_profile_name': ivr_navigation_config.get('profile_name', ''),
            'ivr_profile_source': ivr_navigation_config.get('profile_source', ''),
            'ivr_matched_category': ivr_navigation_config.get('matched_category', ''),
            'ivr_matched_category_source': ivr_navigation_config.get('matched_category_source', ''),
            'run_excel_converter': run_converter,
        }), 200
    except Exception as e:
        logger.exception('api_call_error error=%s', str(e))
        return jsonify({'error': str(e)}), 500


def extract_claim_details_via_python_script(interaction):
    if not isinstance(interaction, list):
        return None
    import re
    import pandas as pd
    
    contacts_path = os.path.join(BASE_DIR, 'input', 'contacts_sheet.xlsx')
    all_mids = []
    if os.path.exists(contacts_path):
        try:
            df = pd.read_excel(contacts_path)
            for col in df.columns:
                clean_col = str(col).strip().lower().replace('_', '').replace(' ', '')
                if clean_col in {'memberid', 'patientid'}:
                    all_mids = [str(x).strip() for x in df[col].tolist() if str(x).strip()]
                    break
        except Exception:
            pass

    def match_member_id_from_text(text):
        cleaned = re.sub(r'[^A-Za-z0-9]', '', str(text or '')).upper()
        for mid in all_mids:
            mid_clean = re.sub(r'[^A-Za-z0-9]', '', str(mid)).upper()
            if mid_clean and mid_clean in cleaned:
                return mid
        match_w = re.search(r'W\s*(?:[0-9]\s*){9}', str(text or ''), re.I)
        if match_w:
            return re.sub(r'\s+', '', match_w.group(0)).upper()
        return None

    member_dialogues = {}
    current_member_id = None
    
    for entry in interaction:
        # Support both state format (speaker/text) and transcript file format (representative/calling_agent)
        speaker = str(entry.get('speaker') or '').strip().lower()
        text = str(entry.get('text') or '').strip()
        rep_text = str(entry.get('representative') or '').strip()
        agent_text = str(entry.get('calling_agent') or '').strip()
        
        # Normalise to rep_text / agent_text for unified processing
        if not rep_text and not agent_text:
            if speaker == 'representative':
                rep_text = text
            else:
                agent_text = text
        combined_text = rep_text or agent_text
        
        if 'Switched context to next member' in agent_text or 'Switched context to next member' in text:
            match = re.search(r'Switched context to next member\s*(\w+)', combined_text or text, re.I)
            if match:
                current_member_id = match.group(1).strip()
        
        if not current_member_id or 'said:' in agent_text or 'said:' in text or 'W' in agent_text or 'W' in text:
            spoken_mid = match_member_id_from_text(agent_text or text)
            if spoken_mid:
                current_member_id = spoken_mid
                
        if not current_member_id and all_mids:
            current_member_id = all_mids[0]
            
        if current_member_id:
            if current_member_id not in member_dialogues:
                member_dialogues[current_member_id] = []
            if rep_text:
                member_dialogues[current_member_id].append(rep_text)
            elif speaker == 'representative' and text:
                member_dialogues[current_member_id].append(text)
            if agent_text:
                member_dialogues[current_member_id].append(agent_text)
            elif speaker != 'representative' and text:
                member_dialogues[current_member_id].append(text)
            if agent_text:
                member_dialogues[current_member_id].append(agent_text)
                
    if not member_dialogues:
        return None
        
    audits = []
    for m_id, lines in member_dialogues.items():
        full_text = ' '.join(lines)
        full_text_clean = re.sub(r'\s+', ' ', full_text)
        
        claim_status = None
        if re.search(r'completed|paid|processed', full_text_clean, re.I):
            claim_status = 'Completed'
        elif re.search(r'rejected|denied|rejection', full_text_clean, re.I):
            claim_status = 'Denied'
        elif re.search(r'pending|in process|review', full_text_clean, re.I):
            claim_status = 'Pending'
            
        comp_date_match = re.search(r'completed on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})', full_text_clean, re.I)
        claim_comp_date = comp_date_match.group(1).strip() if comp_date_match else None
        
        allow_match = re.search(r'allowable amount is\s*(?:\$)?(\d+(?:\.\d{2})?)', full_text_clean, re.I)
        allowable_amt = float(allow_match.group(1)) if allow_match else None
        
        paid_match = re.search(r'paid to the provider[^0-9]*(?:\$)?(\d+(?:\.\d{2})?)', full_text_clean, re.I)
        paid_amt = float(paid_match.group(1)) if paid_match else None
        
        copay_match = re.search(r'co-payment amount\s*(?:\$)?(\d+(?:\.\d{2})?)', full_text_clean, re.I)
        copay_amt = float(copay_match.group(1)) if copay_match else None
        
        coins_match = re.search(r'co-insurance amount\s*(?:\$)?(\d+(?:\.\d{2})?)', full_text_clean, re.I)
        coins_amt = float(coins_match.group(1)) if coins_match else None
        
        issued_match = re.search(r'issued electronically on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})', full_text_clean, re.I)
        issued_date = issued_match.group(1).strip() if issued_match else None
        
        settled_match = re.search(r'settled on\s*([A-Za-z]+\s*\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})', full_text_clean, re.I)
        settled_date = settled_match.group(1).strip() if settled_match else None
        
        eft_match = re.search(r'EFT[\s.]*Trace number is\s*([0-9\s,\.]+)', full_text_clean, re.I)
        eft_num = re.sub(r'[^0-9]', '', eft_match.group(1)).strip() if eft_match else None

        clm_match = re.search(r'claim number is[,\s]*([A-Za-z0-9][A-Za-z0-9\s,\.]+?)(?:\s+say|\.\.\.|   |$)', full_text_clean, re.I)
        claim_num = re.sub(r'[^A-Za-z0-9]', '', clm_match.group(1)).upper() if clm_match else None
        
        m_id_clean = str(m_id).strip()
        
        if claim_status or allowable_amt or paid_amt or eft_num:
            audits.append({
                "member_id": m_id_clean,
                "claim_status": claim_status,
                "claim_number": claim_num,
                "claim_completion_date": claim_comp_date,
                "allowable_amount": allowable_amt,
                "total_paid_amount": paid_amt,
                "co_payment_amount": copay_amt,
                "co_insurance_amount": coins_amt,
                "payment_issued_date": issued_date,
                "settlement_date": settled_date,
                "eft_trace_number": eft_num
            })
    return {"audits": audits} if audits else None


def process_offline_claim_extraction(call_sid):
    """
    Safety-net audit writer that runs at end-of-call.
    Covers the last patient in the campaign (no context switch fires for them).
    All previous patients should already have audit_results.json written
    at context-switch time. This run merges any remaining patients.
    """
    try:
        state = read_call_state(call_sid)
        interaction = state.get('interaction_transcript', [])
        storage_options = state.get('storage_options') or build_storage_options({})
        artifacts = state.get('artifacts', {}) or {}

        if not interaction:
            return

        # 1. Save interaction transcript to artifact folder
        if storage_options.get('save_enabled'):
            try:
                artifact_dir = get_artifact_call_dir(call_sid, storage_options)
                interaction_path = os.path.join(artifact_dir, 'interaction-transcript.json')
                save_json_file(interaction_path, normalize_interaction_transcript(interaction))
                artifacts['interaction_saved_path'] = interaction_path
                update_call_state(call_sid, {'artifacts': artifacts})
            except Exception as artifact_exc:
                logger.warning('failed_to_save_interaction_artifact call_sid=%s error=%s', call_sid, str(artifact_exc))

        # 2. Extract audits via Python script and write to audit_results.json
        logger.info('triggering_offline_audit_json_update call_sid=%s', call_sid)
        extracted_audits = extract_claim_details_via_python_script(interaction)
        if extracted_audits and isinstance(extracted_audits, dict):
            audits_list = extracted_audits.get('audits') or []
            if audits_list:
                save_audit_results_json(audits_list, call_sid=call_sid)
                logger.info('offline_audit_json_saved call_sid=%s patients=%s', call_sid, len(audits_list))
        else:
            logger.info('offline_audit_no_data_extracted call_sid=%s', call_sid)

    except Exception as exc:
        logger.warning('failed_to_process_offline_audit call_sid=%s error=%s', call_sid, str(exc))


@app.route('/api/webhook/call_status', methods=['POST'])
def handle_call_status():
    call_sid = request.form.get('CallSid')
    call_status = (request.form.get('CallStatus') or '').lower()
    logger.info('api_call_status_webhook call_sid=%s twilio_status=%s', call_sid, call_status)
    if not call_sid:
        return ('', 204)

    state = read_call_state(call_sid)
    recording_enabled = bool((state.get('recording') or {}).get('enabled'))

    updates = {'twilio_call_status': call_status}
    if call_status in {'completed', 'busy', 'failed', 'no-answer', 'canceled'}:
        updates['status'] = 'call_ended'
        call_twilio_credentials.pop(call_sid, None)
        
        # Trigger offline audit extractor in a background thread so we don't block Twilio's webhook response!
        import threading
        thread = threading.Thread(target=process_offline_claim_extraction, args=(call_sid,))
        thread.daemon = True
        thread.start()

    update_call_state(call_sid, updates)
    if call_status in {'completed', 'busy', 'failed', 'no-answer', 'canceled'} and not recording_enabled:
        clear_internal_call_secrets(call_sid)
    return ('', 204)


@app.route('/api/webhook/recording_status', methods=['POST'])
def handle_recording_status():
    call_sid = request.form.get('CallSid')
    recording_sid = request.form.get('RecordingSid', '')
    recording_status = (request.form.get('RecordingStatus') or '').lower()
    recording_url = request.form.get('RecordingUrl', '')
    recording_duration = request.form.get('RecordingDuration', '')
    logger.info('api_recording_status_webhook call_sid=%s recording_sid=%s status=%s', call_sid, recording_sid, recording_status)

    state = read_call_state(call_sid)
    recording_state = state.get('recording', {})
    recording_state.update({
        'enabled': True,
        'status': recording_status or recording_state.get('status', 'unknown'),
        'recording_sid': recording_sid or recording_state.get('recording_sid', ''),
        'recording_url': recording_url or recording_state.get('recording_url', ''),
        'duration': recording_duration or recording_state.get('duration', ''),
    })
    update_call_state(call_sid, {'recording': recording_state})

    state = read_call_state(call_sid)
    storage_options = state.get('storage_options') or build_storage_options({})
    artifacts = state.get('artifacts', {})
    logger.info('artifact_recording_status call_sid=%s recording_sid=%s recording_status=%s save_enabled=%s delete_from_twilio=%s subfolder=%s', call_sid, recording_sid, recording_status, storage_options.get('save_enabled'), storage_options.get('delete_from_twilio'), storage_options.get('subfolder'))
    if recording_status == 'completed' and storage_options.get('save_enabled'):
        try:
            saved_path, save_error = save_recording_media_to_local(call_sid, recording_sid, storage_options)
            if save_error:
                artifacts['last_error'] = save_error
                logger.warning('artifact_recording_save_failed call_sid=%s recording_sid=%s error=%s', call_sid, recording_sid, save_error)
            else:
                artifacts['recording_saved'] = True
                artifacts['recording_saved_path'] = saved_path
                artifacts['last_error'] = ''
                logger.info('artifact_recording_save_completed call_sid=%s recording_sid=%s path=%s', call_sid, recording_sid, saved_path)
                if storage_options.get('delete_from_twilio'):
                    twilio_client, _ = get_twilio_context_for_call(call_sid)
                    telephony_provider.delete_recording(twilio_client, recording_sid)
                    artifacts['twilio_deleted'] = True
                    logger.info('artifact_twilio_recording_deleted call_sid=%s recording_sid=%s', call_sid, recording_sid)
        except Exception as exc:
            artifacts['last_error'] = str(exc)
            logger.exception('artifact_recording_save_exception call_sid=%s recording_sid=%s error=%s', call_sid, recording_sid, str(exc))
    update_call_state(call_sid, {'artifacts': artifacts})

    if recording_status in {'completed', 'absent'}:
        clear_internal_call_secrets(call_sid)

    transcription_enabled = bool((state.get('transcription') or {}).get('enabled'))
    if recording_status == 'completed' and transcription_enabled:
        request_recording_transcription(call_sid, recording_sid)
    return ('', 204)


@app.route('/api/webhook/transcription_status', methods=['POST'])
def handle_transcription_status():
    call_sid = request.form.get('CallSid')
    recording_sid = request.form.get('RecordingSid', '')
    transcription_sid = request.form.get('TranscriptionSid', '')
    transcription_status = (request.form.get('TranscriptionStatus') or '').lower()
    transcription_text = request.form.get('TranscriptionText', '')
    logger.info('api_transcription_status_webhook call_sid=%s recording_sid=%s transcription_sid=%s status=%s', call_sid, recording_sid, transcription_sid, transcription_status)

    transcription_state = {
        'enabled': True,
        'status': transcription_status or 'completed',
        'source': 'twilio_native',
        'transcription_sid': transcription_sid,
        'recording_sid': recording_sid,
        'text': transcription_text,
        'error': '',
    }
    state = read_call_state(call_sid)
    artifacts = state.get('artifacts', {})
    storage_options = state.get('storage_options') or build_storage_options({})
    logger.info('artifact_transcription_status call_sid=%s recording_sid=%s transcription_sid=%s status=%s save_enabled=%s', call_sid, recording_sid, transcription_sid, transcription_status, storage_options.get('save_enabled'))
    if storage_options.get('save_enabled') and transcription_text:
        try:
            artifact_dir = get_artifact_call_dir(call_sid, storage_options)
            transcript_path = os.path.join(artifact_dir, f"transcript-{recording_sid or transcription_sid or 'latest'}.txt")
            save_text_file(transcript_path, transcription_text)
            artifacts['transcription_saved_path'] = transcript_path
            artifacts['last_error'] = ''
            logger.info('artifact_transcript_saved call_sid=%s path=%s', call_sid, transcript_path)
        except Exception as exc:
            artifacts['last_error'] = str(exc)
            logger.exception('artifact_transcript_save_exception call_sid=%s error=%s', call_sid, str(exc))

    interaction = state.get('interaction_transcript', [])
    if storage_options.get('save_enabled') and interaction:
        try:
            artifact_dir = get_artifact_call_dir(call_sid, storage_options)
            interaction_path = os.path.join(artifact_dir, 'interaction-transcript.json')
            save_json_file(interaction_path, normalize_interaction_transcript(interaction))
            artifacts['interaction_saved_path'] = interaction_path
            logger.info('artifact_interaction_saved call_sid=%s path=%s entries=%s', call_sid, interaction_path, len(interaction))
            
            try:
                logger.info('triggering_excel_audit_update call_sid=%s', call_sid)
                dialogue_lines = []
                for entry in interaction:
                    speaker = str(entry.get('speaker') or 'unknown')
                    text = str(entry.get('text') or '').strip()
                    dialogue_lines.append(f"{speaker}: {text}")
                dialogue_text = "\n".join(dialogue_lines)
                
                extracted_audits = extract_claim_details_via_gemini(dialogue_text)
                if extracted_audits and isinstance(extracted_audits, dict):
                    audits_list = extracted_audits.get('audits') if isinstance(extracted_audits.get('audits'), list) else [extracted_audits]
                    update_contacts_excel_with_audits(audits_list, call_sid=call_sid)
            except Exception as excel_exc:
                logger.warning('failed_to_process_excel_audit_in_background call_sid=%s error=%s', call_sid, str(excel_exc))
        except Exception as exc:
            artifacts['last_error'] = str(exc)
            logger.exception('artifact_interaction_save_exception call_sid=%s error=%s', call_sid, str(exc))

    update_call_state(call_sid, {'transcription': transcription_state, 'artifacts': artifacts})
    return ('', 204)


@app.route('/api/debug/callback-url', methods=['GET'])
def debug_callback_url():
    return jsonify({
        'callback_base_url': get_public_base_url(),
        'public_base_url': get_public_base_url(),
        'version': APP_VERSION,
    }), 200


@app.route('/api/debug/twilio-config', methods=['GET'])
def debug_twilio_config():
    return jsonify({
        'account_sid_masked': mask_secret(TWILIO_ACCOUNT_SID),
        'phone_number': TWILIO_PHONE_NUMBER,
        'auth_token_set': bool(TWILIO_AUTH_TOKEN),
        'auth_token_length': len(TWILIO_AUTH_TOKEN or ''),
        'webhook_signature_validation': WEBHOOK_SIGNATURE_VALIDATION,
    }), 200


@app.route('/api/payer-profiles', methods=['GET'])
def get_payer_profiles():
    profiles = []
    for profile_name in sorted(list(PAYER_PROFILES.keys())):
        row = PAYER_PROFILES.get(profile_name) or {}
        phone = str(row.get('phone_number') or '').strip()
        profiles.append({
            'profile_name': profile_name,
            'display_name': str(row.get('display_name') or profile_name.replace('_', ' ').title()).strip(),
            'phone_number': phone,
            'configured': bool(phone),
        })
    return jsonify({
        'payer_profile_file': PAYER_PROFILE_FILE,
        'profiles': profiles,
    }), 200


@app.route('/api/artifact-folders', methods=['GET'])
def get_artifact_folders():
    folders = list_artifact_subfolders()
    return jsonify({
        'base_dir': CALL_ARTIFACTS_DIR,
        'folders': folders,
    }), 200


@app.route('/api/webhook/ivr_menu', methods=['POST'])
def handle_ivr_menu():
    call_sid = request.form.get('CallSid')
    speech_result = (request.form.get('SpeechResult') or '').strip()
    logger.info('api_ivr_menu_webhook call_sid=%s speech=%s', call_sid, speech_result)

    state = read_call_state(call_sid)
    voice = (state.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE
    ivr_config = normalize_ivr_navigation_config(state.get('ivr_navigation'))

    if not ivr_config.get('enabled'):
        update_call_state(call_sid, {'status': 'waiting_for_keypress', 'ivr_attempts': 0})
        callback_url = get_public_base_url()
        twiml_response = f'''
        <Response>
            <Gather numDigits="1" action="{callback_url}/api/webhook/keypress" method="POST" timeout="60">
                <Say voice="{DEFAULT_VOICE}">Please press 9 to begin the automated script.</Say>
            </Gather>
            <Pause length="3600"/>
        </Response>
        '''
        return Response(twiml_response, mimetype='text/xml')

    department_keywords = ivr_config['department_keywords']
    submenu_keywords = ivr_config['submenu_keywords']
    representative_keywords = ivr_config['representative_keywords']
    language_keywords = ivr_config['language_keywords']
    expected_welcome = ivr_config.get('welcome_text') or ''
    max_attempts = ivr_config['max_attempts']
    department_selected = bool(state.get('ivr_department_selected'))
    submenu_selected = bool(state.get('ivr_submenu_selected'))
    welcome_detected = bool(state.get('ivr_welcome_detected') or ivr_config.get('welcome_detected'))
    language_selected = bool(state.get('ivr_language_selected') or ivr_config.get('language_selected'))
    attempts = int(state.get('ivr_attempts') or 0) + 1

    if speech_result:
        append_interaction_transcript(call_sid, 'customer_care_ivr', speech_result, source='ivr_menu_prompt')

    if speech_result:
        update_call_state(call_sid, {'ivr_last_heard_text': speech_result})

    static_route = ivr_config.get('static_route') if isinstance(ivr_config.get('static_route'), dict) else {}
    if static_route:
        language_step = static_route.get('language')
        if not language_selected and should_trigger_static_route_step(language_step, speech_result):
            route_digit = str(language_step.get('digit') or '').strip()
            logger.info('api_ivr_static_route_language_selected call_sid=%s digit=%s', call_sid, route_digit)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {route_digit} from static IVR route (language).', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'ivr_navigating',
                'ivr_language_selected': True,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': 'static_route_language',
                'ivr_last_pressed_digit': route_digit,
                'ivr_phase': 'ivr_department',
            })
            return Response(build_ivr_navigation_twiml(play_digits=route_digit), mimetype='text/xml')

        department_step = static_route.get('department')
        if not department_selected and should_trigger_static_route_step(department_step, speech_result):
            route_digit = str(department_step.get('digit') or '').strip()
            logger.info('api_ivr_static_route_department_selected call_sid=%s digit=%s', call_sid, route_digit)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {route_digit} from static IVR route (department).', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'ivr_navigating',
                'ivr_department_selected': True,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': 'static_route_department',
                'ivr_last_pressed_digit': route_digit,
                'ivr_phase': 'ivr_submenu' if (static_route.get('submenu') or submenu_keywords) else 'ivr_representative',
            })
            return Response(build_ivr_navigation_twiml(play_digits=route_digit), mimetype='text/xml')

        submenu_step = static_route.get('submenu')
        if department_selected and not submenu_selected and should_trigger_static_route_step(submenu_step, speech_result):
            route_digit = str(submenu_step.get('digit') or '').strip()
            logger.info('api_ivr_static_route_submenu_selected call_sid=%s digit=%s', call_sid, route_digit)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {route_digit} from static IVR route (submenu).', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'ivr_navigating',
                'ivr_submenu_selected': True,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': 'static_route_submenu',
                'ivr_last_pressed_digit': route_digit,
                'ivr_phase': 'ivr_representative',
            })
            return Response(build_ivr_navigation_twiml(play_digits=route_digit), mimetype='text/xml')

        rep_step = static_route.get('representative')
        if should_trigger_static_route_step(rep_step, speech_result):
            route_digit = str(rep_step.get('digit') or '').strip()
            logger.info('api_ivr_static_route_representative_selected call_sid=%s digit=%s', call_sid, route_digit)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {route_digit} from static IVR route (representative).', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'customer_care_listening',
                'last_speech': '',
                'detected_intent': '',
                'next_action': 'customer_care_qna',
                'menu_digit': route_digit,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': 'static_route_representative',
                'ivr_last_pressed_digit': route_digit,
                'ivr_phase': 'customer_care_intro',
                'voice': voice,
            })
            intro = build_customer_care_intro(state.get('contact_context') or {})
            append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
            twiml_response = build_customer_care_twiml(intro, voice, followup=False, play_digits=route_digit)
            return Response(twiml_response, mimetype='text/xml')

    digit = ''
    matched_keyword = ''
    matched_category = ''

    if not welcome_detected and expected_welcome:
        if is_common_hold_message(speech_result):
            welcome_detected = True
            update_call_state(call_sid, {'ivr_welcome_detected': True, 'ivr_phase': 'ivr_language'})
        if is_welcome_prompt_detected(expected_welcome, speech_result):
            welcome_detected = True
            update_call_state(call_sid, {'ivr_welcome_detected': True, 'ivr_phase': 'ivr_language'})
        else:
            if attempts >= max_attempts:
                logger.warning('api_ivr_menu_welcome_not_detected_max_attempts call_sid=%s attempts=%s', call_sid, attempts)
                update_call_state(call_sid, {'status': 'customer_care_listening', 'ivr_attempts': attempts, 'ivr_phase': 'customer_care_intro'})
                intro = build_customer_care_intro(state.get('contact_context') or {})
                append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
                return Response(build_customer_care_twiml(intro, voice, followup=False), mimetype='text/xml')
            update_call_state(call_sid, {'status': 'ivr_navigating', 'ivr_attempts': attempts, 'ivr_phase': 'ivr_welcome'})
            return Response(build_ivr_navigation_twiml(), mimetype='text/xml')

    if not language_selected and not language_keywords:
        language_selected = True
        update_call_state(call_sid, {'ivr_language_selected': True, 'ivr_phase': 'ivr_department'})

    if not language_selected:
        preferred_language = str(ivr_config.get('language_preferred') or '').strip().lower()
        fallback_language_digit = ''
        if preferred_language and preferred_language in _normalize_text(speech_result):
            fallback_language_digit = extract_first_press_digit(speech_result)
            if fallback_language_digit:
                logger.info('api_ivr_menu_language_selected_by_fallback call_sid=%s language=%s digit=%s speech=%s', call_sid, preferred_language, fallback_language_digit, speech_result)
                append_interaction_transcript(call_sid, 'agent', f'Pressed {fallback_language_digit} for language fallback ({preferred_language}).', source='ivr_navigation')
                update_call_state(call_sid, {
                    'status': 'ivr_navigating',
                    'ivr_language_selected': True,
                    'ivr_attempts': 0,
                    'ivr_last_matched_keyword': preferred_language,
                    'ivr_last_pressed_digit': fallback_language_digit,
                    'ivr_phase': 'ivr_department',
                })
                return Response(build_ivr_navigation_twiml(play_digits=fallback_language_digit), mimetype='text/xml')

        rep_digit, rep_keyword = find_digit_for_keywords(speech_result, representative_keywords)
        if rep_digit:
            logger.info('api_ivr_menu_representative_reached_from_language_phase call_sid=%s digit=%s keyword=%s', call_sid, rep_digit, rep_keyword)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {rep_digit} to reach a representative (matched "{rep_keyword}").', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'customer_care_listening',
                'last_speech': '',
                'detected_intent': '',
                'next_action': 'customer_care_qna',
                'menu_digit': rep_digit,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': rep_keyword,
                'ivr_last_pressed_digit': rep_digit,
                'ivr_phase': 'customer_care_intro',
                'voice': voice,
            })
            intro = build_customer_care_intro(state.get('contact_context') or {})
            append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
            twiml_response = build_customer_care_twiml(intro, voice, followup=False, play_digits=rep_digit)
            return Response(twiml_response, mimetype='text/xml')

        digit, matched_keyword = find_digit_for_keywords(speech_result, language_keywords)
        if digit:
            logger.info('api_ivr_menu_language_selected call_sid=%s digit=%s keyword=%s', call_sid, digit, matched_keyword)
            append_interaction_transcript(call_sid, 'agent', f'Pressed {digit} for language (matched "{matched_keyword}").', source='ivr_navigation')
            update_call_state(call_sid, {
                'status': 'ivr_navigating',
                'ivr_language_selected': True,
                'ivr_attempts': 0,
                'ivr_last_matched_keyword': matched_keyword,
                'ivr_last_pressed_digit': digit,
                'ivr_phase': 'ivr_department',
            })
            return Response(build_ivr_navigation_twiml(play_digits=digit), mimetype='text/xml')
        if attempts >= max_attempts:
            logger.warning('api_ivr_menu_language_not_detected_max_attempts call_sid=%s attempts=%s', call_sid, attempts)
            update_call_state(call_sid, {'ivr_language_selected': True, 'ivr_attempts': 0, 'ivr_phase': 'ivr_department'})
        else:
            update_call_state(call_sid, {'status': 'ivr_navigating', 'ivr_attempts': attempts, 'ivr_phase': 'ivr_language'})
            return Response(build_ivr_navigation_twiml(), mimetype='text/xml')

    if not department_selected:
        digit, matched_keyword = find_digit_for_keywords(speech_result, department_keywords)
        matched_category = 'department' if digit else ''
        if not digit:
            digit, matched_keyword = find_digit_for_keywords(speech_result, representative_keywords)
            matched_category = 'representative' if digit else ''
    elif not submenu_selected:
        digit, matched_keyword = find_digit_for_keywords(speech_result, representative_keywords)
        matched_category = 'representative' if digit else ''
        if not digit and submenu_keywords:
            digit, matched_keyword = find_digit_for_keywords(speech_result, submenu_keywords)
            matched_category = 'submenu' if digit else ''
    else:
        digit, matched_keyword = find_digit_for_keywords(speech_result, representative_keywords)
        matched_category = 'representative' if digit else ''

    if digit and matched_category == 'representative':
        logger.info('api_ivr_menu_representative_reached call_sid=%s digit=%s keyword=%s', call_sid, digit, matched_keyword)
        append_interaction_transcript(call_sid, 'agent', f'Pressed {digit} to reach a representative (matched "{matched_keyword}").', source='ivr_navigation')
        update_call_state(call_sid, {
            'status': 'customer_care_listening',
            'last_speech': '',
            'detected_intent': '',
            'next_action': 'customer_care_qna',
            'menu_digit': digit,
            'ivr_attempts': 0,
            'ivr_last_matched_keyword': matched_keyword,
            'ivr_last_pressed_digit': digit,
            'ivr_phase': 'customer_care_intro',
            'voice': voice,
        })
        intro = build_customer_care_intro(state.get('contact_context') or {})
        append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
        twiml_response = build_customer_care_twiml(intro, voice, followup=False, play_digits=digit)
        return Response(twiml_response, mimetype='text/xml')

    if digit and matched_category == 'department':
        logger.info('api_ivr_menu_department_selected call_sid=%s digit=%s keyword=%s', call_sid, digit, matched_keyword)
        append_interaction_transcript(call_sid, 'agent', f'Pressed {digit} for department (matched "{matched_keyword}").', source='ivr_navigation')
        update_call_state(call_sid, {
            'status': 'ivr_navigating',
            'ivr_department_selected': True,
            'ivr_attempts': 0,
            'ivr_last_matched_keyword': matched_keyword,
            'ivr_last_pressed_digit': digit,
            'ivr_phase': 'ivr_submenu' if submenu_keywords else 'ivr_representative',
        })
        return Response(build_ivr_navigation_twiml(play_digits=digit), mimetype='text/xml')

    if digit and matched_category == 'submenu':
        logger.info('api_ivr_menu_submenu_selected call_sid=%s digit=%s keyword=%s', call_sid, digit, matched_keyword)
        append_interaction_transcript(call_sid, 'agent', f'Pressed {digit} for submenu (matched "{matched_keyword}").', source='ivr_navigation')
        update_call_state(call_sid, {
            'status': 'ivr_navigating',
            'ivr_submenu_selected': True,
            'ivr_attempts': 0,
            'ivr_last_matched_keyword': matched_keyword,
            'ivr_last_pressed_digit': digit,
            'ivr_phase': 'ivr_representative',
        })
        return Response(build_ivr_navigation_twiml(play_digits=digit), mimetype='text/xml')

    if attempts >= max_attempts:
        logger.warning('api_ivr_menu_max_attempts_reached call_sid=%s attempts=%s', call_sid, attempts)
        append_interaction_transcript(call_sid, 'agent', 'Could not confidently navigate the IVR menu; proceeding to live Q&A mode as a best effort.', source='ivr_navigation_fallback')
        update_call_state(call_sid, {
            'status': 'customer_care_listening',
            'next_action': 'customer_care_qna',
            'ivr_attempts': attempts,
            'ivr_phase': 'customer_care_intro',
            'voice': voice,
        })
        intro = build_customer_care_intro(state.get('contact_context') or {})
        append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
        twiml_response = build_customer_care_twiml(intro, voice, followup=False)
        return Response(twiml_response, mimetype='text/xml')

    update_call_state(call_sid, {'status': 'ivr_navigating', 'ivr_attempts': attempts, 'ivr_phase': 'ivr_navigating'})
    return Response(build_ivr_navigation_twiml(), mimetype='text/xml')


@app.route('/api/webhook/keypress', methods=['POST'])
def handle_keypress():
    """Twilio hits this endpoint when the user presses a key on the initial connection."""
    call_sid = request.form.get('CallSid')
    digits = request.form.get('Digits')
    logger.info('api_keypress_webhook call_sid=%s digits=%s', call_sid, digits)
    
    if digits == '9':
        voice = DEFAULT_VOICE
        state = read_call_state(call_sid)
        state_voice = (state.get('voice') or '').strip() if isinstance(state, dict) else ''
        if state_voice:
            voice = state_voice

        update_call_state(call_sid, {
            'status': 'customer_care_listening',
            'last_speech': '',
            'detected_intent': '',
            'next_action': 'customer_care_qna',
            'menu_digit': '9',
            'ivr_phase': 'customer_care_intro',
            'voice': voice,
        })
        intro = build_customer_care_intro(state.get('contact_context') or {})
        append_interaction_transcript(call_sid, 'agent', intro, source='customer_care_intro')
        twiml_response = build_customer_care_twiml(intro, voice, followup=False)
        return Response(twiml_response, mimetype='text/xml')
    else:
        callback_url = get_public_base_url()
        # If they pressed the wrong key, prompt them again
        twiml_response = f'''
        <Response>
            <Gather numDigits="1" action="{callback_url}/api/webhook/keypress" method="POST" timeout="60">
                <Say voice="{DEFAULT_VOICE}">Invalid key. Please press 9 to begin.</Say>
            </Gather>
        </Response>
        '''
        return Response(twiml_response, mimetype='text/xml')


@app.route('/api/ask_and_listen', methods=['POST'])
def ask_and_listen():
    data = request.json or {}
    call_sid = data.get('call_sid')
    twilio_client, _ = get_twilio_context_for_call(call_sid)
    if not twilio_client:
        return jsonify({'error': 'Twilio credentials unavailable for this call.', 'code': 'TWILIO_CONTEXT_MISSING'}), 500
    question = data.get('question', '')
    yes_response = data.get('yes_response', '')
    no_response = data.get('no_response', '')
    menu_options = data.get('menu_options') or {}
    callback_url = get_public_base_url()
    voice = (data.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE
    logger.info('api_ask_and_listen_voice call_sid=%s voice=%s', call_sid, voice)

    if not call_sid:
        logger.warning('api_ask_and_listen_missing_call_sid')
        return jsonify({'error': 'call_sid is required.'}), 400

    try:
        live_call = telephony_provider.fetch_call(twilio_client, call_sid)
        twilio_status = (live_call.status or '').lower()
        update_call_state(call_sid, {'twilio_call_status': twilio_status})
        logger.info('api_ask_and_listen_status call_sid=%s twilio_status=%s', call_sid, twilio_status)
    except Exception as e:
        logger.exception('api_ask_and_listen_fetch_error call_sid=%s error=%s', call_sid, str(e))
        return jsonify({'error': f'Unable to fetch call status: {str(e)}'}), 500

    if twilio_status not in {'queued', 'ringing', 'in-progress'}:
        update_call_state(call_sid, {'twilio_call_status': twilio_status, 'status': 'call_ended'})
        logger.warning('api_ask_and_listen_inactive call_sid=%s twilio_status=%s', call_sid, twilio_status)
        return jsonify({
            'error': 'Call is not active anymore.',
            'code': 'CALL_NOT_ACTIVE',
            'twilio_status': twilio_status,
        }), 409

    # SAFETY FIX: Escape text to prevent special characters (like '&' or '<') from breaking Twilio's XML
    safe_question = safe_text(question)
    append_interaction_transcript(call_sid, 'agent', question, source='ask_and_listen')

    # Store the expected answers in our state
    update_call_state(call_sid, {
        'status': 'listening',
        'yes_response': yes_response,
        'no_response': no_response,
        'menu_options': menu_options,
        'voice': voice,
        'last_speech': '',
        'detected_intent': '',
        'next_action': '',
        'menu_digit': '',
        'twilio_call_status': twilio_status,
    })

    try:
        # Tell Twilio to speak the question, then listen for a reply
        # and POST the transcript to our public URL webhook
        unclear_prompt = safe_text(choose_non_repeating_fallback_reply(call_sid, '', {}))
        twiml_instructions = f'''
        <Response>
            <Gather input="speech dtmf" numDigits="1" action="{callback_url}/api/webhook/speech" method="POST" timeout="5" speechTimeout="3">
                <Say voice="{voice}">{safe_question}</Say>
            </Gather>
            <Say voice="{voice}">{unclear_prompt}</Say>
            <Pause length="3600"/>
        </Response>
        '''
        telephony_provider.update_call_twiml(twilio_client, call_sid, twiml_instructions)
        logger.info('api_ask_and_listen_success call_sid=%s', call_sid)
        return jsonify({'message': 'Listening for response...'}), 200
    except Exception as e:
        message = str(e)
        lowered = message.lower()
        if 'not in-progress' in lowered or 'cannot redirect' in lowered:
            update_call_state(call_sid, {'status': 'call_ended'})
            logger.warning('api_ask_and_listen_redirect_inactive call_sid=%s twilio_status=%s', call_sid, twilio_status)
            return jsonify({
                'error': 'Call is not active anymore.',
                'code': 'CALL_NOT_ACTIVE',
                'twilio_status': twilio_status or 'unknown',
            }), 409
        logger.exception('api_ask_and_listen_error call_sid=%s error=%s', call_sid, message)
        return jsonify({'error': message}), 500


@app.route('/api/webhook/speech', methods=['POST'])
def handle_speech():
    """Twilio hits this endpoint when it hears the user speak."""
    call_sid = request.form.get('CallSid')
    speech_result = request.form.get('SpeechResult', '').lower()
    pressed_digit = (request.form.get('Digits') or '').strip()
    logger.info('api_speech_webhook call_sid=%s speech=%s', call_sid, speech_result)
    
    state = read_call_state(call_sid)
    yes_text = state.get('yes_response', 'Okay.')
    no_text = state.get('no_response', 'Alright.')
    menu_options = state.get('menu_options') or {}
    voice = (state.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE
    logger.info('api_speech_voice call_sid=%s voice=%s', call_sid, voice)

    if (state.get('status') or '').strip().lower() == 'customer_care_listening':
        asked_text = (request.form.get('SpeechResult') or '').strip()
        if not asked_text:
            asked_text = ''

        resolved_context = resolve_contact_context(state.get('contact_context') or {})
        aetna_action = resolve_universal_prompt_action(asked_text, resolved_context, state)
        if aetna_action:
            source = aetna_action.get('source') or 'universal_prompt_action'
            if aetna_action.get('type') == 'listen':
                logger.info('aetna_prompt_action call_sid=%s source=%s type=listen prompt=%s', call_sid, source, asked_text)
                update_call_state(call_sid, {
                    'status': 'customer_care_listening',
                    'last_speech': asked_text.lower(),
                    'detected_intent': 'listen',
                    'next_action': 'wait',
                    'last_question_from_customer_care': asked_text,
                    'last_answer_source': source,
                    'last_answer_text': '',
                })
                append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
                return Response(build_customer_care_twiml('', voice, followup=False), mimetype='text/xml')
            if aetna_action.get('type') == 'dtmf' and aetna_action.get('digits'):
                digits = str(aetna_action.get('digits') or '').strip()
                asked_norm = _normalize_text(asked_text)
                source_lower = str(source).lower()
                
                is_different_patient_press = (
                    # Standard "same patient or different patient" gate (press 2)
                    ('same patient' in asked_norm and 'different patient' in asked_norm and digits == '2') or
                    ('different patient or press 2' in asked_norm and digits == '2') or
                    ('different patient or press two' in asked_norm and digits == '2') or
                    (str(source) == 'coverage_for_whom_different_patient' and digits == '2') or
                    # Campaign rollover rules: e.g. Meritain "try another patient or press 1"
                    # These are campaign_rollover type rules that advance the queue
                    ('rollover' in source_lower and 'try_another_patient' in source_lower)
                )
                
                if (
                    is_different_patient_press
                    and str(state.get('selected_profile_name') or '') == 'aetna_representative'
                ):
                    queue = normalize_campaign_contact_queue(state.get('campaign_contact_queue'))
                    has_next_member = bool(queue)
                    if not has_next_member:
                        has_next_member = bool(state.get('campaign_has_next_member'))
                    if has_next_member:
                        if queue:
                            # ── Write audit for current patient BEFORE switching context ──
                            current_member_id = normalize_member_id(
                                state.get('active_member_id') or
                                extract_member_id_from_context(state.get('contact_context') or {})
                            )
                            if current_member_id:
                                try:
                                    current_state = read_call_state(call_sid)
                                    current_transcript = current_state.get('interaction_transcript') or []
                                    current_audits = extract_claim_details_via_python_script(current_transcript)
                                    if current_audits and current_audits.get('audits'):
                                        # Only keep audit for the current patient
                                        patient_audits = [
                                            a for a in current_audits['audits']
                                            if normalize_member_id(a.get('member_id')) == current_member_id
                                        ]
                                        if patient_audits:
                                            save_audit_results_json(patient_audits, call_sid=call_sid)
                                            logger.info('audit_written_before_context_switch call_sid=%s member_id=%s', call_sid, current_member_id)
                                except Exception as audit_exc:
                                    logger.warning('audit_write_before_switch_failed call_sid=%s error=%s', call_sid, str(audit_exc))

                            next_contact = dict(queue[0])
                            remaining_queue = queue[1:]
                            next_member_id = normalize_member_id(next_contact.get('member_id') or extract_member_id_from_context(next_contact))
                            merged_next_context = resolve_contact_context(next_contact)
                            all_bank = normalize_uploaded_answer_bank(state.get('uploaded_answer_bank_all') or state.get('uploaded_answer_bank') or [])
                            next_bank = filter_uploaded_answer_bank_for_member(all_bank, next_member_id)
                            update_call_state(call_sid, {
                                'contact_context': merged_next_context,
                                'active_member_id': next_member_id,
                                'campaign_member_switched': True,
                                'uploaded_answer_bank': next_bank,
                                'uploaded_answer_bank_all': all_bank,
                                'campaign_contact_queue': remaining_queue,
                                'campaign_remaining_contacts_after_current': len(remaining_queue),
                                'campaign_has_next_member': bool(remaining_queue),
                            })
                            append_interaction_transcript(call_sid, 'agent', f'Switched context to next member {next_member_id} on same call.', source='campaign_member_switch')
                    else:
                        goodbye_text = 'Goodbye. Thank you for your help today.'
                        logger.info('aetna_prompt_action_end_no_next_member call_sid=%s prompt=%s', call_sid, asked_text)
                        update_call_state(call_sid, {
                            'status': 'call_ended',
                            'ivr_phase': 'customer_care_completed',
                            'last_speech': asked_text.lower(),
                            'detected_intent': 'customer_care_qna',
                            'next_action': 'end_call',
                            'last_question_from_customer_care': asked_text,
                            'last_answer_source': 'aetna_campaign_end',
                            'last_answer_text': goodbye_text,
                        })
                        append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
                        append_interaction_transcript(call_sid, 'agent', goodbye_text, source='customer_care_live_answer_aetna_campaign_end')
                        return Response(build_end_call_twiml(goodbye_text, voice), mimetype='text/xml')
                logger.info('aetna_prompt_action call_sid=%s source=%s type=dtmf digits=%s prompt=%s', call_sid, source, digits, asked_text)
                update_call_state(call_sid, {
                    'status': 'customer_care_listening',
                    'last_speech': asked_text.lower(),
                    'detected_intent': 'menu',
                    'next_action': 'aetna_auto_dtmf',
                    'menu_digit': digits,
                    'last_question_from_customer_care': asked_text,
                    'last_answer_source': source,
                    'last_answer_text': f'Pressed {digits}.',
                })
                append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
                append_interaction_transcript(call_sid, 'agent', f'Pressed {digits} for Aetna prompt.', source='customer_care_live_auto_dtmf')
                return Response(build_customer_care_twiml('', voice, followup=False, play_digits=digits), mimetype='text/xml')
            if aetna_action.get('type') == 'speak' and aetna_action.get('text'):
                text = str(aetna_action.get('text') or '').strip()
                logger.info('aetna_prompt_action call_sid=%s source=%s type=speak text=%s prompt=%s', call_sid, source, text, asked_text)
                if aetna_action.get('end_call'):
                    update_call_state(call_sid, {
                        'status': 'call_ended',
                        'ivr_phase': 'customer_care_completed',
                        'last_speech': asked_text.lower(),
                        'detected_intent': 'customer_care_qna',
                        'next_action': 'end_call',
                        'last_question_from_customer_care': asked_text,
                        'last_answer_source': source,
                        'last_answer_text': text,
                    })
                    append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
                    append_interaction_transcript(call_sid, 'agent', text, source=f'customer_care_live_answer_{source}')
                    return Response(build_end_call_twiml(text, voice), mimetype='text/xml')
                update_call_state(call_sid, {
                    'status': 'customer_care_listening',
                    'last_speech': asked_text.lower(),
                    'detected_intent': 'customer_care_qna',
                    'next_action': 'customer_care_qna',
                    'last_question_from_customer_care': asked_text,
                    'last_answer_source': source,
                    'last_answer_text': text,
                })
                append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
                append_interaction_transcript(call_sid, 'agent', text, source=f'customer_care_live_answer_{source}')
                return Response(build_customer_care_twiml(text, voice, followup=False), mimetype='text/xml')

        if _to_bool(state.get('awaiting_rejection_reason'), False) and looks_like_rejection_reason_detail(asked_text):
            thanks_reply = 'Thank you for sharing the rejection reason. We appreciate your help. Goodbye.'
            update_call_state(call_sid, {
                'status': 'call_ended',
                'ivr_phase': 'customer_care_completed',
                'last_speech': asked_text.lower(),
                'detected_intent': 'customer_care_qna',
                'next_action': 'end_call',
                'awaiting_rejection_reason': False,
                'last_question_from_customer_care': asked_text,
                'last_answer_source': 'rejection_reason_closure',
                'last_answer_text': thanks_reply,
            })
            append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
            append_interaction_transcript(call_sid, 'agent', thanks_reply, source='customer_care_live_answer_rejection_reason_closure')
            return Response(build_end_call_twiml(thanks_reply, voice), mimetype='text/xml')

        asked_norm = _normalize_text(asked_text)
        if asked_norm.startswith('the claim was completed on') or asked_norm.startswith('claim was completed on'):
            # Strictly silence and listen if the representative is reading completed claim details
            active_member_id = normalize_member_id(state.get('active_member_id') or extract_member_id_from_context(resolved_context))
            update_call_state(call_sid, {
                'status': 'customer_care_listening',
                'last_speech': asked_text.lower(),
                'detected_intent': 'customer_care_qna',
                'next_action': 'customer_care_qna',
                'last_question_from_customer_care': asked_text,
                'last_answer_source': 'statement_silent_listen',
                'last_answer_text': '',
            })
            persist_claim_status_from_conversation(call_sid, state, asked_text, '', active_member_id)
            maybe_update_claim_status(call_sid, state, asked_text, '', active_member_id)
            append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
            append_interaction_transcript(call_sid, 'agent', '', source='customer_care_live_answer_statement_silent_listen')
            return Response(build_customer_care_twiml('', voice, followup=False), mimetype='text/xml')

        active_member_id = normalize_member_id(state.get('active_member_id') or extract_member_id_from_context(resolved_context))
        confirmation_digit = resolve_confirmation_digit(asked_text, resolved_context, last_answer_text=state.get('last_answer_text') or '', state=state)
        if confirmation_digit:
            if state.get('campaign_member_switched') and confirmation_digit == '2':
                update_call_state(call_sid, {'campaign_member_switched': False})
            answer_text = f'Okay. Confirming now with {confirmation_digit}.'
            answer_source = 'auto_dtmf_confirmation'
            update_call_state(call_sid, {
                'status': 'customer_care_listening',
                'last_speech': asked_text.lower(),
                'detected_intent': 'menu',
                'next_action': 'customer_care_auto_confirm',
                'menu_digit': confirmation_digit,
                'active_member_id': active_member_id,
                'contact_context': resolved_context,
                'last_question_from_customer_care': asked_text,
                'last_answer_source': answer_source,
                'last_answer_text': answer_text,
            })
            append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
            append_interaction_transcript(call_sid, 'agent', f'Pressed {confirmation_digit} for confirmation prompt.', source='customer_care_live_auto_dtmf')
            return Response(build_customer_care_twiml('', voice, followup=False, play_digits=confirmation_digit), mimetype='text/xml')

        answer_text, answer_source = resolve_customer_care_answer(
            asked_text,
            resolved_context,
            state.get('uploaded_answer_bank') or [],
            active_member_id=active_member_id,
            call_sid=call_sid,
        )
        if answer_source == 'fallback_human' and not is_active_question_or_command(asked_text):
            answer_text = ''
            answer_source = 'statement_silent_listen'

        answer_consistency = build_answer_consistency_check(
            asked_text,
            answer_text,
            answer_source,
            resolved_context,
            active_member_id=active_member_id,
            uploaded_answer_bank=state.get('uploaded_answer_bank') or [],
        )
        if answer_consistency.get('status') != 'ok':
            logger.warning('answer_consistency_warning call_sid=%s details=%s', call_sid, json.dumps(answer_consistency, ensure_ascii=True))

        status_followup = build_claim_status_followup(asked_text, answer_text, resolved_context, state=state)
        if status_followup.get('claim_status'):
            answer_text = status_followup.get('reply_text') or answer_text
            answer_source = 'claim_status_followup'
            answer_consistency = {
                'status': 'ok',
                'flags': [],
                'answer_source': answer_source,
                'active_member_id': active_member_id,
                'context_member_id': normalize_member_id(extract_member_id_from_context(resolved_context)),
            }

        update_call_state(call_sid, {
            'status': 'customer_care_listening',
            'last_speech': asked_text.lower(),
            'detected_intent': 'customer_care_qna',
            'next_action': 'customer_care_qna',
            'ivr_phase': 'customer_care_qna',
            'active_member_id': active_member_id,
            'contact_context': resolved_context,
            'last_question_from_customer_care': asked_text,
            'last_answer_source': answer_source,
            'last_answer_text': answer_text,
            'answer_consistency': answer_consistency,
            'awaiting_rejection_reason': bool(status_followup.get('await_rejection_reason')),
        })
        persist_claim_status_from_conversation(call_sid, state, asked_text, answer_text, active_member_id)
        maybe_update_claim_status(call_sid, state, asked_text, answer_text, active_member_id)
        append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_live_question')
        append_interaction_transcript(call_sid, 'agent', answer_text, source=f'customer_care_live_answer_{answer_source}')
        if status_followup.get('end_call'):
            switched, twiml_inst = transition_call_to_next_member_context(call_sid, state, voice, answer_text)
            if switched:
                return Response(twiml_inst, mimetype='text/xml')
            update_call_state(call_sid, {'status': 'call_ended', 'ivr_phase': 'customer_care_completed'})
            return Response(twiml_inst, mimetype='text/xml')
        return Response(build_customer_care_twiml(answer_text, voice, followup=False), mimetype='text/xml')

    # Logic to match keywords in what the user said
    yes_keywords = ['yes', 'yeah', 'sure', 'yep', 'ok', 'okay', 'correct', 'right', 'please', 'course', 'absolutely']
    no_keywords = ['no', 'nope', 'not', 'wrong', 'incorrect', 'stop']

    intent = 'unknown'
    reply_text = "I'm sorry, I didn't understand if that was a yes or a no. Please hold."
    next_action = ''

    if pressed_digit and isinstance(menu_options, dict):
        option_value = menu_options.get(pressed_digit)
        if isinstance(option_value, dict):
            next_action = str(option_value.get('action') or '').strip().lower()
            reply_text = str(option_value.get('reply') or option_value.get('label') or f'You selected option {pressed_digit}. Please hold.')
        elif isinstance(option_value, str):
            next_action = option_value.strip().lower()
            reply_text = f'You selected option {pressed_digit}. Please hold.'
        if next_action:
            intent = 'menu'

    if intent != 'menu':
        if any(word in speech_result for word in yes_keywords):
            intent = 'yes'
            reply_text = yes_text
        elif any(word in speech_result for word in no_keywords):
            intent = 'no'
            reply_text = no_text

    # Update our server state so the frontend can display it
    update_call_state(call_sid, {
        'status': 'answered',
        'last_speech': speech_result,
        'detected_intent': intent,
        'next_action': next_action,
        'menu_digit': pressed_digit,
    })

    # SAFETY FIX: Escape the reply text as well
    safe_reply = safe_text(reply_text)
    append_interaction_transcript(call_sid, 'customer', speech_result, source='gather_speech')
    append_interaction_transcript(call_sid, 'agent', reply_text, source='speech_reply')

    # Tell Twilio what to say back to the user automatically
    twiml_response = f'''
    <Response>
        <Say voice="{voice}">{safe_reply}</Say>
        <Pause length="3600"/>
    </Response>
    '''
    return Response(twiml_response, mimetype='text/xml')


@app.route('/api/call/<call_sid>/end', methods=['POST'])
def end_active_call(call_sid):
    logger.info('api_call_end_request call_sid=%s', call_sid)
    try:
        twilio_client, _ = get_twilio_context_for_call(call_sid)
        if twilio_client:
            # Update Twilio call status to completed to terminate the active call
            twilio_client.calls(call_sid).update(status='completed')
            logger.info('api_call_ended_successfully call_sid=%s', call_sid)
            return jsonify({'message': 'Call ended successfully', 'call_sid': call_sid}), 200
        else:
            logger.warning('api_call_end_failed_client_missing call_sid=%s', call_sid)
            return jsonify({'error': 'Twilio client missing or credentials not configured', 'code': 'TWILIO_CLIENT_MISSING'}), 500
    except Exception as exc:
        logger.exception('api_call_end_exception call_sid=%s error=%s', call_sid, str(exc))
        return jsonify({'error': f'Failed to end call: {str(exc)}', 'code': 'TWILIO_CALL_END_FAILED'}), 500


@app.route('/api/call_state/<call_sid>', methods=['GET'])
def get_call_state(call_sid):
    """Frontend polls this to see if the user answered."""
    state = read_call_state(call_sid)
    if state.get('interaction_transcript') and (state.get('storage_options') or {}).get('save_enabled'):
        try:
            artifact_dir = get_artifact_call_dir(call_sid, state.get('storage_options'))
            interaction_path = os.path.join(artifact_dir, 'interaction-transcript.json')
            save_json_file(interaction_path, normalize_interaction_transcript(state.get('interaction_transcript')))
            artifacts = state.get('artifacts', {})
            artifacts['interaction_saved_path'] = interaction_path
            state['artifacts'] = artifacts
            write_call_state(call_sid, state)
        except Exception:
            pass
    response_state = strip_internal_state_fields(state)
    logger.debug('api_call_state call_sid=%s status=%s twilio_status=%s', call_sid, response_state.get('status'), response_state.get('twilio_call_status'))
    return jsonify(response_state), 200


@app.route('/api/customer_care_qna', methods=['POST'])
def customer_care_qna():
    data = request.json or {}
    call_sid = (data.get('call_sid') or '').strip()
    asked_text = (data.get('asked_text') or '').strip()
    contact_context = resolve_contact_context(data.get('contact_context') or {})
    voice = (data.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE

    if not call_sid:
        return jsonify({'error': 'call_sid is required.'}), 400
    if not asked_text:
        return jsonify({'error': 'asked_text is required.'}), 400

    twilio_client, _ = get_twilio_context_for_call(call_sid)
    if not twilio_client:
        return jsonify({'error': 'Twilio credentials unavailable for this call.', 'code': 'TWILIO_CONTEXT_MISSING'}), 500

    state = read_call_state(call_sid)
    uploaded_answer_bank = state.get('uploaded_answer_bank') or []
    active_member_id = normalize_member_id(state.get('active_member_id') or extract_member_id_from_context(contact_context))
    status_followup = {'claim_status': '', 'reply_text': '', 'end_call': False}

    confirmation_digit = resolve_confirmation_digit(asked_text, contact_context, state=state)
    if confirmation_digit:
        if state.get('campaign_member_switched') and confirmation_digit == '2':
            update_call_state(call_sid, {'campaign_member_switched': False})
        answer_text = f'Okay. Confirming now with {confirmation_digit}.'
        answer_source = 'auto_dtmf_confirmation'
        twiml = build_customer_care_twiml('', voice, followup=False, play_digits=confirmation_digit)
    else:
        answer_text, answer_source = resolve_customer_care_answer(
            asked_text,
            contact_context,
            uploaded_answer_bank,
            active_member_id=active_member_id,
            call_sid=call_sid,
        )
        if answer_source == 'fallback_human' and not is_active_question_or_command(asked_text):
            answer_text = ''
            answer_source = 'statement_silent_listen'

        answer_consistency = build_answer_consistency_check(
            asked_text,
            answer_text,
            answer_source,
            contact_context,
            active_member_id=active_member_id,
            uploaded_answer_bank=uploaded_answer_bank,
        )
        if answer_consistency.get('status') != 'ok':
            logger.warning('answer_consistency_warning call_sid=%s details=%s', call_sid, json.dumps(answer_consistency, ensure_ascii=True))
        status_followup = build_claim_status_followup(asked_text, answer_text, contact_context, state=state)
        if status_followup.get('claim_status'):
            answer_text = status_followup.get('reply_text') or answer_text
            answer_source = 'claim_status_followup'
            answer_consistency = {
                'status': 'ok',
                'flags': [],
                'answer_source': answer_source,
                'active_member_id': active_member_id,
                'context_member_id': normalize_member_id(extract_member_id_from_context(contact_context)),
            }
        switched = False
        if status_followup.get('end_call'):
            switched, twiml = transition_call_to_next_member_context(call_sid, state, voice, answer_text)
        else:
            twiml = build_customer_care_twiml(answer_text, voice, followup=False)
    if confirmation_digit:
        switched = False
        answer_consistency = {
            'status': 'ok',
            'flags': [],
            'answer_source': answer_source,
            'active_member_id': active_member_id,
            'context_member_id': normalize_member_id(extract_member_id_from_context(contact_context)),
        }

    try:
        telephony_provider.update_call_twiml(twilio_client, call_sid, twiml)
    except Exception as exc:
        logger.exception('api_customer_care_qna_error call_sid=%s error=%s', call_sid, str(exc))
        return jsonify({'error': str(exc)}), 500

    if not switched:
        update_call_state(call_sid, {
            'status': 'call_ended' if status_followup.get('end_call') else 'customer_care_listening',
            'ivr_phase': 'customer_care_completed' if status_followup.get('end_call') else 'customer_care_qna',
            'active_member_id': active_member_id,
            'contact_context': contact_context,
            'last_question_from_customer_care': asked_text,
            'last_answer_source': answer_source,
            'last_answer_text': answer_text,
            'answer_consistency': answer_consistency,
            'menu_digit': confirmation_digit,
            'awaiting_rejection_reason': bool(status_followup.get('await_rejection_reason')),
        })
    else:
        update_call_state(call_sid, {
            'last_question_from_customer_care': asked_text,
            'last_answer_source': answer_source,
            'last_answer_text': answer_text,
            'answer_consistency': answer_consistency,
            'menu_digit': confirmation_digit,
        })
    persist_claim_status_from_conversation(call_sid, state, asked_text, answer_text, active_member_id)
    maybe_update_claim_status(call_sid, state, asked_text, answer_text, active_member_id)
    append_interaction_transcript(call_sid, 'customer', asked_text, source='customer_care_qna_question')
    if confirmation_digit:
        append_interaction_transcript(call_sid, 'agent', f'Pressed {confirmation_digit} for confirmation prompt.', source='customer_care_qna_auto_dtmf')
    else:
        append_interaction_transcript(call_sid, 'agent', answer_text, source=f'customer_care_qna_{answer_source}')

    return jsonify({
        'message': 'Answered question and resumed listening.',
        'answer_source': answer_source,
        'answer_text': answer_text,
    }), 200


@app.route('/api/customer_care/reload_qa', methods=['POST'])
def reload_customer_care_qa():
    return jsonify({
        'message': 'Static QA reload is not used. Answering uses uploaded JSON from UI per call.',
        'qa_file': CUSTOMER_CARE_QA_FILE,
        'count': 0,
    }), 200


@app.route('/api/customer_care/reload_customer_data', methods=['POST'])
def reload_customer_data():
    global CUSTOMER_DATA_INDEX
    CUSTOMER_DATA_INDEX = load_customer_data_index()
    return jsonify({
        'message': 'Customer data reloaded.',
        'customer_data_file': CUSTOMER_DATA_FILE,
        'members': len(CUSTOMER_DATA_INDEX),
    }), 200


@app.route('/api/ivr/reload_profile', methods=['POST'])
def reload_ivr_profile():
    global IVR_PROFILE
    IVR_PROFILE = load_ivr_profile()
    return jsonify({
        'message': 'IVR profile reloaded.',
        'ivr_profile_file': IVR_PROFILE_FILE,
        'profile_name': IVR_PROFILE.get('profile_name', ''),
        'language': (IVR_PROFILE.get('language') or {}).get('preferred', ''),
        'menu_goals': IVR_PROFILE.get('menu_goals') or [],
        'profiles': sorted(list((IVR_PROFILE.get('profiles') or {}).keys())) if isinstance(IVR_PROFILE.get('profiles'), dict) else [IVR_PROFILE.get('profile_name', '')],
        'default_profile': IVR_PROFILE.get('default_profile', IVR_PROFILE.get('profile_name', '')),
        'number_map_count': len(IVR_PROFILE.get('number_map') or {}) if isinstance(IVR_PROFILE.get('number_map'), dict) else 0,
    }), 200


@app.route('/api/version', methods=['GET'])
def get_version():
    return jsonify({'version': APP_VERSION, 'backend_version': APP_VERSION, 'pod_name': POD_NAME}), 200


# Keep standard text-to-speech for manual overriding
@app.route('/api/speak', methods=['POST'])
def speak_in_call():
    data = request.json
    call_sid = data.get('call_sid')
    twilio_client, _ = get_twilio_context_for_call(call_sid)
    if not twilio_client:
        return jsonify({'error': 'Twilio credentials unavailable for this call.', 'code': 'TWILIO_CONTEXT_MISSING'}), 500
    text = data.get('text', '')
    voice = (data.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE
    safe_speech = safe_text(text)
    append_interaction_transcript(call_sid, 'agent', text, source='manual_speak')
    logger.info('api_speak_request call_sid=%s', call_sid)
    logger.info('api_speak_voice call_sid=%s voice=%s', call_sid, voice)
    if not call_sid:
        return jsonify({'error': 'call_sid is required.'}), 400

    twilio_status = 'unknown'
    try:
        live_call = telephony_provider.fetch_call(twilio_client, call_sid)
        twilio_status = (live_call.status or '').lower()
        update_call_state(call_sid, {'twilio_call_status': twilio_status})
        logger.info('api_speak_status call_sid=%s twilio_status=%s', call_sid, twilio_status)
    except Exception as e:
        logger.exception('api_speak_fetch_error call_sid=%s error=%s', call_sid, str(e))
        return jsonify({'error': f'Unable to fetch call status: {str(e)}'}), 500

    if twilio_status not in {'queued', 'ringing', 'in-progress'}:
        update_call_state(call_sid, {'status': 'call_ended'})
        logger.warning('api_speak_inactive call_sid=%s twilio_status=%s', call_sid, twilio_status)
        return jsonify({
            'error': 'Call is not active anymore.',
            'code': 'CALL_NOT_ACTIVE',
            'twilio_status': twilio_status,
        }), 409

    try:
        twiml = f'<Response><Say voice="{voice}">{safe_speech}</Say><Pause length="3600"/></Response>'
        telephony_provider.update_call_twiml(twilio_client, call_sid, twiml)
        logger.info('api_speak_success call_sid=%s', call_sid)
        return jsonify({'message': 'Text spoken successfully'}), 200
    except Exception as e:
        message = str(e)
        lowered = message.lower()
        if 'not in-progress' in lowered or 'cannot redirect' in lowered:
            update_call_state(call_sid, {'status': 'call_ended'})
            logger.warning('api_speak_redirect_inactive call_sid=%s twilio_status=%s', call_sid, twilio_status)
            return jsonify({
                'error': 'Call is not active anymore.',
                'code': 'CALL_NOT_ACTIVE',
                'twilio_status': twilio_status,
            }), 409
        logger.exception('api_speak_error call_sid=%s error=%s', call_sid, message)
        return jsonify({'error': message}), 500


@app.route('/api/dtmf', methods=['POST'])
def send_dtmf_in_call():
    data = request.json or {}
    call_sid = (data.get('call_sid') or '').strip()
    digits = (data.get('digits') or '').strip()
    voice = (data.get('voice') or DEFAULT_VOICE).strip() or DEFAULT_VOICE
    menu_prompt = (data.get('menu_prompt') or '').strip()

    if not call_sid:
        return jsonify({'error': 'call_sid is required.'}), 400
    if not digits or not re.fullmatch(r'[0-9#*]{1,8}', digits):
        return jsonify({'error': 'digits must be 1-8 chars of 0-9, # or *.'}), 400

    twilio_client, _ = get_twilio_context_for_call(call_sid)
    if not twilio_client:
        return jsonify({'error': 'Twilio credentials unavailable for this call.', 'code': 'TWILIO_CONTEXT_MISSING'}), 500

    callback_url = get_public_base_url()
    safe_prompt = safe_text(menu_prompt or 'Please choose an option.')
    try:
        twiml = f'''
        <Response>
            <Gather input="dtmf" numDigits="1" action="{callback_url}/api/webhook/speech" method="POST" timeout="8">
                <Say voice="{voice}">{safe_prompt}</Say>
            </Gather>
            <Say voice="{voice}">No selection received. Please hold.</Say>
            <Pause length="3600"/>
        </Response>
        '''
        telephony_provider.update_call_twiml(twilio_client, call_sid, twiml)
        return jsonify({'message': 'DTMF menu sent successfully'}), 200
    except Exception as exc:
        logger.exception('api_dtmf_error call_sid=%s error=%s', call_sid, str(exc))
        return jsonify({'error': str(exc)}), 500

if __name__ == '__main__':
    print(f"Starting Twilio Call Server on http://0.0.0.0:{APP_PORT} ({APP_VERSION})")
    app.run(host='0.0.0.0', debug=False, port=APP_PORT)
